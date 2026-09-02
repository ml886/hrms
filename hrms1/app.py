"""HRMS 人力资源管理系统 - Flask 主应用"""
import os
import io
import json
import re
import csv as _csv
import time as _time
from datetime import date, datetime, timedelta
from sqlalchemy import or_, func, create_engine, inspect, text
import pandas as pd
from openpyxl import Workbook, load_workbook
from flask import (
    Flask, render_template, request, redirect, url_for, jsonify,
    flash, session, send_file, abort, current_app, Response, g
)
from flask_login import LoginManager, login_user, logout_user, login_required, current_user

from models import (db, current_company_engine, User, Employee, Contract, ContractRenew, Attachment,
                    InsuranceDetail, SalaryRecord, Todo, DictItem, TransferRecord, OperationLog, Setting,
                    PayslipTemplate, Company, UserCompany, SysSetting)
import models as _models  # 用于向多租户会话注入占位引擎兜底
from auth import (role_required, can_view_employee, can_view_module,
    can_manage_module, employee_scope_filter, parse_depts, _deny, MODULES)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 系统版本号（登录页与主页面页脚展示）
APP_VERSION = '2.1.2'

# 公司（独立库）模型对应的表对象集合，用于在公司库文件中单独建表
COMPANY_TABLES = [m.__table__ for m in
                  (Employee, Contract, ContractRenew, Attachment, InsuranceDetail,
                   SalaryRecord, TransferRecord, PayslipTemplate, Setting, OperationLog, DictItem)]

# 工资表汇总：最近一次导入的源文件缓存（校验接口写入，导出接口按 token 复用其样式）
# 结构：{token: (filename, bytes)}；仅保留最近若干份，避免内存堆积
PAYSLIP_SRC_STORE = {}


def due_todo_query(user_id):
    """顶部铃铛角标与弹窗共用：当前公司下所有「过期未办」待办（未完成且截止日期<=今天）。
    角标数字与弹窗内容来自同一查询，保证一致。"""
    today = date.today()
    cid = current_company().id if current_company() else None
    q = Todo.query.filter_by(user_id=user_id, done=False)
    if cid is not None:
        q = q.filter(Todo.company_id == cid)
    return q.filter(Todo.due_date <= today)


def _safe_json_load(s, default):
    """安全解析 JSON 文本，失败返回默认值。"""
    if not s:
        return default
    try:
        v = json.loads(s)
        return v if v is not None else default
    except Exception:
        return default


def client_unit_options():
    """客户单位下拉选项：以代码字段（字典）「客户单位」分类为主（按 sort 排序），
    并合并工资明细中已存在但未登记到代码字段的单位，保证历史数据可筛选。"""
    dict_units = [d.label for d in DictItem.query.filter_by(category='客户单位', enabled=True)
                  .order_by(DictItem.sort.asc(), DictItem.id.asc()).all()]
    existing = [u[0] for u in db.session.query(SalaryRecord.client_unit)
                .filter(SalaryRecord.client_unit != '', SalaryRecord.client_unit.isnot(None))
                .distinct().all()]
    seen = set(dict_units)
    opts = list(dict_units)
    for u in existing:
        if u not in seen:
            seen.add(u)
            opts.append(u)
    return opts


# ---------------- 多公司（账套）引擎与助手 ----------------
_CDB_ENGINES = {}   # company_id -> Engine 缓存


def company_engine(cid):
    """获取指定公司的数据库引擎（按需建库建表并缓存）。"""
    if cid in _CDB_ENGINES:
        return _CDB_ENGINES[cid]
    comp = Company.query.get(cid)
    if not comp:
        return None
    eng = create_engine('sqlite:///' + comp.db_path())
    try:
        db.metadata.create_all(eng, tables=COMPANY_TABLES)
    except Exception:
        pass
    _CDB_ENGINES[cid] = eng
    return eng


def current_company():
    cid = session.get('company_id')
    if not cid:
        return None
    return Company.query.get(cid)


def accessible_companies(user):
    """当前用户可管理的公司列表：admin=全部；hr/employee=权限关联的公司。"""
    if user.role == 'admin':
        return Company.query.filter_by(enabled=True).order_by(Company.id.asc()).all()
    links = UserCompany.query.filter_by(user_id=user.id).all()
    ids = [l.company_id for l in links]
    if not ids:
        return []
    return Company.query.filter(Company.id.in_(ids), Company.enabled == True).order_by(Company.id.asc()).all()


def user_can_access_company(user, cid):
    if user.role == 'admin':
        return True
    return UserCompany.query.filter_by(user_id=user.id, company_id=cid).first() is not None


def create_app():
    app = Flask(__name__)
    app.config['SECRET_KEY'] = 'hrms-secret-key-2026'
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(BASE_DIR, 'data.db')
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
    app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
    app.config['HRMS_BASE_DIR'] = BASE_DIR
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    db.init_app(app)
    # 多公司（独立 SQLite 文件）：占位库仅作兜底（未选公司/未登录时 Setting 等查询不报错）
    _ph = create_engine('sqlite:///' + os.path.join(BASE_DIR, 'company_core.db'))
    try:
        db.metadata.create_all(_ph, tables=COMPANY_TABLES)
    except Exception:
        pass
    app.config['CDB_PLACEHOLDER_ENGINE'] = _ph
    _models.DEFAULT_COMPANY_ENGINE = _ph  # 供 MultiTenantSession.get_bind 兜底

    login_mgr = LoginManager(app)
    login_mgr.login_view = 'login'
    login_mgr.login_message = '请先登录'

    @login_mgr.user_loader
    def load_user(uid):
        return db.session.get(User, int(uid))

    # 每请求：已登录但未选定公司时，引导到选择页（仅 1 个公司则自动选定）
    @app.before_request
    def ensure_company_selected():
        if not current_user.is_authenticated:
            return
        ep = request.endpoint
        # 以下端点无需已选公司
        if ep in ('login', 'logout', 'after_login', 'select_company', 'set_company',
                  'companies', 'api_companies', 'api_company_create',
                  'api_set_company', 'api_user_companies', 'static'):
            return
        # 公司 Logo 图片公开路由：未选公司（如选择公司页/左上角 img）也需可加载
        if request.path.startswith('/company-logo'):
            return
        cid = session.get('company_id')
        if cid and Company.query.get(cid):
            session.pop('pending_company_select', None)
            return
        comps = accessible_companies(current_user)
        if len(comps) == 1:
            session['company_id'] = comps[0].id
            session.pop('pending_company_select', None)
            return
        # 多公司且未选定：若是登录后的弹窗流程，则允许进入页面并叠加选择弹窗
        if session.get('pending_company_select'):
            g.show_company_modal = True
            return
        # 需要选择；页面请求跳转，API 请求返回 401
        if request.path.startswith('/api/'):
            return  # 具体接口在无公司时本就不该被调用（页面会先跳选择页）
        return redirect(url_for('select_company'))

    # 每请求：把当前选中的公司库引擎写入 contextvar，供 MultiTenantSession.get_bind 动态路由
    @app.before_request
    def bind_company_engine_ctx():
        if not current_user.is_authenticated:
            return
        cid = session.get('company_id')
        eng = company_engine(cid) if cid else current_app.config['CDB_PLACEHOLDER_ENGINE']
        current_company_engine.set(eng)

    # ---------------- 页面路由 ----------------
    @app.route('/')
    @login_required
    def dashboard():
        today = date.today()
        in30 = today + timedelta(days=30)

        # 合同到期提醒：30天内到期且未终止（员工按范围过滤）
        contracts_q = Contract.query.filter(
            Contract.status == '生效',
            Contract.end_date.isnot(None),
            Contract.end_date <= in30
        ).order_by(Contract.end_date.asc())
        if current_user.role == 'employee':
            allowed_ids = [e.id for e in employee_scope_filter(Employee.query).all()]
            contracts_q = contracts_q.filter(
                Contract.employee_id.in_(allowed_ids)) if allowed_ids else contracts_q.filter(Contract.employee_id == -1)
        contracts_q = contracts_q.all()
        expiring = []
        for c in contracts_q:
            days = c.days_to_expiry()
            if days is None:
                continue
            level = 'red' if days < 0 else ('yellow' if days <= 7 else 'normal')
            expiring.append({
                'id': c.id, 'emp_name': c.employee.name,
                'contract_no': c.contract_no or '',
                'department': c.employee.department or '-',
                'end_date': c.end_date.strftime('%Y-%m-%d'),
                'days': days, 'level': level,
            })

        # 入在离统计（员工按范围过滤）
        emp_base = employee_scope_filter(Employee.query)
        in_service = emp_base.filter_by(status='在职').count()
        resigned = emp_base.filter_by(status='离职').count()
        retired = emp_base.filter_by(status='退休').count()
        month_start = today.replace(day=1)
        new_hires = emp_base.filter(
            Employee.hire_date >= month_start, Employee.hire_date <= today
        ).count()
        leavers = emp_base.filter(
            Employee.leave_date.isnot(None),
            Employee.leave_date >= month_start, Employee.leave_date <= today
        ).count()

        # 当月生日提醒：生日月份 == 当前月（员工按范围过滤）
        birth_q = employee_scope_filter(Employee.query).filter(
            Employee.birthday.isnot(None), Employee.status != '离职'
        )
        birthdays = []
        for e in birth_q.all():
            if e.birthday.month != today.month:
                continue
            # 防止 2/29 在非闰年 replace 抛 ValueError 导致首页 500
            try:
                yb = today.replace(month=e.birthday.month, day=e.birthday.day)
            except ValueError:
                yb = today.replace(month=e.birthday.month, day=28)
            days = (yb - today).days  # 今年生日距今天数（负数表示已过的本月生日）
            birthdays.append({
                'id': e.id, 'name': e.name, 'department': e.department or '-',
                'birthday': e.birthday.strftime('%Y-%m-%d'),
                'day': e.birthday.day, 'days': days,
            })
        birthdays.sort(key=lambda x: (x['days'] < 0, x['day']))

        # 我的待办 + 今日提醒（仅当前公司）
        cid = current_company().id if current_company() else None
        todos_q = Todo.query.filter_by(user_id=current_user.id, done=False)
        if cid is not None:
            todos_q = todos_q.filter(Todo.company_id == cid)
        todos_today = todos_q.filter(Todo.due_date <= today).order_by(Todo.due_date.asc()).all()
        upcoming = todos_q.filter(Todo.due_date > today).order_by(Todo.due_date.asc()).limit(8).all()

        return render_template('dashboard.html',
                               expiring=expiring, in_service=in_service,
                               resigned=resigned, retired=retired,
                               new_hires=new_hires,
                               leavers=leavers, todos_today=todos_today,
                               birthdays=birthdays, upcoming=upcoming,
                               today=today)

    # 登录防爆破计数：{ "username|ip": [失败次数, 首次失败时间戳] }（内存态，进程重启即清零）
    _login_fails = {}

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('after_login'))
        if request.method == 'POST':
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            # 登录防爆破：按「用户名+IP」计数，5 次失败锁定 10 分钟
            key = f"{username}|{request.remote_addr or ''}"
            now = _time.time()
            rec = _login_fails.get(key)
            if rec and rec[0] >= 5 and now - rec[1] < 600:
                flash('尝试次数过多，请 10 分钟后再试', 'danger')
                return render_template('login.html')
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                _login_fails.pop(key, None)
                login_user(user)
                return redirect(url_for('after_login'))
            # 记录失败（成功登录或超 10 分钟后重置）
            if rec and now - rec[1] < 600:
                _login_fails[key] = [rec[0] + 1, rec[1]]
            else:
                _login_fails[key] = [1, now]
            flash('用户名或密码错误', 'danger')
        return render_template('login.html')

    @app.route('/after-login')
    @login_required
    def after_login():
        """登录后：按可访问公司数决定是否强制弹出「选择公司」浮层。"""
        comps = accessible_companies(current_user)
        if len(comps) <= 1:
            # 0/1 家公司：自动选定（无则交由页面提示新建），不弹窗
            if comps:
                session['company_id'] = comps[0].id
            session.pop('pending_company_select', None)
            return redirect(url_for('dashboard'))
        # 多公司：清除已选公司，强制弹出选择浮层（每次登录均可重新选择）
        session.pop('company_id', None)
        session['pending_company_select'] = '1'
        return redirect(url_for('dashboard'))

    @app.route('/select-company')
    @login_required
    def select_company():
        comps = accessible_companies(current_user)
        is_admin = (current_user.role == 'admin')
        has_any = Company.query.filter_by(enabled=True).first() is not None
        return render_template('select_company.html',
                               companies=comps, is_admin=is_admin,
                               need_create=(is_admin and not has_any))

    @app.route('/set-company/<int:cid>')
    @login_required
    def set_company(cid):
        if not user_can_access_company(current_user, cid):
            flash('无权访问该公司', 'danger')
            return redirect(url_for('select_company'))
        session['company_id'] = cid
        return redirect(url_for('dashboard'))

    @app.route('/api/set-company', methods=['POST'])
    @login_required
    def api_set_company():
        cid = request.get_json(force=True).get('company_id')
        if not cid or not user_can_access_company(current_user, int(cid)):
            return jsonify({'error': '无权访问该公司'}), 403
        session['company_id'] = int(cid)
        session.pop('pending_company_select', None)
        return jsonify({'msg': '已切换', 'company_id': int(cid)})

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('login'))

    # ---- 劳动合同 ----
    @app.route('/contracts')
    @login_required
    def contracts_page():
        return render_template('contracts.html')

    # ---- 入在离管理 ----
    def compute_dept_archives():
        """每家客户单位最后一个（最大）档案编号：按编号数字后缀自然排序取最大，
        避免 DF-KG-9 > DF-KG-10 这类字符串比较错误。"""
        try:
            rows = Employee.query.with_entities(
                Employee.department, Employee.archive_no
            ).filter(Employee.department.isnot(None),
                     Employee.department != '',
                     Employee.archive_no.isnot(None),
                     Employee.archive_no != '').all()
            def _no_key(no):
                nums = re.findall(r'\d+', no or '')
                return (int(nums[-1]), no) if nums else (0, no or '')
            m = {}
            for dept, no in rows:
                if _no_key(no) > _no_key(m.get(dept, '')):
                    m[dept] = no
            return [{'department': d, 'last_archive_no': m[d]} for d in sorted(m)]
        except Exception:
            return []

    @app.route('/api/dept-archives')
    @login_required
    def api_dept_archives():
        return jsonify({'dept_archives': compute_dept_archives()})

    @app.route('/lifecycle')
    @login_required
    def lifecycle_page():
        return render_template('lifecycle.html', dept_archives=compute_dept_archives(),
                               is_admin=(current_user.role == 'admin'))

    # ---- 五险一金导入 ----
    @app.route('/insurance/import')
    @role_required('admin', 'hr', 'employee')
    def insurance_import_page():
        return render_template('insurance_import.html')

    # ---- 五险一金汇总 ----
    @app.route('/insurance/summary')
    @login_required
    def insurance_summary_page():
        return render_template('insurance_summary.html')

    # ---- 待办 ----
    @app.route('/todos')
    @login_required
    def todos_page():
        return render_template('todos.html')

    # ---- 操作记录查询 ----
    @app.route('/logs')
    @role_required('admin', 'hr')
    def logs_page():
        return render_template('logs.html')

    # ---- 账号管理 ----
    @app.route('/users')
    @login_required
    def users_page():
        return render_template('users.html')

    # ---- 代码字段维护 ----
    @app.route('/dicts')
    @role_required('admin', 'hr')
    def dicts_page():
        return render_template('dicts.html')

    # ---- 数据备份 ----
    @app.route('/backup')
    @role_required('admin', 'hr')
    def backup_page():
        return render_template('backup.html')

    # ================== API ==================
    # ---------- 劳动合同 ----------
    @app.route('/api/contracts')
    @login_required
    def api_contracts_list():
        q = Contract.query.join(Employee)
        # 员工只看自己可见范围（含未绑定员工账号：employee_scope_filter 兜底为空集）
        if current_user.role == 'employee':
            scope_ids = employee_scope_filter(Employee.query).with_entities(Employee.id).subquery()
            q = q.filter(Contract.employee_id.in_(scope_ids))

        kw = request.args.get('keyword')
        if kw:
            q = q.filter(Employee.name.contains(kw) | Contract.contract_no.contains(kw))
        status = request.args.get('status')
        if status:
            q = q.filter(Contract.status == status)
        # 到期筛选
        expiry = request.args.get('expiry')  # 30/7/overdue
        today = date.today()
        if expiry == '30':
            q = q.filter(Contract.end_date.isnot(None),
                         Contract.end_date <= today + timedelta(days=30),
                         Contract.end_date >= today)
        elif expiry == '7':
            q = q.filter(Contract.end_date.isnot(None),
                         Contract.end_date <= today + timedelta(days=7),
                         Contract.end_date >= today)
        elif expiry == 'overdue':
            q = q.filter(Contract.end_date.isnot(None), Contract.end_date < today,
                         Contract.status == '生效')

        rows = q.order_by(Contract.end_date.asc().nullslast()).all()
        return jsonify({'items': [serialize_contract(c) for c in rows]})

    @app.route('/api/contracts/export')
    @login_required
    def api_contracts_export():
        """导出合同台账（按当前筛选条件），列：员工姓名/档案编号/合同类型/开始日期/到期日期/剩余天数/状态/备注"""
        q = Contract.query.join(Employee)
        if current_user.role == 'employee':
            scope_ids = employee_scope_filter(Employee.query).with_entities(Employee.id).subquery()
            q = q.filter(Contract.employee_id.in_(scope_ids))
        kw = request.args.get('keyword')
        if kw:
            q = q.filter(Employee.name.contains(kw) | Contract.contract_no.contains(kw))
        status = request.args.get('status')
        if status:
            q = q.filter(Contract.status == status)
        expiry = request.args.get('expiry')
        today = date.today()
        if expiry == '30':
            q = q.filter(Contract.end_date.isnot(None),
                         Contract.end_date <= today + timedelta(days=30),
                         Contract.end_date >= today)
        elif expiry == '7':
            q = q.filter(Contract.end_date.isnot(None),
                         Contract.end_date <= today + timedelta(days=7),
                         Contract.end_date >= today)
        elif expiry == 'overdue':
            q = q.filter(Contract.end_date.isnot(None), Contract.end_date < today,
                         Contract.status == '生效')
        rows = q.order_by(Contract.end_date.asc().nullslast()).all()

        data = []
        for c in rows:
            days = ''
            if c.end_date:
                d = (c.end_date - today).days
                days = f'{d}天' if d >= 0 else f'已过期 {-d}天'
            data.append({
                '员工姓名': c.employee.name if c.employee else '',
                '档案编号': c.contract_no or '',
                '合同类型': c.contract_type or '',
                '开始日期': c.start_date.strftime('%Y-%m-%d') if c.start_date else '',
                '到期日期': c.end_date.strftime('%Y-%m-%d') if c.end_date else '无固定期限',
                '剩余天数': days,
                '状态': c.status or '',
                '备注': c.remark or '',
            })
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df = pd.DataFrame(data)
            if not data:
                df = pd.DataFrame([{k: '' for k in
                                    ['员工姓名', '档案编号', '合同类型', '开始日期',
                                     '到期日期', '剩余天数', '状态', '备注']}])
            df.to_excel(writer, index=False, sheet_name='劳动合同台账')
        buf.seek(0)
        fname = f'劳动合同台账_{today.strftime("%Y%m%d")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/contracts', methods=['POST'])
    @role_required('admin', 'hr')
    def api_contracts_create():
        data = request.get_json(force=True)
        emp = db.session.get(Employee, data['employee_id'])
        # 档案编号始终取员工基本信息档案编号（新增合同不再单独维护、取消手动关联）
        contract_no = (data.get('contract_no') or '').strip() or (emp.archive_no if emp else '')
        c = Contract(
            employee_id=data['employee_id'],
            contract_no=contract_no,
            start_date=parse_date(data.get('start_date')),
            end_date=parse_date(data.get('end_date')) if data.get('end_date') else None,
            contract_type=data.get('contract_type', '固定期限'),
            status=data.get('status', '生效'),
            renew_count=data.get('renew_count', 0),
            remark=data.get('remark', '')
        )
        db.session.add(c)
        db.session.flush()
        db.session.commit()
        log_operation('合同', '新增', f'{emp.name if emp else ""} {c.contract_no}',
                      f'{c.contract_type}，{c.start_date}~{c.end_date or "无固定期限"}')
        return jsonify({'id': c.id, 'msg': '已创建'}), 201

    @app.route('/api/contracts/<int:cid>', methods=['PUT'])
    @role_required('admin', 'hr')
    def api_contracts_update(cid):
        c = db.session.get(Contract, cid)
        if not c:
            return jsonify({'error': '不存在'}), 404
        emp = db.session.get(Employee, c.employee_id)
        data = request.get_json(force=True)
        for f in ['employee_id', 'contract_type', 'status', 'remark']:
            if f in data:
                setattr(c, f, data[f])
        # 档案编号不再由前端维护：始终与员工基本信息档案编号保持一致（取消手动关联）
        if emp and emp.archive_no:
            c.contract_no = emp.archive_no
        if 'renew_count' in data:
            c.renew_count = data['renew_count']
        # 合同时间（起止日期）修改：仅管理员有权限
        new_start = parse_date(data['start_date']) if 'start_date' in data else c.start_date
        new_end = (parse_date(data['end_date']) if data['end_date'] else None) if 'end_date' in data else c.end_date
        if (new_start != c.start_date or new_end != c.end_date) and current_user.role != 'admin':
            return jsonify({'error': '劳动合同起止时间仅管理员可修改，如需变更请联系管理员'}), 403
        if 'start_date' in data:
            c.start_date = new_start
        if 'end_date' in data:
            c.end_date = new_end
        db.session.commit()
        log_operation('合同', '编辑', f'{emp.name if emp else ""} {c.contract_no}',
                      f'{c.contract_type}，{c.start_date}~{c.end_date or "无固定期限"}')
        return jsonify({'msg': '已更新'})

    @app.route('/api/contracts/<int:cid>', methods=['DELETE'])
    @role_required('admin', 'hr')
    def api_contracts_delete(cid):
        c = db.session.get(Contract, cid)
        if not c:
            return jsonify({'error': '不存在'}), 404
        emp = db.session.get(Employee, c.employee_id)
        log_operation('合同', '删除', f'{emp.name if emp else ""} {c.contract_no}', '删除合同记录')
        db.session.delete(c)
        db.session.commit()
        return jsonify({'msg': '已删除'})

    # ---------- 合同续签记录 ----------
    @app.route('/api/contracts/<int:cid>/renews', methods=['GET'])
    @login_required
    def api_contract_renews(cid):
        c = db.session.get(Contract, cid)
        if not c:
            return jsonify({'error': '不存在'}), 404
        # 员工角色：仅能查看自己可见范围内的合同续签记录
        if current_user.role == 'employee':
            scope_q = employee_scope_filter(Employee.query).filter(Employee.id == c.employee_id)
            if not scope_q.first():
                return jsonify({'error': '无权查看该合同续签记录'}), 403
        return jsonify({'items': [serialize_renew(r) for r in c.renews.all()]})

    @app.route('/api/contracts/<int:cid>/renews', methods=['POST'])
    @role_required('admin', 'hr')
    def api_contract_renew_create(cid):
        c = db.session.get(Contract, cid)
        if not c:
            return jsonify({'error': '不存在'}), 404
        data = request.get_json(force=True)
        old_end = c.end_date  # 记录续签前的到期日
        r = ContractRenew(
            contract_id=c.id,
            renew_date=parse_date(data.get('renew_date')) or date.today(),
            old_end_date=parse_date(data.get('old_end_date')) if data.get('old_end_date') else old_end,
            new_end_date=parse_date(data.get('new_end_date')) if data.get('new_end_date') else None,
            remark=data.get('remark', ''),
            operator=current_user.name
        )
        db.session.add(r)
        # 更新合同到期日和续签次数
        if data.get('new_end_date'):
            c.end_date = parse_date(data['new_end_date'])
        c.renew_count = (c.renew_count or 0) + 1
        if not data.get('old_end_date'):
            r.old_end_date = old_end
        db.session.commit()
        emp = db.session.get(Employee, c.employee_id)
        log_operation('合同', '续签', f'{emp.name if emp else ""} {c.contract_no}',
                      f'到期日 {r.old_end_date} → {r.new_end_date}，第 {c.renew_count} 次续签')
        return jsonify({'id': r.id, 'renew_count': c.renew_count, 'msg': '续签成功'}), 201

    @app.route('/api/contract_renews/<int:rid>', methods=['DELETE'])
    @role_required('admin', 'hr')
    def api_contract_renew_delete(rid):
        r = db.session.get(ContractRenew, rid)
        if not r:
            return jsonify({'error': '不存在'}), 404
        c = db.session.get(Contract, r.contract_id)
        emp = db.session.get(Employee, c.employee_id) if c else None
        log_operation('合同', '删除续签', f'{emp.name if emp else ""} {c.contract_no if c else ""}',
                      f'续签日期 {r.renew_date}，到期日 {r.old_end_date} → {r.new_end_date}')
        if c:
            c.renew_count = max((c.renew_count or 0) - 1, 0)
        db.session.delete(r)
        db.session.commit()
        return jsonify({'msg': '已删除'})

    # ---------- 附件管理 ----------
    @app.route('/api/attachments')
    @login_required
    def api_attachments_list():
        emp_id = request.args.get('employee_id', type=int)
        q = Attachment.query
        if current_user.role == 'employee':
            # 员工只能看自己可见范围内的附件（含未绑定账号：兜底为空集）
            scope_ids = employee_scope_filter(Employee.query).with_entities(Employee.id).subquery()
            q = q.filter(Attachment.employee_id.in_(scope_ids))
            if emp_id:
                q = q.filter_by(employee_id=emp_id)
        elif emp_id:
            q = q.filter_by(employee_id=emp_id)
        rows = q.order_by(Attachment.id.desc()).all()
        return jsonify({'items': [serialize_attachment(a) for a in rows]})

    @app.route('/api/attachments', methods=['POST'])
    @role_required('admin', 'hr')
    def api_attachments_upload():
        emp_id = request.form.get('employee_id', type=int)
        if not emp_id:
            return jsonify({'error': '缺少员工ID'}), 400
        emp = db.session.get(Employee, emp_id)
        if not emp:
            return jsonify({'error': '员工不存在'}), 404
        f = request.files.get('file')
        if not f or not f.filename:
            return jsonify({'error': '请选择文件'}), 400
        # 安全存储：保留扩展名，用时间戳+随机数命名
        import uuid
        ext = os.path.splitext(f.filename)[1].lower()
        stored = f"{emp_id}_{uuid.uuid4().hex[:10]}{ext}"
        upload_dir = current_app.config['UPLOAD_FOLDER']
        f.save(os.path.join(upload_dir, stored))
        a = Attachment(
            employee_id=emp_id,
            filename=stored,
            orig_name=f.filename,
            size=f.seek(0, 2) or 0,  # 获取文件大小
            category=request.form.get('category', '其他'),
            uploader=current_user.name
        )
        db.session.add(a)
        db.session.commit()
        log_operation('附件', '上传', f'{emp.name} {f.filename}', f'大小 {a.size} 字节，分类 {a.category}')
        return jsonify({'id': a.id, 'msg': '上传成功'}), 201

    @app.route('/api/attachments/<int:aid>/download')
    @login_required
    def api_attachments_download(aid):
        a = db.session.get(Attachment, aid)
        if not a:
            return abort(404)
        # 员工角色：仅能下载自己可见范围内员工的附件（含未绑定账号：兜底空集）
        if current_user.role == 'employee':
            scope_q = employee_scope_filter(Employee.query).filter(Employee.id == a.employee_id)
            if not scope_q.first():
                return abort(403)
        path = os.path.join(current_app.config['UPLOAD_FOLDER'], a.filename)
        if not os.path.exists(path):
            return jsonify({'error': '文件不存在'}), 404
        return send_file(path, as_attachment=True, download_name=a.orig_name)

    @app.route('/api/attachments/<int:aid>', methods=['DELETE'])
    @role_required('admin', 'hr')
    def api_attachments_delete(aid):
        a = db.session.get(Attachment, aid)
        if not a:
            return jsonify({'error': '不存在'}), 404
        emp = db.session.get(Employee, a.employee_id)
        log_operation('附件', '删除', f'{emp.name if emp else ""} {a.orig_name}', '删除附件文件')
        path = os.path.join(current_app.config['UPLOAD_FOLDER'], a.filename)
        if os.path.exists(path):
            os.remove(path)
        db.session.delete(a)
        db.session.commit()
        return jsonify({'msg': '已删除'})

    # ---------- 员工（入在离） ----------
    @app.route('/api/employees')
    @login_required
    def api_employees_list():
        q = employee_scope_filter(Employee.query)
        kw = request.args.get('keyword')
        if kw:
            q = q.filter(Employee.name.contains(kw) | Employee.id_card.contains(kw))
        st = request.args.get('status')
        if st:
            q = q.filter(Employee.status == st)
        dept = request.args.get('department')
        if dept:
            q = q.filter(Employee.department == dept)
        # 本月入职/本月离职筛选（period=month_hired / month_left）
        period = request.args.get('period')
        today = date.today()
        month_start = today.replace(day=1)
        next_month = (month_start + timedelta(days=31)).replace(day=1)
        if period == 'month_hired':
            q = q.filter(Employee.hire_date >= month_start, Employee.hire_date < next_month)
        elif period == 'month_left':
            q = q.filter(Employee.leave_date >= month_start, Employee.leave_date < next_month)
        rows = q.order_by(Employee.hire_date.desc()).all()
        return jsonify({'items': [serialize_employee(e) for e in rows]})

    @app.route('/api/employees/export')
    @login_required
    def api_employees_export():
        """导出员工信息台账（按当前筛选条件全量导出）"""
        q = employee_scope_filter(Employee.query)
        kw = request.args.get('keyword')
        if kw:
            q = q.filter(Employee.name.contains(kw) | Employee.id_card.contains(kw))
        st = request.args.get('status')
        if st:
            q = q.filter(Employee.status == st)
        dept = request.args.get('department')
        if dept:
            q = q.filter(Employee.department == dept)
        period = request.args.get('period')
        today = date.today()
        month_start = today.replace(day=1)
        next_month = (month_start + timedelta(days=31)).replace(day=1)
        if period == 'month_hired':
            q = q.filter(Employee.hire_date >= month_start, Employee.hire_date < next_month)
        elif period == 'month_left':
            q = q.filter(Employee.leave_date >= month_start, Employee.leave_date < next_month)
        rows = q.order_by(Employee.hire_date.desc()).all()

        def _s(v):
            return v.strftime('%Y-%m-%d') if v else ''
        data = []
        for e in rows:
            data.append({
                '姓名': e.name or '',
                '身份证号': e.id_card or '',
                '性别': e.gender or '',
                '出生日期': _s(e.birthday),
                '客户单位': e.department or '',
                '职位': e.position or '',
                '入职日期': _s(e.hire_date),
                '状态': e.status or '',
                '离职日期': _s(e.leave_date),
                '档案编号': e.archive_no or '',
                '电话': e.phone or '',
                '住址': e.address or '',
                '邮箱': e.email or '',
                '籍贯': e.native_place or '',
                '民族': e.ethnicity or '',
                '婚姻状况': e.marital_status or '',
                '学历': e.education or '',
                '毕业院校': e.school or '',
                '政治面貌': e.political_status or '',
                '紧急联系人': e.emergency_contact or '',
                '紧急联系电话': e.emergency_phone or '',
            })
        buf = io.BytesIO()
        cols = ['姓名', '身份证号', '性别', '出生日期', '客户单位', '职位', '入职日期',
                '状态', '离职日期', '档案编号', '电话', '住址', '邮箱', '籍贯', '民族',
                '婚姻状况', '学历', '毕业院校', '政治面貌', '紧急联系人', '紧急联系电话']
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df = pd.DataFrame(data, columns=cols)
            if not data:
                df = pd.DataFrame([{k: '' for k in cols}], columns=cols)
            df.to_excel(writer, index=False, sheet_name='员工信息台账')
        buf.seek(0)
        fname = f'员工信息台账_{today.strftime("%Y%m%d")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/employees/departments')
    @login_required
    def api_departments():
        # 员工角色自动按可见范围过滤客户单位（admin/hr 为全量）
        q = employee_scope_filter(Employee.query)
        rows = q.with_entities(Employee.department).distinct().all()
        return jsonify({'items': [r[0] for r in rows if r[0]]})

    @app.route('/api/employees', methods=['POST'])
    @role_required('admin', 'hr')
    def api_employees_create():
        data = request.get_json(force=True)
        e = Employee(
            name=data['name'],
            id_card=data.get('id_card', ''),
            phone=data.get('phone', ''),
            department=data.get('department', ''),
            position=data.get('position', ''),
            hire_date=parse_date(data.get('hire_date')) or date.today(),
            status='在职',
            gender=data.get('gender'),
            birthday=parse_date(data.get('birthday')) if data.get('birthday') else None,
            address=data.get('address', ''),
            archive_no=data.get('archive_no', ''),
            native_place=data.get('native_place', ''),
            ethnicity=data.get('ethnicity', ''),
            education=data.get('education', ''),
            school=data.get('school', ''),
            political_status=data.get('political_status', ''),
            marital_status=data.get('marital_status', ''),
            email=data.get('email', ''),
            emergency_contact=data.get('emergency_contact', ''),
            emergency_phone=data.get('emergency_phone', ''),
        )
        db.session.add(e)
        db.session.flush()
        # 若填了档案编号，自动同步创建/更新该员工的合同档案编号
        if e.archive_no:
            sync_contract_no(e)
        db.session.commit()
        log_operation('员工', '入职', e.name, f'客户单位 {e.department}，档案编号 {e.archive_no}')
        return jsonify({'id': e.id, 'msg': '入职登记成功'}), 201

    @app.route('/api/employees/<int:eid>', methods=['PUT'])
    @role_required('admin', 'hr')
    def api_employees_update(eid):
        e = db.session.get(Employee, eid)
        if not e:
            return jsonify({'error': '不存在'}), 404
        data = request.get_json(force=True)
        for f in ['name', 'id_card', 'phone', 'department', 'position', 'gender',
                  'address', 'archive_no', 'native_place', 'ethnicity', 'education',
                  'school', 'political_status', 'marital_status', 'email',
                  'emergency_contact', 'emergency_phone',
                  'pension_person_no', 'injury_person_no', 'medical_person_no',
                  'maternity_person_no', 'unemployment_person_no', 'fund_person_no']:
            if f in data:
                setattr(e, f, data[f])
        if 'hire_date' in data:
            e.hire_date = parse_date(data['hire_date']) or e.hire_date
        if 'birthday' in data and data['birthday']:
            e.birthday = parse_date(data['birthday'])
        # 参保时间（各险种）单独处理日期
        for f in ['pension_enroll_date', 'injury_enroll_date', 'medical_enroll_date',
                  'maternity_enroll_date', 'unemployment_enroll_date', 'fund_enroll_date']:
            if f in data:
                setattr(e, f, parse_date(data[f]) if data[f] else None)
        # 档案编号变更时同步合同档案编号
        if 'archive_no' in data:
            sync_contract_no(e)
        db.session.commit()
        log_operation('员工', '编辑', e.name, f'客户单位 {e.department}，职位 {e.position}')
        return jsonify({'msg': '已更新'})

    @app.route('/api/employees/<int:eid>/resign', methods=['POST'])
    @role_required('admin', 'hr')
    def api_employees_resign(eid):
        e = db.session.get(Employee, eid)
        if not e:
            return jsonify({'error': '不存在'}), 404
        data = request.get_json(force=True)
        new_status = data.get('status', '离职')
        if new_status not in ('离职', '退休'):
            new_status = '离职'
        e.status = new_status
        e.leave_date = parse_date(data.get('leave_date')) or date.today()
        if data.get('remark'):
            e.remark = data['remark']
        # 同步终止其生效合同
        for c in e.contracts.filter_by(status='生效').all():
            c.status = '终止'
        db.session.commit()
        log_operation('员工', new_status, e.name,
                      f'{new_status}日期 {e.leave_date}，备注 {data.get("remark") or "无"}')
        return jsonify({'msg': f'{new_status}登记成功'})

    @app.route('/api/employees/<int:eid>', methods=['DELETE'])
    @role_required('admin')
    def api_employees_delete(eid):
        e = db.session.get(Employee, eid)
        if not e:
            return jsonify({'error': '不存在'}), 404
        # 级联删除关联记录：外键 employee_id 为 NOT NULL，必须先清理子表再删员工，
        # 否则会触发 NOT NULL constraint failed: contracts.employee_id 等错误。
        # 工资记录(salary_records)的 employee_id 可为空且刻意保留冗余姓名，故不删除。
        # 注：contracts/insurance_details/renews 为 lazy='dynamic'，attachments/transfer_records
        # 为普通列表，统一用 list() 兼容两种关系。
        for c in list(e.contracts):               # 合同续签记录（引用 contracts.id）
            for r in list(c.renews):
                db.session.delete(r)
        for c in list(e.contracts):               # 劳动合同
            db.session.delete(c)
        for d in list(e.insurance_details):       # 五险一金明细
            db.session.delete(d)
        for a in list(e.attachments):             # 附件
            db.session.delete(a)
        for t in list(e.transfer_records):        # 转岗记录
            db.session.delete(t)
        db.session.delete(e)
        db.session.commit()
        log_operation('员工', '删除', e.name or '', f'客户单位 {e.department or ""}')
        return jsonify({'msg': '已删除'})

    # ---------- 员工批量导入 ----------
    EMPLOYEE_IMPORT_COLS = [
        '姓名', '身份证号', '性别', '出生日期', '客户单位', '职位', '入职日期',
        '电话', '住址', '邮箱', '籍贯', '民族', '婚姻状况', '学历', '毕业院校',
        '政治面貌', '紧急联系人', '紧急联系电话', '档案编号', '开始日期', '到期日期'
    ]

    @app.route('/api/employees/template')
    @role_required('admin', 'hr', 'employee')
    def api_employees_template():
        """下载员工批量导入模板"""
        sample = pd.DataFrame([{
            '姓名': '张三', '身份证号': '110101199001011234', '性别': '男',
            '出生日期': '1990-01-01', '客户单位': '研发部', '职位': '工程师',
            '入职日期': '2023-03-01', '电话': '13800001111', '住址': '北京市海淀区',
            '邮箱': 'zhangsan@example.com', '籍贯': '北京', '民族': '汉族',
            '婚姻状况': '已婚', '学历': '本科', '毕业院校': '某大学',
            '政治面貌': '群众', '紧急联系人': '张父', '紧急联系电话': '13900001111',
            '档案编号': 'HT-2023-001', '开始日期': '2023-03-01', '到期日期': '2026-02-28'
        }], columns=EMPLOYEE_IMPORT_COLS)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as w:
            sample.to_excel(w, index=False, sheet_name='员工导入模板')
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='员工批量导入模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/employees/import', methods=['POST'])
    @role_required('admin', 'hr')
    def api_employees_import():
        """批量导入员工"""
        f = request.files.get('file')
        if not f or not f.filename:
            return jsonify({'error': '请选择文件'}), 400
        name = f.filename.lower()
        try:
            if name.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(f.read()), dtype=str)
            else:
                df = pd.read_excel(io.BytesIO(f.read()), dtype=str)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {e}'}), 400

        df.columns = [str(c).strip() for c in df.columns]
        colmap = {
            '姓名': 'name', '员工姓名': 'name', '名字': 'name',
            '身份证号': 'id_card', '身份证': 'id_card', '身份证号码': 'id_card',
            '证件号码': 'id_card', '证件号': 'id_card',
            '性别': 'gender', '出生日期': 'birthday', '出生': 'birthday',
            '客户单位': 'department', '部门': 'department', '单位': 'department', '所在单位': 'department',
            '职位': 'position', '岗位': 'position',
            '入职日期': 'hire_date', '入职时间': 'hire_date',
            '电话': 'phone', '手机号': 'phone', '手机号码': 'phone',
            '住址': 'address', '家庭住址': 'address',
            '邮箱': 'email', '电子邮件': 'email',
            '籍贯': 'native_place', '民族': 'ethnicity',
            '婚姻状况': 'marital_status', '学历': 'education',
            '毕业院校': 'school', '政治面貌': 'political_status',
            '紧急联系人': 'emergency_contact', '紧急联系电话': 'emergency_phone',
            '档案编号': 'archive_no', '合同编号': 'archive_no',
            '开始日期': 'contract_start_date', '合同开始日期': 'contract_start_date',
            '到期日期': 'contract_end_date', '合同到期日期': 'contract_end_date',
        }
        rename = {c: colmap[c] for c in df.columns if c in colmap}
        df = df.rename(columns=rename)
        if 'name' not in df.columns or 'id_card' not in df.columns:
            return jsonify({'error': '缺少必需列：姓名/身份证号（请使用下载的模板；当前表头：' + '、'.join(map(str, df.columns)) + '）'}), 400

        success = 0
        skipped = 0
        duplicates = []
        errors = []

        def safe_str(row, v):
            val = str(row.get(v) or '').strip()
            return val if val and val.lower() != 'nan' else ''

        for ridx, (_, row) in enumerate(df.iterrows(), start=2):  # 第1行为表头
            name_v = str(row.get('name') or '').strip()
            if name_v.lower() == 'nan':
                name_v = ''
            id_card = str(row.get('id_card') or '').strip()
            if id_card.lower() == 'nan':
                id_card = ''
            if not name_v or not id_card:
                skipped += 1
                continue
            # 身份证号已存在则跳过（不覆盖）
            if Employee.query.filter_by(id_card=id_card).first():
                duplicates.append({'name': name_v, 'id_card': id_card})
                skipped += 1
                continue
            try:
                e = Employee(
                    name=name_v, id_card=id_card,
                    gender=safe_str(row, 'gender'),
                    birthday=parse_date(row.get('birthday')) if safe_str(row, 'birthday') else None,
                    department=safe_str(row, 'department'),
                    position=safe_str(row, 'position'),
                    hire_date=parse_date(row.get('hire_date')) or date.today(),
                    status='在职',
                    phone=safe_str(row, 'phone'),
                    address=safe_str(row, 'address'),
                    email=safe_str(row, 'email'),
                    native_place=safe_str(row, 'native_place'),
                    ethnicity=safe_str(row, 'ethnicity'),
                    marital_status=safe_str(row, 'marital_status'),
                    education=safe_str(row, 'education'),
                    school=safe_str(row, 'school'),
                    political_status=safe_str(row, 'political_status'),
                    emergency_contact=safe_str(row, 'emergency_contact'),
                    emergency_phone=safe_str(row, 'emergency_phone'),
                    archive_no=safe_str(row, 'archive_no'),
                )
                db.session.add(e)
                db.session.flush()
                if e.archive_no:
                    sync_contract_no(e)
                # 劳动合同开始/到期日期：若提供开始日期，则创建/更新该员工劳动合同
                # 到期日期未填写 → 无固定期限（end_date 留空、contract_type=无固定期限）
                cs_raw = row.get('contract_start_date')
                cs_date = parse_date(cs_raw) if cs_raw and str(cs_raw).strip() and str(cs_raw).strip().lower() != 'nan' else None
                if cs_date:
                    ce_raw = row.get('contract_end_date')
                    ce_date = parse_date(ce_raw) if ce_raw and str(ce_raw).strip() and str(ce_raw).strip().lower() != 'nan' else None
                    c = e.contracts.filter_by(status='生效').first() or e.contracts.first()
                    if not c:
                        c = Contract(
                            employee_id=e.id,
                            contract_no=e.archive_no or f'HT-{e.id}',
                            start_date=cs_date,
                            end_date=ce_date,  # None = 无固定期限
                            contract_type=('固定期限' if ce_date else '无固定期限'),
                            status='生效'
                        )
                        db.session.add(c)
                    else:
                        c.start_date = cs_date
                        c.end_date = ce_date
                        c.contract_type = '固定期限' if ce_date else '无固定期限'
                    db.session.flush()
                db.session.commit()  # 逐行提交，避免单条异常连累整体导入
                success += 1
            except Exception as ex:
                db.session.rollback()
                errors.append({'row': ridx, 'name': name_v, 'id_card': id_card, 'error': str(ex)})
                continue

        log_operation('员工', '批量导入', f'导入 {success} 人',
                      f'成功 {success}，跳过 {skipped}（含重复 {len(duplicates)}），异常 {len(errors)}')
        msg = f'导入完成，成功 {success} 人，跳过 {skipped} 人'
        if duplicates:
            msg += f'，重复 {len(duplicates)} 人'
        if errors:
            msg += f'，异常 {len(errors)} 行'
        return jsonify({
            'msg': msg,
            'success': success, 'skipped': skipped,
            'duplicates': duplicates,
            'errors': errors,
        })

    @app.route('/api/employees/<int:eid>/photo', methods=['POST'])
    @role_required('admin', 'hr')
    def api_employees_photo(eid):
        """上传/修改个人证件照"""
        e = db.session.get(Employee, eid)
        if not e:
            return jsonify({'error': '员工不存在'}), 404
        f = request.files.get('file')
        if not f or not f.filename:
            return jsonify({'error': '请选择图片文件'}), 400
        ALLOWED = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
        ext = os.path.splitext(f.filename)[1].lower().lstrip('.')
        if ext not in ALLOWED:
            return jsonify({'error': '仅支持图片文件（png/jpg/jpeg/gif/bmp/webp）'}), 400
        # 删除旧照片（若存在）
        if e.photo:
            old_path = os.path.join(current_app.config['UPLOAD_FOLDER'], e.photo)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except OSError:
                    pass
        import uuid
        stored = f"emp_{eid}_{uuid.uuid4().hex[:10]}.{ext}"
        f.save(os.path.join(current_app.config['UPLOAD_FOLDER'], stored))
        e.photo = stored
        db.session.commit()
        log_operation('员工', '上传证件照', e.name, '上传/修改个人证件照')
        return jsonify({'photo': stored, 'msg': '证件照已更新'})

    @app.route('/api/employees/<int:eid>/photo')
    @login_required
    def api_employees_photo_get(eid):
        """读取个人证件照（员工角色仅限可见范围内）"""
        e = db.session.get(Employee, eid)
        if not e or not e.photo:
            return abort(404)
        if current_user.role == 'employee':
            scope_q = employee_scope_filter(Employee.query).filter(Employee.id == eid)
            if not scope_q.first():
                return abort(403)
        path = os.path.join(current_app.config['UPLOAD_FOLDER'], e.photo)
        if not os.path.exists(path):
            return abort(404)
        return send_file(path, mimetype='image/jpeg' if not e.photo.lower().endswith('png') else 'image/png')

    # ---------- 五险一金：期间锁定（汇总数据确认） ----------
    def insurance_locked_periods():
        """已锁定的社保期间列表（存于 Setting，JSON 数组）"""
        try:
            return json.loads(Setting.get('insurance_locked_periods', '[]') or '[]')
        except Exception:
            return []

    def insurance_period_locked(period):
        return period in insurance_locked_periods()

    @app.route('/api/insurance/lock', methods=['POST'])
    @role_required('admin', 'hr')
    def api_insurance_lock():
        """汇总数据确认：锁定某期间社保数据（三次弹窗确认由前端完成）"""
        data = request.get_json(force=True) or {}
        period = str(data.get('period') or '').strip()
        if not period:
            return jsonify({'error': '请选择期间'}), 400
        if insurance_period_locked(period):
            return jsonify({'locked': True, 'period': period, 'msg': f'{period} 已锁定'})
        locked = insurance_locked_periods()
        locked.append(period)
        Setting.set('insurance_locked_periods', json.dumps(locked))
        db.session.commit()
        log_operation('五险一金', '锁定', period, '汇总数据确认，锁定期间社保数据（仅管理员可修改）')
        return jsonify({'locked': True, 'period': period, 'msg': f'{period} 已锁定，仅管理员可修改'})

    @app.route('/api/insurance/unlock', methods=['POST'])
    @role_required('admin')
    def api_insurance_unlock():
        """仅管理员可解除锁定"""
        data = request.get_json(force=True) or {}
        period = str(data.get('period') or '').strip()
        locked = [p for p in insurance_locked_periods() if p != period]
        Setting.set('insurance_locked_periods', json.dumps(locked))
        db.session.commit()
        log_operation('五险一金', '解锁', period, '解除期间社保数据锁定')
        return jsonify({'locked': False, 'period': period, 'msg': f'{period} 已解锁'})

    @app.route('/api/insurance/lock/status')
    @login_required
    def api_insurance_lock_status():
        period = request.args.get('period', '').strip()
        return jsonify({'locked': insurance_period_locked(period) if period else False,
                        'period': period})

    @app.route('/api/insurance/lock/list')
    @login_required
    def api_insurance_lock_list():
        """返回全部已锁定期间（供导入月份选择器禁用已锁定月份）"""
        return jsonify({'locked_periods': insurance_locked_periods()})

    # ---------- 五险一金：明细列表 ----------
    @app.route('/api/insurance/details')
    @login_required
    def api_insurance_list():
        q = InsuranceDetail.query.join(Employee)
        if current_user.role == 'employee':
            q = employee_scope_filter(q)
        period = request.args.get('period')
        if period:
            q = q.filter(InsuranceDetail.period == period)
        kw = request.args.get('keyword')
        if kw:
            q = q.filter(Employee.name.contains(kw))
        rows = q.order_by(InsuranceDetail.period.desc()).all()
        return jsonify({'items': [serialize_insurance(d) for d in rows]})

    @app.route('/api/insurance/periods')
    @login_required
    def api_periods():
        q = db.session.query(InsuranceDetail.period)
        if current_user.role == 'employee':
            q = q.join(Employee)
            q = employee_scope_filter(q)
        rows = q.distinct().order_by(InsuranceDetail.period.desc()).all()
        return jsonify({'items': [r[0] for r in rows]})

    # ---------- 五险一金：导入（按险种分表） ----------
    # 险种类型 → (目标字段, 是否涉及公积金基数/比例)
    INSURANCE_TYPES = {
        '养老单位': ('pension_emp', False),
        '养老个人': ('pension_per', False),
        '工伤单位': ('injury_emp', False),
        '医疗单位': ('medical_emp', False),
        '医疗个人': ('medical_per', False),
        '大额医疗补助单位': ('extra_medical_emp', False),
        '大额医疗补助个人': ('extra_medical_per', False),
        '生育单位': ('maternity_emp', False),
        '失业单位': ('unemployment_emp', False),
        '失业个人': ('unemployment_per', False),
        '公积金单位': ('fund_emp', True),
        '公积金个人': ('fund_per', False),
    }
    # 险种类型 → 参保信息字段前缀（参保时间 + 个人编号）
    # 首次导入时填充，后续导入不覆盖
    ENROLLMENT_MAP = {
        '养老单位': 'pension', '养老个人': 'pension',
        '工伤单位': 'injury',
        '医疗单位': 'medical', '医疗个人': 'medical',
        '大额医疗补助单位': 'medical', '大额医疗补助个人': 'medical',
        '生育单位': 'maternity',
        '失业单位': 'unemployment', '失业个人': 'unemployment',
        '公积金单位': 'fund', '公积金个人': 'fund',
    }

    def leave_remark(emp, base=''):
        """员工备注：退休员工固定填入『退休人员』；离职员工在原有备注后追加离职时间"""
        parts = [p for p in (base or '').split('；') if p.strip()]
        if emp.status == '退休':
            if '退休人员' not in parts:
                parts.append('退休人员')
        if emp.status == '离职' and emp.leave_date:
            parts.append('离职时间：{:04d}年{:02d}月{:02d}日'.format(
                emp.leave_date.year, emp.leave_date.month, emp.leave_date.day))
        return '；'.join(parts)

    def apply_first_enrollment(emp, period, person_no, ins_type):
        """首次导入自动关联参保时间与个人编号；已存在则不覆盖"""
        prefix = ENROLLMENT_MAP.get(ins_type)
        if not prefix:
            return
        if not getattr(emp, f'{prefix}_enroll_date'):
            try:
                enroll_date = datetime.strptime(period, '%Y-%m').date().replace(day=1)
            except (ValueError, TypeError):
                enroll_date = date.today()
            setattr(emp, f'{prefix}_enroll_date', enroll_date)
        if person_no and not getattr(emp, f'{prefix}_person_no'):
            setattr(emp, f'{prefix}_person_no', person_no)

    @app.route('/api/insurance/import', methods=['POST'])
    @role_required('admin', 'hr')
    def api_insurance_import():
        f = request.files.get('file')
        if not f:
            return jsonify({'error': '未上传文件'}), 400
        period = (request.form.get('period') or '').strip()
        if not period:
            return jsonify({'error': '请选择导入月份'}), 400
        ins_type = (request.form.get('insurance_type') or '').strip()
        if ins_type not in INSURANCE_TYPES:
            return jsonify({'error': '请选择正确的险种类型'}), 400
        target_field, is_fund = INSURANCE_TYPES[ins_type]

        # 解析文件（模板列：序号/员工姓名/身份证号/缴费基数/应缴费额/人员编号）
        name = f.filename.lower()
        try:
            if name.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(f.read()), dtype=str)
            else:
                df = pd.read_excel(io.BytesIO(f.read()), dtype=str)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {e}'}), 400

        df.columns = [str(c).strip() for c in df.columns]
        colmap = {
            '序号': 'seq', '员工姓名': 'name', '姓名': 'name',
            '身份证号': 'id_card', '身份证': 'id_card',
            '缴费基数': 'base', '基数': 'base',
            '缴费比例': 'fund_rate', '比例': 'fund_rate',
            '应缴费额': 'amount', '缴费金额': 'amount', '金额': 'amount',
            '人员编号': 'person_no', '人员号': 'person_no', '编号': 'person_no',
        }
        # 处理重复列名：旧模板中可能存在多列「缴费基数」，仅保留第一个作为 base（避免 rename 冲突）
        _seen_base = False
        rename = {}
        for c in df.columns:
            if c in colmap:
                if colmap[c] == 'base' and _seen_base:
                    continue
                rename[c] = colmap[c]
                if colmap[c] == 'base':
                    _seen_base = True
        df = df.rename(columns=rename)
        if 'name' not in df.columns or 'id_card' not in df.columns:
            return jsonify({'error': '缺少必需列：员工姓名/身份证号（请使用下载的模板）'}), 400

        rows_data = []
        for _, row in df.iterrows():
            name_v = str(row.get('name') or '').strip()
            id_card = str(row.get('id_card') or '').strip()
            if not name_v or not id_card:
                continue
            def to_float(v):
                try:
                    f = float(str(v).replace(',', '').replace('，', '').strip() or 0)
                    if f != f:  # NaN 安全兜底
                        return 0.0
                    return round(f, 2)
                except (ValueError, TypeError):
                    return 0.0
            rows_data.append({
                'name': name_v,
                'id_card': id_card,
                'base': to_float(row.get('base')),
                'amount': to_float(row.get('amount')),
                'fund_rate': to_float(row.get('fund_rate')),
                'person_no': str(row.get('person_no') or '').strip(),
            })
        if not rows_data:
            return jsonify({'error': '文件中没有有效数据行'}), 400

        # 期间已锁定（汇总数据确认后）：仅管理员可继续导入
        if insurance_period_locked(period) and current_user.role != 'admin':
            return jsonify({'error': f'{period} 社保数据已锁定，仅管理员可修改'}), 403

        # 匹配员工（在职及离职）：身份证号存在即匹配；员工表中不存在的才列为未匹配。
        # 库内已存在同期间记录、或文件内同一身份证重复出现，均归入 duplicates 供弹窗确认合并
        matched = []      # 新导入（该员工该期间无记录，且文件内首次出现）
        duplicates = []   # 重复需确认：已导入 或 文件内重复行
        unmatched = []    # 员工表中不存在
        emp_cache = {}
        mask_cache = {}
        file_seen = set()
        for rd in rows_data:
            raw_ic = str(rd.get('id_card') or '').strip()
            emp = None
            if is_masked_id_card(raw_ic):
                # 脱敏身份证（如 532224********1125）：按「姓名 + 未脱敏部分」匹配
                emp, reason = match_employee_by_masked(rd['name'], raw_ic, mask_cache)
                if not emp:
                    unmatched.append({**rd, 'reason': reason})
                    continue
            else:
                # 身份证格式校验：非法格式直接列为未匹配，避免误匹配脏数据
                if not valid_id_card(raw_ic):
                    unmatched.append({**rd, 'reason': '身份证号格式不正确（应为15或18位）'})
                    continue
                if raw_ic in emp_cache:
                    emp = emp_cache[raw_ic]
                else:
                    emp = Employee.query.filter_by(id_card=raw_ic).first()
                    emp_cache[raw_ic] = emp
                if not emp:
                    unmatched.append({**rd, 'reason': '未匹配到人员（员工表中无此身份证）'})
                    continue
            det = InsuranceDetail.query.filter_by(employee_id=emp.id, period=period).first()
            # 同期间同险种已导入过（目标字段已有值）才视为重复；不同险种（字段为 0）直接导入互不影响
            if det is not None and float(getattr(det, target_field) or 0) != 0:
                duplicates.append({'employee_id': emp.id, **rd, 'emp_name': emp.name,
                                   'dup_type': '已导入过'})
            elif rd['id_card'] in file_seen:
                duplicates.append({'employee_id': emp.id, **rd, 'emp_name': emp.name,
                                   'dup_type': '文件内重复'})
            else:
                matched.append({'employee_id': emp.id, **rd, 'emp_name': emp.name})
                file_seen.add(rd['id_card'])

        # 检查阶段：返回重复项供前端确认（含未匹配列表）
        confirm = (request.form.get('confirm') or '') == '1'
        if duplicates and not confirm:
            return jsonify({
                'need_confirm': True,
                'matched_count': len(matched),
                'duplicates': duplicates,
                'unmatched': unmatched,
                'msg': f'检测到 {len(duplicates)} 条重复记录',
            })

        # 正式导入：按员工分组合并（同员工同期间只产生一条记录，避免 UNIQUE 冲突）
        # 备注按员工独立：remarks 为 {"员工id": "备注"}，只写入对应员工记录
        remarks = {}
        try:
            _r = request.form.get('remarks') or '{}'
            remarks = {int(k): str(v).strip() for k, v in json.loads(_r).items() if str(v).strip()}
        except Exception:
            remarks = {}
        groups = {}
        for rd in matched + duplicates:
            g = groups.setdefault(rd['employee_id'], {'base': 0.0, 'amount': 0.0, 'fund_rate': 0.0, 'person_no': '', 'override': False})
            g['base'] = max(g['base'], rd['base'])
            g['fund_rate'] = max(g['fund_rate'], rd.get('fund_rate') or 0)
            if rd['person_no'] and not g['person_no']:
                g['person_no'] = rd['person_no']
            if rd.get('dup_type') == '已导入过':
                # 库内已存在该期间记录：覆盖而非累加（避免同一文件重复导入金额翻倍）
                g['override'] = True
                g['amount'] = rd['amount']
            else:
                # 新导入 + 文件内重复行：合并计算（累加）
                g['amount'] = round(g['amount'] + rd['amount'], 2)
        success = 0
        for emp_id, g in groups.items():
            det = InsuranceDetail.query.filter_by(employee_id=emp_id, period=period).first()
            if det is None:
                det = InsuranceDetail(employee_id=emp_id, period=period,
                                      base=g['base'], person_no=g['person_no'])
                db.session.add(det)
            else:
                det.base = max(det.base or 0, g['base'])
                if g['person_no']:
                    det.person_no = g['person_no']
            # 累加金额；基数取较大值。已导入过的重复行走覆盖逻辑（不累加）
            if g.get('override'):
                setattr(det, target_field, round(g['amount'], 2))
            else:
                cur = float(getattr(det, target_field) or 0)
                setattr(det, target_field, round(cur + g['amount'], 2))
            if g.get('fund_rate'):
                # 模板中直接提供的缴费比例
                det.fund_rate = g['fund_rate']
            elif is_fund:
                # 公积金单位：未提供缴费比例时由单位额/基数推算
                if g['base']:
                    det.fund_base = max(det.fund_base or 0, g['base'])
                    det.fund_rate = round(float(getattr(det, target_field) or 0) / det.fund_base * 100, 2) if det.fund_base else 0
            # 仅对填写了备注的员工追加（相同备注只保留 1 条）
            remark_text = remarks.get(emp_id, '')
            if remark_text:
                parts = [p.strip() for p in (det.remark or '').split('；') if p.strip()]
                if remark_text not in parts:
                    parts.append(remark_text)
                det.remark = '；'.join(parts)
            success += 1

        # 首次导入自动关联参保时间与个人编号（后续导入不覆盖）
        for emp_id, g in groups.items():
            emp = db.session.get(Employee, emp_id)
            if emp:
                apply_first_enrollment(emp, period, g['person_no'], ins_type)

        db.session.commit()
        log_operation('五险一金', '导入', f'{period} {ins_type}',
                      f'成功 {success} 条，未匹配 {len(unmatched)} 条')
        return jsonify({
            'msg': f'导入完成，共 {success} 条（{ins_type}，{period}）',
            'success': success,
            'duplicates_count': len(duplicates),
            'unmatched': unmatched,
        })

    # ---------- 五险一金：导入后「修改人员」— 将未匹配行指定到现有员工 ----------
    @app.route('/api/insurance/assign', methods=['POST'])
    @role_required('admin', 'hr')
    def api_insurance_assign():
        """导入后对未匹配到人员的行，手动指定到现有员工。
        按导入时使用的 insurance_type，在该员工的 InsuranceDetail(period) 上写入对应金额。"""
        data = request.get_json(silent=True) or {}
        period = (data.get('period') or '').strip()
        ins_type = (data.get('insurance_type') or '').strip()
        target_emp_id = data.get('target_employee_id')
        base = data.get('base')
        amount = data.get('amount')
        person_no = (data.get('person_no') or '').strip()
        fund_rate = data.get('fund_rate')

        if not period or not ins_type or not target_emp_id:
            return jsonify({'error': '缺少必填参数（期间/险种/目标员工）'}), 400
        if ins_type not in INSURANCE_TYPES:
            return jsonify({'error': '险种类型不正确'}), 400
        target_field, is_fund = INSURANCE_TYPES[ins_type]

        emp = db.session.get(Employee, int(target_emp_id))
        if not emp:
            return jsonify({'error': '目标员工不存在'}), 404

        # 期间已锁定：仅管理员可修改
        if insurance_period_locked(period) and current_user.role != 'admin':
            return jsonify({'error': f'{period} 社保数据已锁定，仅管理员可修改'}), 403

        det = InsuranceDetail.query.filter_by(employee_id=emp.id, period=period).first()
        if not det:
            det = InsuranceDetail(employee_id=emp.id, period=period)
            db.session.add(det)

        def _f(v):
            try: return round(float(v or 0), 2)
            except (ValueError, TypeError): return 0.0

        b = _f(base)
        a = _f(amount)
        det.base = b
        if person_no:
            det.person_no = person_no
        if is_fund:
            det.fund_base = b
            fr = _f(fund_rate)
            if fr: det.fund_rate = fr
        setattr(det, target_field, a)

        # 首次参保信息关联
        apply_first_enrollment(emp, period, person_no, ins_type)

        db.session.commit()
        log_operation('五险一金', '指定人员', f'{period} {ins_type}',
                      f'将 {emp.name}(id={emp.id}) 的 {ins_type} 金额 {a} 写入')
        return jsonify({'msg': f'已将 {emp.name} 的 {ins_type}（{period}）金额 {a:.2f} 写入'})

    # ---------- 五险一金：汇总 ----------
    @app.route('/api/insurance/summary')
    @login_required
    def api_insurance_summary():
        period = request.args.get('period')

        # 参与单位/个人合计的字段（不含公积金基数与比例）
        unit_fields = ['pension_emp', 'medical_emp', 'extra_medical_emp',
                       'unemployment_emp', 'injury_emp', 'maternity_emp', 'fund_emp']
        per_fields = ['pension_per', 'medical_per', 'extra_medical_per',
                      'unemployment_per', 'fund_per']
        amount_fields = unit_fields + per_fields

        # 员工基础（在职+离职全量）；员工角色按查看范围过滤
        emp_q = Employee.query
        if current_user.role == 'employee':
            emp_q = employee_scope_filter(emp_q)
        emps = emp_q.order_by(Employee.id.asc()).all()

        # 期间明细映射 employee_id -> 明细
        det_q = InsuranceDetail.query
        if current_user.role == 'employee':
            det_q = det_q.join(Employee)
            det_q = employee_scope_filter(det_q)
        if period:
            det_q = det_q.filter(InsuranceDetail.period == period)
        det_map = {}
        # 未选择期间时按最新期间优先（period 降序），避免同员工多期间取到任意月份
        for d in det_q.order_by(InsuranceDetail.period.desc()).all():
            det_map.setdefault(d.employee_id, d)

        items = []
        total = {k: 0.0 for k in amount_fields}
        total['base'] = 0.0
        total['fund_base'] = 0.0
        seq = 0
        for emp in emps:
            d = det_map.get(emp.id)
            if d is None and not period:
                # 未选择期间时只展示有明细的员工（保持原行为）
                continue
            seq += 1
            if d is not None:
                item = serialize_insurance(d)
            else:
                # 该期间无数据：各金额显示 0
                item = {k: 0.0 for k in amount_fields}
                item.update({'base': 0.0, 'fund_base': 0.0, 'period': period,
                             'remark': ''})
            item['seq'] = seq
            item['emp_name'] = emp.name
            item['department'] = emp.department or ''
            # 离职员工：备注栏追加离职时间
            item['remark'] = leave_remark(emp, item.get('remark') or '')
            item['unit_total'] = round(sum(float(item.get(f) or 0) for f in unit_fields), 2)
            item['per_total'] = round(sum(float(item.get(f) or 0) for f in per_fields), 2)
            item['grand_total'] = round(item['unit_total'] + item['per_total'], 2)
            for k in amount_fields:
                v = float(item.get(k) or 0)
                total[k] += v
                item[k] = v
            total['base'] += float(item.get('base') or 0)
            total['fund_base'] += float(item.get('fund_base') or 0)
            items.append(item)

        total['unit_total'] = round(sum(total[f] for f in unit_fields), 2)
        total['per_total'] = round(sum(total[f] for f in per_fields), 2)
        total['grand_total'] = round(total['unit_total'] + total['per_total'], 2)
        return jsonify({'items': items, 'total': total, 'count': len(items)})

    @app.route('/api/insurance/summary/export')
    @login_required
    def api_insurance_export():
        """导出：按人员+时间段 或 按月份全员。表头含『某年某月五险一金缴费明细表』"""
        period = request.args.get('period', '')
        start = request.args.get('start', '')
        end = request.args.get('end', '')
        emp_id = request.args.get('employee_id', type=int)

        q = InsuranceDetail.query.join(Employee)
        if current_user.role == 'employee':
            q = employee_scope_filter(q)
        if period:
            q = q.filter(InsuranceDetail.period == period)
        else:
            if start:
                q = q.filter(InsuranceDetail.period >= start)
            if end:
                q = q.filter(InsuranceDetail.period <= end)
        if emp_id:
            q = q.filter(InsuranceDetail.employee_id == emp_id)
        rows = q.order_by(InsuranceDetail.period.asc(), Employee.id.asc()).all()

        data = []
        unit_fields = ['pension_emp', 'medical_emp', 'extra_medical_emp',
                       'unemployment_emp', 'injury_emp', 'maternity_emp', 'fund_emp']
        per_fields = ['pension_per', 'medical_per', 'extra_medical_per',
                      'unemployment_per', 'fund_per']
        for d in rows:
            unit_total = round(sum(float(getattr(d, f) or 0) for f in unit_fields), 2)
            per_total = round(sum(float(getattr(d, f) or 0) for f in per_fields), 2)
            data.append({
                '客户单位': d.employee.department or '',
                '期间': d.period,
                '员工姓名': d.employee.name,
                '缴费基数': round(d.base or 0, 2),
                '养老单位缴费金额': round(d.pension_emp or 0, 2),
                '养老个人缴费金额': round(d.pension_per or 0, 2),
                '工伤单位缴费金额': round(d.injury_emp or 0, 2),
                '医疗单位缴费金额': round(d.medical_emp or 0, 2),
                '医疗个人缴费金额': round(d.medical_per or 0, 2),
                '大额医疗补助单位缴费金额': round(d.extra_medical_emp or 0, 2),
                '大额医疗补助个人缴费金额': round(d.extra_medical_per or 0, 2),
                '生育单位缴费金额': round(d.maternity_emp or 0, 2),
                '失业单位缴费金额': round(d.unemployment_emp or 0, 2),
                '失业个人缴费金额': round(d.unemployment_per or 0, 2),
                '公积金缴费基数': round(d.fund_base or 0, 2),
                '缴费比例': round(d.fund_rate or 0, 2),
                '公积金单位缴费金额': round(d.fund_emp or 0, 2),
                '公积金个人缴费金额': round(d.fund_per or 0, 2),
                '单位缴费合计': unit_total,
                '个人缴费合计': per_total,
                '总计': round(unit_total + per_total, 2),
                '备注': leave_remark(d.employee, d.remark or ''),
            })

        # 生成表头标题（某年某月五险一金缴费明细表）
        if period and '-' in period:
            y, m = period.split('-')
            title = f'{y}年{int(m)}月五险一金缴费明细表'
        else:
            y1 = (start or '')[:4] or (period or '')[:4] or ''
            if start and end:
                title = f'{start}至{end}五险一金缴费明细表'
            else:
                title = f'{y1}年五险一金缴费明细表' if y1 else '五险一金缴费明细表'

        buf = io.BytesIO()
        # 合计行：数值列求和（备注等文本列留空，第一列标注"合计"）
        num_cols = ['缴费基数', '养老单位缴费金额', '养老个人缴费金额', '工伤单位缴费金额',
                    '医疗单位缴费金额', '医疗个人缴费金额', '大额医疗补助单位缴费金额',
                    '大额医疗补助个人缴费金额', '生育单位缴费金额', '失业单位缴费金额',
                    '失业个人缴费金额', '公积金缴费基数', '公积金单位缴费金额',
                    '公积金个人缴费金额', '单位缴费合计', '个人缴费合计', '总计']
        if data:
            total_row = {k: '' for k in data[0].keys()}
            total_row['客户单位'] = '合计'
            for col in num_cols:
                total_row[col] = round(sum(float(d.get(col) or 0) for d in data), 2)
            data.append(total_row)
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df = pd.DataFrame(data)
            if not data:
                df = pd.DataFrame([{k: '' for k in [
                    '客户单位', '期间', '员工姓名', '缴费基数', '养老单位缴费金额', '养老个人缴费金额',
                    '工伤单位缴费金额', '医疗单位缴费金额', '医疗个人缴费金额',
                    '大额医疗补助单位缴费金额', '大额医疗补助个人缴费金额',
                    '生育单位缴费金额', '失业单位缴费金额', '失业个人缴费金额',
                    '公积金缴费基数', '缴费比例', '公积金单位缴费金额', '公积金个人缴费金额',
                    '单位缴费合计', '个人缴费合计', '总计', '备注']}])
            df.to_excel(writer, index=False, sheet_name='五险一金缴费明细')
            ws = writer.sheets['五险一金缴费明细']
            # 在数据上方插入表头标题行（合并 A1:V1）
            ws.insert_rows(1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = title_cell.font.copy(size=14, bold=True)
            title_cell.alignment = title_cell.alignment.copy(horizontal='center')
            ws.row_dimensions[1].height = 24
            # 合计行加粗
            if data:
                last_row = ws.max_row
                for cell in ws[last_row]:
                    cell.font = cell.font.copy(bold=True)

        buf.seek(0)
        fname = f'{title}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/insurance/template')
    @login_required
    def api_insurance_template():
        """下载导入模板（序号/员工姓名/身份证号/缴费基数/缴费比例/应缴费额/人员编号）"""
        cols = ['序号', '员工姓名', '身份证号', '缴费基数', '缴费比例', '应缴费额', '人员编号']
        sample = pd.DataFrame(
            [[1, '张三', '110101199001011234', 10000, 16, 1600, 'P001']],
            columns=cols
        )
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as w:
            sample.to_excel(w, index=False, sheet_name='社保明细导入模板')
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='社保明细导入模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ===================== 薪酬绩效管理 =====================
    # 工资汇总明细表 列定义（key 对应 SalaryRecord 字段；前端可按需自定义显示）
    SALARY_COLUMNS = [
        {'key': 'seq', 'label': '序号', 'type': 'str', 'default': True},
        {'key': 'department', 'label': '客户单位', 'type': 'str', 'default': True},
        {'key': 'period', 'label': '期间', 'type': 'str', 'default': True},
        {'key': 'name', 'label': '员工姓名', 'type': 'str', 'default': True},
        {'key': 'id_card', 'label': '身份证号', 'type': 'str', 'default': True},
        {'key': 'base_salary', 'label': '基本工资', 'type': 'num', 'default': True},
        {'key': 'post_salary', 'label': '岗位工资', 'type': 'num', 'default': True},
        {'key': 'performance_salary', 'label': '绩效工资', 'type': 'num', 'default': True},
        {'key': 'allowance', 'label': '津贴补贴', 'type': 'num', 'default': True},
        {'key': 'overtime_pay', 'label': '加班工资', 'type': 'num', 'default': True},
        {'key': 'bonus', 'label': '奖金', 'type': 'num', 'default': True},
        {'key': 'should_pay', 'label': '应发工资', 'type': 'num', 'default': True},
        {'key': 'social_personal', 'label': '社保个人', 'type': 'num', 'default': True},
        {'key': 'fund_personal', 'label': '公积金个人', 'type': 'num', 'default': True},
        {'key': 'tax', 'label': '个税', 'type': 'num', 'default': True},
        {'key': 'deduct_total', 'label': '扣款合计', 'type': 'num', 'default': True},
        {'key': 'net_pay', 'label': '实发工资', 'type': 'num', 'default': True},
        {'key': 'remark', 'label': '备注', 'type': 'str', 'default': True},
    ]
    SALARY_NUM_FIELDS = [c['key'] for c in SALARY_COLUMNS if c['type'] == 'num']

    def salary_visible_columns():
        """当前设置的可见列（持久化于 Setting 'salary_columns'），默认全显示"""
        raw = Setting.get('salary_columns', '')
        if raw:
            try:
                vis = json.loads(raw)
                if isinstance(vis, list) and vis:
                    valid = {c['key'] for c in SALARY_COLUMNS}
                    return [k for k in vis if k in valid]
            except Exception:
                pass
        return [c['key'] for c in SALARY_COLUMNS if c['default']]

    def serialize_salary(rec, department='-'):
        d = {
            'id': rec.id,
            'employee_id': rec.employee_id,
            'matched': bool(rec.employee_id),
            'period': rec.period,
            'name': rec.name or '',
            'id_card': rec.id_card or '',
            # 客户单位优先取自匹配员工的「所在单位」，未匹配员工回退导入/工资表表头值（可手动维护）
            'department': (rec.client_unit or '').strip() or department or '-',
            'client_unit': (rec.client_unit or '').strip(),
            'remark': rec.remark or '',
        }
        for f in SALARY_NUM_FIELDS:
            d[f] = round(float(getattr(rec, f) or 0), 2)
        d['tax'] = abs(d['tax'])  # 个税统一为正（扣款金额），兼容历史负值数据
        # 原始导入/校验表格内容（按原表列名与逐行值展示，不再套用固定模板）
        d['headers'] = _safe_json_load(rec.headers_json, [])
        d['values'] = _safe_json_load(rec.values_json, {})
        return d

    @app.route('/salary')
    @login_required
    def salary_detail_page():
        return render_template('salary.html')

    @app.route('/api/salary/records')
    @login_required
    def api_salary_list():
        period = request.args.get('period', '').strip()
        keyword = request.args.get('keyword', '').strip()
        # 多客户单位筛选：units 为逗号分隔列表；兼容旧 unit 单值参数
        units_param = request.args.get('units', '').strip()
        unit = request.args.get('unit', '').strip()
        unit_list = [u for u in units_param.split(',') if u] if units_param else ([unit] if unit else [])
        try:
            page = int(request.args.get('page', 1))
            size = int(request.args.get('size', 10))
        except (ValueError, TypeError):
            page, size = 1, 10
        if size <= 0:
            size = 10
        q = SalaryRecord.query
        if period:
            q = q.filter(SalaryRecord.period == period)
        if unit_list:
            q = q.filter(SalaryRecord.client_unit.in_(unit_list))
        if keyword:
            like = f'%{keyword}%'
            q = q.filter(or_(SalaryRecord.name.like(like), SalaryRecord.id_card.like(like)))
        total = q.count()
        # 按客户单位（空置排末尾）+ 期间升序 + 姓名升序 + id 升序，分组连续、人员正序
        rows = q.order_by(
            func.coalesce(SalaryRecord.client_unit, '') == '',
            func.coalesce(SalaryRecord.client_unit, ''),
            SalaryRecord.period.asc(), SalaryRecord.name.asc(), SalaryRecord.id.asc()
        ).limit(size).offset((page - 1) * size).all()
        emp_cache = {}
        def dept_of(eid):
            if not eid:
                return '-'
            if eid not in emp_cache:
                e = db.session.get(Employee, eid)
                emp_cache[eid] = e.department if e else '-'
            return emp_cache[eid]
        items = [serialize_salary(r, dept_of(r.employee_id)) for r in rows]
        # 客户单位下拉选项：关联代码字段「客户单位」分类，并合并历史单位（保证可筛选）
        units = client_unit_options()
        periods = [r[0] for r in db.session.query(SalaryRecord.period).distinct().order_by(SalaryRecord.period.desc()).all()]
        default_period = periods[0] if periods else ''
        # 默认客户单位：在最近月份中，按选项排序第一个「有数据」的单位；无则取选项第一个
        default_unit = ''
        if units and default_period:
            have = {u[0] for u in db.session.query(SalaryRecord.client_unit)
                    .filter(SalaryRecord.period == default_period,
                            SalaryRecord.client_unit.isnot(None), SalaryRecord.client_unit != '')
                    .distinct().all()}
            for u in units:
                if u in have:
                    default_unit = u
                    break
        if not default_unit and units:
            default_unit = units[0]
        return jsonify({'items': items, 'total': total, 'page': page, 'size': size,
                        'units': units, 'periods': periods,
                        'default_period': default_period, 'default_unit': default_unit,
                        'columns': SALARY_COLUMNS, 'visible': salary_visible_columns()})

    @app.route('/api/salary/periods')
    @login_required
    def api_salary_periods():
        rows = db.session.query(SalaryRecord.period).distinct().order_by(SalaryRecord.period.desc()).all()
        return jsonify({'items': [r[0] for r in rows]})

    @app.route('/api/salary/columns', methods=['GET', 'POST'])
    @login_required
    def api_salary_columns():
        if request.method == 'POST':
            if not can_manage_module('salary'):
                return jsonify({'error': '权限不足'}), 403
            data = request.get_json(silent=True) or {}
            vis = data.get('visible') or []
            valid = {c['key'] for c in SALARY_COLUMNS}
            vis = [k for k in vis if k in valid]
            Setting.set('salary_columns', json.dumps(vis))
            db.session.commit()
            return jsonify({'visible': vis})
        return jsonify({'columns': SALARY_COLUMNS, 'visible': salary_visible_columns()})

    @app.route('/api/salary/template')
    @login_required
    def api_salary_template():
        cols = ['序号', '客户单位', '员工姓名', '身份证号', '期间', '基本工资', '岗位工资', '绩效工资',
                '津贴补贴', '加班工资', '奖金', '应发工资', '社保个人', '公积金个人', '个税',
                '扣款合计', '实发工资', '备注']
        sample = pd.DataFrame([{
            '序号': 1, '客户单位': '示例科技有限公司', '员工姓名': '张三', '身份证号': '110101199001011234', '期间': '2026-08',
            '基本工资': 8000, '岗位工资': 2000, '绩效工资': 3000, '津贴补贴': 500,
            '加班工资': 300, '奖金': 1000, '应发工资': 14800, '社保个人': 800,
            '公积金个人': 500, '个税': 200, '扣款合计': 1500, '实发工资': 13300, '备注': ''
        }], columns=cols)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as w:
            sample.to_excel(w, index=False, sheet_name='工资汇总明细导入模板')
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='工资汇总明细导入模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def _import_salary_rows(rows_data):
        """把 [{name,id_card,period,remark,数值字段...}] 写入工资明细：
        按 身份证+期间 去重（存在则更新、否则新增）；身份证匹配员工表（未匹配不阻断）。
        返回 (success, updated, unmatched)。"""
        emp_cache = {}
        mask_cache = {}
        success = 0
        updated = 0
        unmatched = []
        for rd in rows_data:
            raw_ic = str(rd.get('id_card') or '').strip()
            emp = None
            if is_masked_id_card(raw_ic):
                # 脱敏身份证（如 532224********1125）：按「姓名 + 未脱敏部分」匹配员工
                emp, reason = match_employee_by_masked(rd['name'], raw_ic, mask_cache)
                if not emp:
                    unmatched.append({'name': rd['name'], 'id_card': raw_ic,
                                      'period': rd['period'], 'reason': reason})
                    continue
                rd['id_card'] = emp.id_card  # 归一化为库中真实身份证号后入账
            else:
                # 身份证格式校验：非法格式不写入，列为未匹配，避免脏数据入账
                if not valid_id_card(raw_ic):
                    unmatched.append({'name': rd['name'], 'id_card': raw_ic,
                                      'period': rd['period'], 'reason': '身份证号格式不正确（应为15或18位）'})
                    continue
                emp = emp_cache.get(raw_ic)
                if emp is None:
                    emp = Employee.query.filter_by(id_card=raw_ic).first()
                    emp_cache[raw_ic] = emp
            rec = SalaryRecord.query.filter_by(id_card=rd['id_card'], period=rd['period']).first()
            if rec is None:
                rec = SalaryRecord(id_card=rd['id_card'], period=rd['period'])
                db.session.add(rec)
                success += 1
            else:
                updated += 1
            rec.employee_id = emp.id if emp else None
            rec.name = rd['name']
            # 客户单位：优先取匹配员工「所在单位」（员工台账 department），未匹配员工回退导入/工资表表头值
            rec.client_unit = (emp.department.strip() if (emp and emp.department) else (rd.get('department') or '').strip())
            rec.remark = rd.get('remark', '')
            for fld in SALARY_NUM_FIELDS:
                setattr(rec, fld, rd.get(fld, 0.0) or 0.0)
            # 原始导入/校验表格内容（按原表列名与逐行值存储，供查询页按原表展示）
            rec.headers_json = json.dumps(rd.get('headers') or [], ensure_ascii=False)
            rec.values_json = json.dumps(rd.get('values') or {}, ensure_ascii=False)
            if emp is None:
                unmatched.append({'name': rd['name'], 'id_card': rd['id_card'], 'period': rd['period']})
        db.session.commit()
        return success, updated, unmatched

    @app.route('/api/salary/import', methods=['POST'])
    @role_required('admin', 'hr')
    def api_salary_import():
        f = request.files.get('file')
        if not f:
            return jsonify({'error': '未上传文件'}), 400
        default_period = (request.form.get('period') or '').strip()
        name = f.filename.lower()
        try:
            if name.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(f.read()), dtype=str)
            else:
                df = pd.read_excel(io.BytesIO(f.read()), dtype=str)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {e}'}), 400
        df.columns = [str(c).strip() for c in df.columns]
        orig_cols = list(df.columns)
        colmap = {
            '序号': 'seq', '员工姓名': 'name', '姓名': 'name',
            '身份证号': 'id_card', '身份证': 'id_card',
            '期间': 'period', '月份': 'period',
            '基本工资': 'base_salary', '岗位工资': 'post_salary',
            '绩效工资': 'performance_salary', '津贴补贴': 'allowance',
            '加班工资': 'overtime_pay', '奖金': 'bonus', '应发工资': 'should_pay',
            '社保个人': 'social_personal', '公积金个人': 'fund_personal',
            '个税': 'tax', '扣款合计': 'deduct_total', '实发工资': 'net_pay',
            '备注': 'remark', '客户单位': 'department', '部门': 'department',
        }
        # 每个标准字段对应的「首选原始列名」（避免重复映射；保留导入表原始列名）
        std_col = {}
        for c in orig_cols:
            k = colmap.get(c)
            if k and k not in std_col:
                std_col[k] = c
        if 'name' not in std_col or 'id_card' not in std_col:
            return jsonify({'error': '缺少必需列：员工姓名/身份证号'}), 400

        def to_float(v):
            try:
                f = float(str(v).replace(',', '').replace('，', '').strip() or 0)
                if f != f:  # NaN 安全兜底
                    return 0.0
                return round(f, 2)
            except (ValueError, TypeError):
                return 0.0

        def gv(row, key):
            col = std_col.get(key)
            return str(row.get(col) or '').strip() if col else ''

        rows_data = []
        for _, row in df.iterrows():
            nm = gv(row, 'name')
            ic = gv(row, 'id_card')
            if not nm or not ic:
                continue
            per = gv(row, 'period') or default_period or ''
            if not per:
                continue
            # 原始表格内容：逐列名 -> 逐行值（按导入表实际内容存储，供查询页原样展示）
            values = {}
            for c in orig_cols:
                v = row.get(c)
                values[c] = '' if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)
            rows_data.append({
                'name': nm, 'id_card': ic, 'period': per,
                'base_salary': to_float(gv(row, 'base_salary')),
                'post_salary': to_float(gv(row, 'post_salary')),
                'performance_salary': to_float(gv(row, 'performance_salary')),
                'allowance': to_float(gv(row, 'allowance')),
                'overtime_pay': to_float(gv(row, 'overtime_pay')),
                'bonus': to_float(gv(row, 'bonus')),
                'should_pay': to_float(gv(row, 'should_pay')),
                'social_personal': to_float(gv(row, 'social_personal')),
                'fund_personal': to_float(gv(row, 'fund_personal')),
                'tax': abs(to_float(gv(row, 'tax'))),   # 个税统一为正（扣款金额）
                'deduct_total': to_float(gv(row, 'deduct_total')),
                'net_pay': to_float(gv(row, 'net_pay')),
                'remark': gv(row, 'remark'),
                'department': gv(row, 'department'),
                'headers': orig_cols,
                'values': values,
            })
        if not rows_data:
            return jsonify({'error': '文件中没有有效数据行'}), 400

        success, updated, unmatched = _import_salary_rows(rows_data)
        log_operation('薪酬绩效', '导入', f'{default_period or "多期间"}',
                      f'新增 {success} 条，更新 {updated} 条，未匹配员工 {len(unmatched)} 条')
        return jsonify({'msg': f'导入完成：新增 {success} 条，更新 {updated} 条',
                        'success': success, 'updated': updated, 'unmatched': unmatched})

    def _extract_client_unit(grid, hf_lines):
        """从表头表脚文本 / 网格中提取客户单位（单位名称：XXX）。"""
        candidates = []
        if hf_lines:
            for line in hf_lines:
                if isinstance(line, list):
                    candidates.extend(str(c) for c in line)
                else:
                    candidates.append(str(line))
        for row in (grid or []):
            if isinstance(row, list):
                candidates.extend(str(c) for c in row)
        pat = re.compile(r'(?:客户单位|单位名称)\s*[:：]\s*(.+)')
        for s in candidates:
            m = pat.search(str(s))
            if m:
                v = m.group(1).strip()
                v = re.sub(r'^[A-Za-z]\s*[:：]\s*', '', v).strip()
                if v and v not in ('单位名称', '客户单位', ''):
                    return v
        return ''

    @app.route('/api/payslip/to-salary', methods=['POST'])
    @role_required('admin', 'hr')
    def api_payslip_to_salary():
        """工资汇总制作 → 一键生成工资明细：把校验后的工资表正文区写入 SalaryRecord，
        个税管理等基于工资明细的功能随之有数据。按表头关键字识别列；期间可来自网格列或请求参数；
        客户单位（单位名称）从表头表脚配置或网格中提取。"""
        payload = request.get_json(silent=True) or {}
        sheets = payload.get('sheets')
        period = (payload.get('period') or '').strip()
        hf = payload.get('headfoot') or {}
        per_sheet_hf = (hf.get('per_sheet') or {}) if isinstance(hf, dict) else {}
        if not sheets or not isinstance(sheets, list):
            return jsonify({'error': '缺少表格数据'}), 400
        if not re.match(r'^\d{4}-\d{1,2}$', period):
            return jsonify({'error': '请填写期间（YYYY-MM，如 2026-08）'}), 400
        rows_data = []
        skipped = []
        for sh in sheets:
            grid = sh.get('grid') or []
            if not grid:
                continue
            # 客户单位：优先从本表表头表脚配置的单位名称取，回退扫描网格
            sh_hf = per_sheet_hf.get(str(sh.get('name') or '')) or {}
            hf_lines = (sh_hf.get('header_after') or []) + (sh_hf.get('footer_after') or [])
            unit = _extract_client_unit(grid, hf_lines)
            hidx = _detect_body_start(grid)
            if hidx >= len(grid):
                skipped.append(str(sh.get('name') or ''))
                continue
            hdr = grid[hidx]

            def _fc(*kws):
                for j, h in enumerate(hdr):
                    if any(k in str(h) for k in kws):
                        return j
                return None

            c_name = _fc('姓名')
            c_id = _fc('身份证')
            c_period = _fc('期间', '月份')
            c_should = _fc('应发')
            c_social = _fc('社保')
            c_fund = _fc('公积金')
            c_tax = _fc('个税', '所得税')
            c_deduct = _fc('扣款')
            c_net = _fc('实发')
            c_remark = _fc('备注')
            if c_name is None or c_id is None:
                skipped.append(str(sh.get('name') or ''))
                continue
            body = _cut_original_footer(grid[hidx + 1:])
            while body and all(c in (None, '') for c in body[-1]):
                body.pop()
            for row in body:
                if not row or all(c in (None, '') for c in row):
                    continue

                def _get(cidx):
                    return str(row[cidx]).strip() if (cidx is not None and cidx < len(row)) else ''

                nm = _get(c_name)
                ic = _get(c_id)
                if not nm or not ic:
                    continue
                per = _get(c_period) or period
                if not re.match(r'^\d{4}-\d{1,2}$', per):
                    continue
                # 原始表格内容：用校验后网格的表头与各列实际值，供查询页按原表展示
                _hdr = [str(h).strip() for h in hdr]
                _vals = {}
                for _j, _h in enumerate(_hdr):
                    if _j < len(row):
                        _v = row[_j]
                        _vals[_h] = '' if _v is None else str(_v)
                rows_data.append({
                    'name': nm, 'id_card': ic, 'period': per,
                    'should_pay': _payslip_to_float(_get(c_should)) or 0.0,
                    'social_personal': _payslip_to_float(_get(c_social)) or 0.0,
                    'fund_personal': _payslip_to_float(_get(c_fund)) or 0.0,
                    'tax': abs(_payslip_to_float(_get(c_tax)) or 0.0),
                    'deduct_total': _payslip_to_float(_get(c_deduct)) or 0.0,
                    'net_pay': _payslip_to_float(_get(c_net)) or 0.0,
                    'remark': _get(c_remark),
                    'department': unit,
                    'headers': _hdr,
                    'values': _vals,
                })
        if not rows_data:
            return jsonify({'error': '未从表格正文识别到有效数据行（需包含「姓名/身份证号」列）'}), 400
        success, updated, unmatched = _import_salary_rows(rows_data)
        log_operation('薪酬绩效', '工资表生成明细', period,
                      f'新增 {success} 条，更新 {updated} 条，未匹配员工 {len(unmatched)} 条')
        return jsonify({'msg': f'已生成工资明细：新增 {success} 条，更新 {updated} 条',
                        'success': success, 'updated': updated, 'unmatched': unmatched,
                        'skipped': skipped})

    @app.route('/api/salary/export')
    @login_required
    def api_salary_export():
        period = request.args.get('period', '').strip()
        keyword = request.args.get('keyword', '').strip()
        units_param = request.args.get('units', '').strip()
        unit = request.args.get('unit', '').strip()
        unit_list = [u for u in units_param.split(',') if u] if units_param else ([unit] if unit else [])
        q = SalaryRecord.query
        if period:
            q = q.filter(SalaryRecord.period == period)
        if unit_list:
            q = q.filter(SalaryRecord.client_unit.in_(unit_list))
        if keyword:
            like = f'%{keyword}%'
            q = q.filter(or_(SalaryRecord.name.like(like), SalaryRecord.id_card.like(like)))
        rows = q.order_by(SalaryRecord.period.desc(), SalaryRecord.id.desc()).all()
        emp_cache = {}
        def dept_of(eid):
            if not eid:
                return '-'
            if eid not in emp_cache:
                e = db.session.get(Employee, eid)
                emp_cache[eid] = e.department if e else '-'
            return emp_cache[eid]
        # 动态列：取结果集中各记录「原始导入/校验表格」表头的并集（按首次出现顺序）
        all_headers = []
        seen = set()
        for r in rows:
            hs = _safe_json_load(r.headers_json, []) or []
            for h in hs:
                if h not in seen:
                    seen.add(h)
                    all_headers.append(h)
        # 存在回退记录（无原始内容）时，并入结构化字段的标签作为表头
        for r in rows:
            if not (_safe_json_load(r.values_json, {}) or {}):
                for c in SALARY_COLUMNS:
                    if c['key'] == 'seq':
                        continue
                    if c['label'] not in seen:
                        seen.add(c['label'])
                        all_headers.append(c['label'])
                break

        def record_values(r):
            vals = _safe_json_load(r.values_json, {}) or {}
            if vals:
                return vals
            d = serialize_salary(r, dept_of(r.employee_id))
            return {c['label']: d.get(c['key'], '') for c in SALARY_COLUMNS if c['key'] != 'seq'}

        data = []
        for r in rows:
            vals = record_values(r)
            row = {h: vals.get(h, '') for h in all_headers}
            data.append(row)
        cols_label = all_headers
        df = pd.DataFrame(data, columns=cols_label) if data else pd.DataFrame(columns=cols_label)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as w:
            df.to_excel(w, index=False, sheet_name='工资汇总明细表')
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='工资汇总明细表.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/salary/clear', methods=['POST'])
    @role_required('admin', 'hr')
    def api_salary_clear():
        """清空工资明细数据（含关联的个税数据，因个税实时取自工资明细）。"""
        n = SalaryRecord.query.count()
        SalaryRecord.query.delete()
        db.session.commit()
        log_operation('薪酬绩效', '清空工资明细', '', f'清空 {n} 条')
        return jsonify({'msg': f'已清空工资明细 {n} 条', 'deleted': n})

    @app.route('/api/salary/<int:sid>', methods=['DELETE'])
    @role_required('admin', 'hr')
    def api_salary_delete(sid):
        rec = db.session.get(SalaryRecord, sid)
        if not rec:
            return jsonify({'error': '不存在'}), 404
        db.session.delete(rec)
        db.session.commit()
        log_operation('薪酬绩效', '删除', f'{rec.period} {rec.name}', '')
        return jsonify({'msg': '已删除'})

    @app.route('/api/salary/<int:sid>', methods=['PUT'])
    @role_required('admin', 'hr')
    def api_salary_update(sid):
        """手动维护工资明细（未匹配员工的客户单位等）：可改姓名/身份证/期间/客户单位/备注，
        并按身份证重新匹配员工；匹配到的记录客户单位自动对齐员工「所在单位」。"""
        rec = db.session.get(SalaryRecord, sid)
        if not rec:
            return jsonify({'error': '不存在'}), 404
        data = request.get_json(force=True, silent=True) or {}
        rec.name = (data.get('name') or '').strip() or rec.name
        rec.id_card = (data.get('id_card') or '').strip() or rec.id_card
        rec.period = (data.get('period') or '').strip() or rec.period
        if 'remark' in data:
            rec.remark = data.get('remark') or ''
        # 按身份证重新匹配员工
        emp = Employee.query.filter_by(id_card=rec.id_card).first() if rec.id_card else None
        rec.employee_id = emp.id if emp else None
        # 匹配到的记录：客户单位对齐员工「所在单位」；未匹配（无员工）：保留手动维护值
        if emp and emp.department:
            rec.client_unit = emp.department.strip()
        else:
            rec.client_unit = (data.get('client_unit') or '').strip()
        # 同步原始表格内容（若导入表含对应列）
        try:
            vals = _safe_json_load(rec.values_json, {}) or {}
            if vals:
                if '客户单位' in vals:
                    vals['客户单位'] = rec.client_unit
                if '员工姓名' in vals:
                    vals['员工姓名'] = rec.name
                elif '姓名' in vals:
                    vals['姓名'] = rec.name
                if '身份证号' in vals:
                    vals['身份证号'] = rec.id_card
                elif '身份证' in vals:
                    vals['身份证'] = rec.id_card
                rec.values_json = json.dumps(vals, ensure_ascii=False)
        except Exception:
            pass
        db.session.commit()
        log_operation('薪酬绩效', '编辑明细', f'{rec.period} {rec.name}', f'客户单位={rec.client_unit}')
        return jsonify({'msg': '已更新', 'matched': bool(rec.employee_id)})

    @app.route('/api/salary/batch-delete', methods=['POST'])
    @role_required('admin', 'hr')
    def api_salary_batch_delete():
        """批量删除：按期间 + 多客户单位（均可选）。两个筛选均为空时删除全部。"""
        data = request.get_json(force=True, silent=True) or {}
        period = (data.get('period') or '').strip()
        units = [u for u in (data.get('units') or []) if u]
        q = SalaryRecord.query
        if period:
            q = q.filter(SalaryRecord.period == period)
        if units:
            q = q.filter(SalaryRecord.client_unit.in_(units))
        n = q.count()
        q.delete()
        db.session.commit()
        scope = (period or '全部期间') + ' / ' + (('、'.join(units)) if units else '全部客户单位')
        log_operation('薪酬绩效', '批量删除工资明细', scope, f'删除 {n} 条')
        return jsonify({'msg': f'已删除 {n} 条', 'deleted': n})

    @app.route('/api/salary/rematch-units', methods=['POST'])
    @role_required('admin', 'hr')
    def api_salary_rematch():
        """按员工信息匹配客户单位：将已匹配员工的工资明细客户单位，按员工「所在单位」回填；
        未匹配记录不动，由用户手动维护。"""
        n = 0
        recs = SalaryRecord.query.filter(SalaryRecord.employee_id.isnot(None)).all()
        for r in recs:
            emp = db.session.get(Employee, r.employee_id)
            if emp and emp.department and emp.department.strip() != (r.client_unit or '').strip():
                r.client_unit = emp.department.strip()
                n += 1
        db.session.commit()
        log_operation('薪酬绩效', '按员工信息匹配客户单位', '', f'更新 {n} 条')
        return jsonify({'msg': f'已按员工所在单位更新 {n} 条', 'updated': n})

    # ===================== 工资表汇总（导入校验 / 自定义生成） =====================
    def _payslip_to_float(v):
        try:
            return float(str(v).replace(',', '').replace('，', '').strip())
        except (ValueError, TypeError):
            return None

    # 日期字符串（含 datetime 转 str 带来的 00:00:00 尾部）匹配
    _DATE_RE = re.compile(r'^(\d{4}-\d{2}-\d{2})( \d{2}:\d{2}:\d{2})?$')
    # 身份证号特征：15 位，或 17 位 + 校验位（数字或 X）
    _ID_RE = re.compile(r'^\d{15}$|^\d{17}[\dXx]$')

    def _looks_like_id(v):
        return isinstance(v, str) and bool(_ID_RE.match(v.strip()))

    def _payslip_cell(v):
        if v is None:
            return ''
        if isinstance(v, bool):
            return str(v)
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        if isinstance(v, (int, float)):
            return str(v)
        return str(v).strip()

    def _pad_grid(rows):
        """统一所有行的列宽（补空），返回 list[list[str]]"""
        maxc = max((len(r) for r in rows), default=0)
        for r in rows:
            while len(r) < maxc:
                r.append('')
        return rows

    def _parse_payslip_grid(file_bytes, name):
        """解析 Excel/CSV 为 sheets（list[{name, grid, style}]），统一列宽。
        Excel 含多个 sheet 时全部返回；CSV 视为单 sheet。
        style：从源表捕获的格式（列宽 / 行高 / 逐格字体·边框·对齐·数字格式），导出时正文区复用。"""
        from openpyxl.utils import get_column_letter, column_index_from_string
        if name.endswith('.csv'):
            text = file_bytes.decode('utf-8-sig', errors='ignore')
            rows = [[_payslip_cell(c) for c in r] for r in _csv.reader(io.StringIO(text))]
            grid = _pad_grid(rows)
            return [{'name': '工资表', 'grid': grid, 'style': None}]
        else:
            # 不使用 read_only，以便读取单元格字体/边框/对齐/数字格式等样式
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            # 工作簿默认字体（多数中文表的默认字体，如宋体），用于无显式字体单元格的兜底
            try:
                _df = wb._fonts[0]
                default_font = {'name': _df.name or '宋体', 'size': _df.size or 11}
            except Exception:
                default_font = {'name': '宋体', 'size': 11}
            sheets = []
            for ws in wb.worksheets:
                max_r = ws.max_row or 0
                max_c = ws.max_column or 0
                if max_r == 0 or max_c == 0:
                    sheets.append({'name': ws.title or ('sheet%d' % (len(sheets) + 1)),
                                   'grid': [[]], 'style': None})
                    continue
                max_c = min(max_c, 300)  # 防止维度异常导致超大循环
                rows = []
                style = {'widths': {}, 'heights': {}, 'cells': {}, 'hidden_cols': set(), 'default_font': default_font}
                # 工作表级默认行高/列宽：Excel 中“全表统一设置”时只存于此，逐行逐列并无显式值
                style['default_row_height'] = getattr(ws.sheet_format, 'defaultRowHeight', None)
                style['default_col_width'] = getattr(ws.sheet_format, 'defaultColWidth', None)
                # 打印/页面设置：纸张方向、纸型、缩放、边距（导出后打印与导入表保持一致）
                _ps = ws.page_setup
                _fp = getattr(ws.sheet_properties, 'pageSetUpPr', None)
                style['page_setup'] = {
                    'orientation': getattr(_ps, 'orientation', None),
                    'paperSize': getattr(_ps, 'paperSize', None),
                    'scale': getattr(_ps, 'scale', None),
                    'fitToWidth': getattr(_ps, 'fitToWidth', None),
                    'fitToHeight': getattr(_ps, 'fitToHeight', None),
                    'fitToPage': bool(getattr(_fp, 'fitToPage', False)) if _fp is not None else None,
                }
                _pm = ws.page_margins
                if _pm is not None:
                    style['page_margins'] = {
                        'left': _pm.left, 'right': _pm.right, 'top': _pm.top,
                        'bottom': _pm.bottom, 'header': _pm.header, 'footer': _pm.footer,
                    }
                else:
                    style['page_margins'] = None
                # 列宽/隐藏列（支持范围维度 <col min max>：Excel 会把相邻同属性列合并成范围写，
                # openpyxl 只保留 min 列的维度 key，必须用 dim.min/dim.max 展开，否则范围里的列会漏掉）
                for k, dim in ws.column_dimensions.items():
                    try:
                        _base = column_index_from_string(k)  # 1-based
                    except Exception:
                        _base = 1
                    lo = dim.min if dim.min else _base
                    hi = dim.max if dim.max else lo
                    lo, hi = max(1, lo), max(lo, hi)
                    for ci in range(lo, hi + 1):
                        idx = ci - 1
                        if dim.width is not None:
                            style['widths'][idx] = dim.width
                        # 隐藏列：标准 hidden 标记，或“列宽被拖成 ≈0”的伪隐藏（部分工具用 0 宽藏列）
                        if dim.hidden or (dim.width is not None and dim.width < 0.5):
                            style['hidden_cols'].add(idx)
                # 行高（仅捕获显式设置过行高的行）
                for k, dim in ws.row_dimensions.items():
                    if dim.height is not None:
                        style['heights'][k - 1] = dim.height
                for r in range(1, max_r + 1):
                    row_vals = []
                    for c in range(1, max_c + 1):
                        cell = ws.cell(row=r, column=c)
                        row_vals.append(_payslip_cell(cell.value))
                        f = cell.font
                        b = cell.border
                        # openpyxl 默认每个单元格的四边都是 Side(style=None) 实例（非 None），
                        # 因此必须用 side.style 是否非 None 来判断“是否有实际边框”，否则 has_border 恒为 False。
                        def _side_drawn(sd):
                            return sd is not None and getattr(sd, 'style', None) is not None
                        has_border = any(_side_drawn(getattr(b, s)) for s in ('left', 'right', 'top', 'bottom'))
                        al = cell.alignment
                        nf = cell.number_format
                        al_h = al.horizontal if al else None
                        wrap = bool(al and al.wrap_text)
                        # 仅捕获与默认值不同的样式，减小体积
                        nondefault = (
                            (f and (f.name not in (None, 'Calibri', '宋体', '') or
                                    f.size not in (None, 11) or f.bold)) or
                            has_border or
                            (nf not in (None, 'General', 'general', '@')) or
                            (al_h not in (None, 'center')) or wrap
                        )
                        if nondefault:
                            style['cells'][(r - 1, c - 1)] = {
                                'name': f.name or '宋体',
                                'size': f.size or 11,
                                'bold': bool(f.bold),
                                'border': has_border,
                                'align': al_h or 'center',
                                'wrap': wrap,
                                'num_fmt': nf or 'General',
                            }
                    rows.append(row_vals)
                grid = _pad_grid(rows)
                sheets.append({'name': ws.title or ('sheet%d' % (len(sheets) + 1)),
                               'grid': grid, 'style': style})
            return sheets

    _TOTAL_KEYWORDS = ['合计', '总计', '小计', '实发合计', '应发合计', '合 计', 'total', 'sum']

    def _detect_payslip(grid):
        """自动识别序号列、金额列、合计行"""
        n = len(grid)
        m = max((len(r) for r in grid), default=0)
        seq_col = None
        for j in range(m):
            ints = []
            for i in range(n):
                f = _payslip_to_float(grid[i][j])
                if f is not None and f == int(f) and abs(f) < 1e9:
                    ints.append(int(f))
            if len(ints) >= 2:
                s = set(ints)
                if s == set(range(min(s), max(s) + 1)) and len(s) >= max(2, int(0.5 * n)):
                    seq_col = j
                    break
        amount_cols = []
        for j in range(m):
            if j == seq_col:
                continue
            cnt = 0
            total = 0
            for i in range(n):
                v = grid[i][j]
                if v == '':
                    continue
                total += 1
                if _payslip_to_float(v) is not None:
                    cnt += 1
            if total >= 2 and cnt / total >= 0.6:
                amount_cols.append(j)
        total_row = None
        # 仅当存在连续序号列（规范表格）时，才将含“合计”字样的行视为待校验合计行；
        # 通知单/增减表等非表格文本表不识别合计行，避免误校验。
        if seq_col is not None:
            for i in range(n):
                head = ''
                for j in range(m):
                    if grid[i][j] != '':
                        head = grid[i][j]
                        break
                if head and any(kw in head.lower() for kw in _TOTAL_KEYWORDS):
                    total_row = i
        return {'seq_col': seq_col, 'amount_cols': amount_cols, 'total_row': total_row, 'rows': n, 'cols': m}

    def _validate_payslip(grid, det):
        """校验序号连续性、各金额列合计=明细求和"""
        errors = []
        n = len(grid)
        total_row = det['total_row']
        sc = det['seq_col']
        if sc is not None:
            ints = []
            for i in range(n):
                f = _payslip_to_float(grid[i][sc])
                if f is not None and f == int(f):
                    ints.append(int(f))
            if ints:
                s = set(ints)
                k = max(s)
                missing = [x for x in range(1, k + 1) if x not in s]
                dups = [x for x in s if ints.count(x) > 1]
                if missing or dups:
                    errors.append({
                        'type': 'seq', 'col': sc, 'row': -1,
                        'label': '序号列（第%d列）' % (sc + 1),
                        'wrong': '、'.join(str(x) for x in ints),
                        'correct': '、'.join(str(x) for x in range(1, k + 1)),
                        'msg': '序号不连续（缺失：%s%s）' % (
                            '、'.join(map(str, missing)) if missing else '无',
                            '；重复：' + '、'.join(map(str, dups)) if dups else '')
                    })
        # 合计校验仅在“存在连续序号列”的规范表格上执行，避免通知单/增减表等
        # 非表格文本表把“人数/年月”等列误判为金额列而产生假错误。
        if total_row is not None and sc is not None:
            for j in det['amount_cols']:
                total_val = _payslip_to_float(grid[total_row][j])
                if total_val is None:
                    continue
                ssum = 0.0
                ok = False
                for i in range(n):
                    if i == total_row:
                        continue
                    f = _payslip_to_float(grid[i][j])
                    if f is not None:
                        ssum += f
                        ok = True
                if ok and abs(ssum - total_val) > 0.01:
                    fmt = ('%.2f' % ssum).rstrip('0').rstrip('.')
                    errors.append({
                        'type': 'sum', 'col': j, 'row': total_row,
                        'label': '第%d列' % (j + 1),
                        'wrong': grid[total_row][j],
                        'correct': fmt,
                        'msg': '合计应为 %s，当前为 %s' % (fmt, grid[total_row][j])
                    })
        return errors

    # 个税税率表：累计预扣法（年度）与按月扣缴（月度）
    TAX_CUM = [(36000, 0.03, 0), (144000, 0.10, 2520), (300000, 0.20, 16920),
               (420000, 0.25, 31920), (660000, 0.30, 52920), (960000, 0.35, 85920),
               (float('inf'), 0.45, 181920)]
    TAX_MONTH = [(1500, 0.03, 0), (4500, 0.10, 105), (9000, 0.20, 555),
                 (35000, 0.25, 1005), (55000, 0.30, 2755), (80000, 0.35, 5505),
                 (float('inf'), 0.45, 13505)]

    def calc_income_tax(amount, method='cum', threshold=0.0):
        """计算个税：method='cum' 累计预扣法(年度税率表)；'month' 按月扣缴(月度税率表)。
        threshold 为起征点(基本减除费用)，默认 0；应纳税所得额 = max(0, amount - threshold)。"""
        table = TAX_CUM if method == 'cum' else TAX_MONTH
        x = max(0.0, float(amount or 0) - float(threshold or 0))
        for cap, rate, ded in table:
            if x <= cap:
                return max(0.0, round(x * rate - ded, 2))
        return 0.0

    def _dept_of(eid):
        if not eid:
            return '-'
        e = db.session.get(Employee, eid)
        return e.department if e else '-'

    def _parse_headfoot_template(data, name):
        """解析表头/表脚模板：表头表、表脚表（含合并单元格），返回 {header:{grid,merges}, footer:{...}}"""
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        out = {}
        for key, sheet_name in (('header', '表头'), ('footer', '表脚')):
            if sheet_name not in wb.sheetnames:
                out[key] = {'grid': [], 'merges': []}
                continue
            ws = wb[sheet_name]
            maxr, maxc = ws.max_row, ws.max_column
            grid = []
            for i in range(1, maxr + 1):
                row = []
                for j in range(1, maxc + 1):
                    v = ws.cell(row=i, column=j).value
                    row.append('' if v is None else str(v))
                grid.append(row)
            # 去掉末尾全空行
            while grid and all(c == '' for c in grid[-1]):
                grid.pop()
            merges = []
            for mr in getattr(ws, 'merged_cells', None).ranges if hasattr(ws, 'merged_cells') else []:
                merges.append([mr.min_row - 1, mr.min_col - 1, mr.max_row - 1, mr.max_col - 1])
            out[key] = {'grid': grid, 'merges': merges}
        return out

    def _render_hf_region(ws, region, maxc, start_row, border, bold_title):
        """将表头/表脚区域写入工作表（合并单元格钳制到数据宽度，解决溢出问题）"""
        from openpyxl.styles import Alignment, Font
        grid = region.get('grid') or []
        merges = region.get('merges') or []
        nrows = len(grid)
        for i, row in enumerate(grid):
            rr = start_row + i
            for j, v in enumerate(row):
                if j >= maxc:
                    continue
                cell = ws.cell(row=rr, column=j + 1, value=(v if v != '' else None))
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if bold_title:
                    cell.font = Font(bold=True, size=12)
        for (r1, c1, r2, c2) in merges:
            sc = c1 + 1
            ec = min(c2 + 1, maxc)
            sr = r1 + start_row
            er = r2 + start_row
            if sc > maxc or sr > er or sc > ec:
                continue
            try:
                ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)
            except Exception:
                pass
        return start_row + nrows

    def _write_payslip_sheet(ws, grid, header_rows, footer_rows, headfoot):
        """将 grid + 自定义表头/表脚写入一个已存在的工作表；headfoot 为模板(含合并单元格)时优先使用"""
        from openpyxl.styles import Alignment, Font, Border, Side
        maxc = max((len(r) for r in grid), default=1)
        thin = Side(style='thin', color='D0D0D0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        r = 1
        if headfoot and headfoot.get('header'):
            r = _render_hf_region(ws, headfoot['header'], maxc, r, border, bold_title=True)
        else:
            for h in (header_rows or []):
                cell = ws.cell(row=r, column=1, value=h)
                if maxc > 1:
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=maxc)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=12)
                r += 1
        for row in grid:
            for j, v in enumerate(row, 1):
                f = _payslip_to_float(v)
                cell = ws.cell(row=r, column=j, value=(f if f is not None else v))
                cell.border = border
                if f is None:
                    cell.alignment = Alignment(horizontal='center')
            r += 1
        if headfoot and headfoot.get('footer'):
            r = _render_hf_region(ws, headfoot['footer'], maxc, r, border, bold_title=False)
        else:
            for frow in (footer_rows or []):
                cell = ws.cell(row=r, column=1, value=frow)
                if maxc > 1:
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=maxc)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                r += 1

    def _build_payslip_workbook(sheets, header_rows, footer_rows, headfoot=None):
        """将多张表（list[{name, grid}]）构建为一个多 sheet xlsx 字节流；
        headfoot 模板(含合并单元格)对每张表生效。"""
        wb = Workbook()
        wb.remove(wb.active)
        for idx, sh in enumerate(sheets):
            grid = sh.get('grid') or []
            title = (sh.get('name') or ('工资表%d' % (idx + 1)))[:31]
            ws = wb.create_sheet(title=title)
            _write_payslip_sheet(ws, grid, header_rows, footer_rows, headfoot)
        if not wb.sheetnames:
            wb.create_sheet('工资表')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ===== 方式一统一模板：按 sheet 两套表头表脚 + 个税 =====
    def _is_num_str(s):
        try:
            float(str(s).replace(',', '').strip()); return True
        except Exception:
            return False

    def _detect_body_start(grid):
        """定位表格正文（列头行）起始行索引：优先含'序号'的行；否则首行含较多非数字文本的表头行"""
        for i, row in enumerate(grid):
            for c in row:
                if c is not None and '序号' in str(c):
                    return i
        for i, row in enumerate(grid):
            ne = [c for c in row if c not in (None, '')]
            if len(ne) >= 5 and any(not _is_num_str(x) for x in ne):
                return i
        return 0

    def _extract_report_hf(grid):
        """从一张表的网格提取校验后表头/表脚：正文上方的装饰行 -> header_after；底部含签章关键字的连续块 -> footer_after"""
        hidx = _detect_body_start(grid)
        header_after = []
        for i in range(hidx):
            row = [str(c) if c not in (None, '') else '' for c in grid[i]]
            if any(x != '' for x in row):
                header_after.append(row)
        footer_after = []
        kw = ['负责人', '审核人', '制表人', '签收', '制表']
        end_row = None
        for i in range(len(grid) - 1, -1, -1):
            row = [str(c) if c not in (None, '') else '' for c in grid[i]]
            if any(any(k in x for k in kw) for x in row):
                end_row = i
                break
        if end_row is not None:
            # 向上扩展至连续签章块顶部（遇到非签章且非空白的数据行即止）
            start = end_row
            while start - 1 >= 0:
                r = grid[start - 1]
                if any(any(k in str(c) for k in kw) for c in r if c not in (None, '')) or \
                   all(str(c).strip() == '' for c in r):
                    start -= 1
                else:
                    break
            footer_after = [[str(c) if c not in (None, '') else '' for c in grid[i]]
                            for i in range(start, end_row + 1)]
        return {'header_after': header_after, 'footer_after': footer_after}

    def _suggest_tax(grid):
        """若表内含'税额-所得税'与'税前合计'列，返回默认个税配置"""
        hidx = _detect_body_start(grid)
        if hidx >= len(grid):
            return None
        hdr = grid[hidx]
        has_t = any(str(c).strip() == '税额-所得税' for c in hdr if c is not None)
        has_s = any(str(c).strip() == '税前合计' for c in hdr if c is not None)
        if has_t and has_s:
            return {'enabled': True, 'target': '税额-所得税', 'source': '税前合计',
                    'method': 'cum', 'threshold': 5000, 'mode': 'keep'}
        return None

    def _fill_tax(body, tax_cfg):
        """在正文网格上按个税配置填充目标列（仅 after 版本）：
        1) mode='keep'（默认）：导入数据中已有的税额**原样保留**（人工/税务系统填写的精确值优先，
           其中包含专项附加扣除、累计预扣等公式无法复现的因素），仅对空缺行按公式计算兜底；
           mode='calc'：忽略原值，全部按公式重算；
        2) 计算公式：个税基数 = max(0, 税前合计 - 起征点)，起征点缺失或为 0 时默认 5000，
           避免“工资不到5000也计税”；税额按原表惯例以负数表示扣款（0 记为 0，不写 -0.0）；
        3) 同步重算“实发合计” = 税前合计 + 工会会费 + 税额（工会会费/税额均为扣款负数）；
        4) “合计/总计/小计”行不套公式，其税额/实发改为各明细行之和，保证合计一致。"""
        if not tax_cfg or not tax_cfg.get('enabled'):
            return body
        if not body:
            return body
        hdr = body[0]
        tcol, scol = tax_cfg.get('target'), tax_cfg.get('source')
        mode = (tax_cfg.get('mode') or 'keep').lower()  # keep=保留原值(默认) calc=全部重算
        tgt = src = None
        sf = gh = None  # 实发合计、工会会费 列（重算实发用）
        for j, h in enumerate(hdr):
            if h is None:
                continue
            hs = str(h).strip()
            if tcol and hs == str(tcol):
                tgt = j
            if scol and hs == str(scol):
                src = j
            if hs == '实发合计':
                sf = j
            if hs == '工会会费':
                gh = j
        if tgt is None or src is None:
            return body
        method = tax_cfg.get('method', 'cum')
        # 起征点(基本减除费用)：税前合计已扣除五险两金等专项扣除；缺失或显式 0 时默认 5000
        raw_th = tax_cfg.get('threshold')
        try:
            threshold = 5000.0 if (raw_th in (None, '') or float(raw_th) == 0) else float(raw_th)
        except Exception:
            threshold = 5000.0

        def _fmt_tax(x):
            """正数应纳税额 -> 负数扣款字符串；0 记为 '0'"""
            x = round(float(x), 2)
            return '0' if x == 0 else str(-x)

        def _is_total_row(row):
            txts = [str(c) for c in row if c not in (None, '')]
            if not txts:
                return False
            if any(k in t for t in txts for k in ('合计', '总计', '小计')):
                return True
            if not _is_num_str(txts[0]):
                return True
            return False

        total_tax = 0.0   # 累计员工明细税额（已为负）
        total_sf = 0.0    # 累计员工明细实发
        # 第一遍：明细行处理（保留原值或计算）；合计行标记跳过
        for i in range(1, len(body)):
            row = body[i]
            if _is_total_row(row):
                continue
            v = _payslip_to_float(row[src]) if src < len(row) else None
            orig = row[tgt] if tgt < len(row) else None
            orig_num = _payslip_to_float(orig)
            # keep 模式：原值非空则完全保留（人工/税务系统精确值优先，保持原符号），仅空缺才计算
            if mode == 'keep' and orig not in (None, ''):
                if orig_num is not None:
                    row[tgt] = str(round(orig_num, 2))  # 数字则规范化两位小数
            elif v is None:
                # 税前合计为空：无原值则置空，有原值则保留
                if orig in (None, ''):
                    row[tgt] = ''
                continue
            else:
                tax_val = calc_income_tax(v, method, threshold)  # 正数应纳税额
                row[tgt] = _fmt_tax(tax_val)  # 原表以负数表示扣款
            # 重算实发合计 = 税前合计 + 工会会费 + 税额（税前为空则跳过）
            if sf is not None and v is not None:
                pre = _payslip_to_float(row[src]) or 0.0
                gh_v = _payslip_to_float(row[gh]) if (gh is not None and gh < len(row)) else 0.0
                gh_v = gh_v or 0.0
                tv = _payslip_to_float(row[tgt]) if tgt < len(row) else 0.0
                tv = tv or 0.0
                sf_val = round(pre + gh_v + tv, 2)
                row[sf] = str(sf_val)
                total_tax += tv
                total_sf += sf_val
        # 第二遍：填合计行（税额/实发 = 各明细之和，保持符号）
        for i in range(1, len(body)):
            if _is_total_row(body[i]):
                if tgt is not None:
                    body[i][tgt] = str(round(total_tax, 2))
                if sf is not None:
                    body[i][sf] = str(round(total_sf, 2))
        return body

    def _sheet_hf(hf_v2, name):
        if not isinstance(hf_v2, dict):
            return {'header_after': [], 'footer_after': []}
        ps = hf_v2.get('per_sheet') or {}
        cand = ps.get(name) or ps.get('*') or {}
        if cand:
            # 透传整条配置：header_after/footer_after 以及 font/header_align/footer_align
            return {**{'header_after': [], 'footer_after': []}, **cand}
        return {'header_after': hf_v2.get('header_after', []), 'footer_after': hf_v2.get('footer_after', [])}

    def _col_to_idx(letters):
        """将 Excel 列字母（A/B/.../Z/AA...）转为 1 开始的列序号"""
        idx = 0
        for ch in str(letters).upper():
            if 'A' <= ch <= 'Z':
                idx = idx * 26 + (ord(ch) - 64)
        return idx

    def _parse_col_text_cells(line):
        """逐列定位解析：line 为单元格列表；每个非空单元格内按空格拆分多个「列字母:文本」片段，分别落位。
        返回 [(col_1based, text), ...]。无「列字母」前缀的片段按顺序落到后续列。"""
        out = []
        k = 1
        for raw_cell in line:
            if raw_cell is None or raw_cell == '':
                k += 1
                continue
            for tok in re.split(r'\s+', str(raw_cell).strip()):
                if tok == '':
                    continue
                col = k
                text = tok
                m = re.match(r'^([A-Za-z]{1,2})[:：]\s*(.*)$', tok)
                if m and 0 < _col_to_idx(m.group(1)) <= 16384:
                    col = _col_to_idx(m.group(1))
                    text = m.group(2)
                out.append((col, text))
                if not m:
                    k = col + 1
        return out

    def _has_decoration(hf_v2):
        if not isinstance(hf_v2, dict):
            return False
        if hf_v2.get('header_after') or hf_v2.get('footer_after'):
            return True
        for v in (hf_v2.get('per_sheet') or {}).values():
            if v and (v.get('header_after') or v.get('footer_after')):
                return True
        return False

    def _cut_original_footer(body):
        """截掉正文网格底部原文件自带的表脚区（签章/空行/残留行），只保留数据区（数据行 + 合计行）。
        从下向上找最后的数据行（序号为整数字的行）或合计行，其下方一律视为表脚区丢弃。"""
        if not body:
            return body
        data_end = None
        for i in range(len(body) - 1, -1, -1):
            row = body[i]
            if not row:
                continue
            txts = [str(c) for c in row if c not in (None, '')]
            if not txts:
                continue
            if any('合计' in t or '总计' in t for t in txts):
                data_end = i
                break
            if row[0] not in (None, '') and _is_num_str(row[0]):
                try:
                    if float(str(row[0]).replace(',', '')) == int(float(str(row[0]).replace(',', ''))):
                        data_end = i
                        break
                except Exception:
                    pass
        if data_end is None:
            return body
        return body[:data_end + 1]

    def _apply_style(cell, font_style, align=None, wrap=None):
        """应用字体/对齐样式（font_style: {name,size,bold}；align: left/center/right；wrap: True/False/None 默认 True）"""
        from openpyxl.styles import Alignment, Font
        if font_style:
            cell.font = Font(name=font_style.get('name') or '宋体',
                             size=int(font_style.get('size') or 11),
                             bold=bool(font_style.get('bold')))
        if align is not None or wrap is not None:
            cell.alignment = Alignment(
                horizontal=(align if align is not None else 'center'),
                vertical='center',
                wrap_text=(wrap if wrap is not None else True)
            )

    def _build_report_workbook(sheets, hf_v2, variant, style_map=None):
        """按统一模板构建工作簿：after 版本对每张表加模板表头表脚并填充个税；before 版本正文透传。
        style_map：{工作表名: style}，正文区（数据区）复用导入表格的列宽/行高/字体/边框/对齐/数字格式。"""
        from openpyxl.styles import Alignment, Font, Border, Side
        from openpyxl.utils import get_column_letter
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        wb = Workbook()
        wb.remove(wb.active)
        for idx, sh in enumerate(sheets):
            grid = sh.get('grid') or []
            title = (sh.get('name') or ('工资表%d' % (idx + 1)))[:31]
            ws = wb.create_sheet(title=title)
            hidx = _detect_body_start(grid)
            body = grid[hidx:] if variant == 'after' else grid
            if variant == 'after':
                # 截掉原文件自带的表脚区（避免与模板表脚重复），并去掉正文末尾全空行
                body = _cut_original_footer(body)
            # 去掉正文末尾全空行，避免数据区与表脚之间出现空隙
            while body and all(c == '' for c in body[-1]):
                body.pop()
            hf = _sheet_hf(hf_v2, sh.get('name'))
            font_style = hf.get('font') if isinstance(hf, dict) else None
            if variant == 'after':
                _fill_tax(body, (hf_v2 or {}).get('tax'))
            hf = _sheet_hf(hf_v2, sh.get('name'))
            maxc = max((len(r) for r in body), default=1)
            # 正文区样式（导入表格格式）：按工作表名匹配（名称精确 → strip 比对 → 按顺序回退）
            style = (style_map or {}).get(sh.get('name')) if style_map else None
            if style is None and style_map:
                _sname = str(sh.get('name') or '').strip()
                for _k, _v in style_map.items():
                    if str(_k).strip() == _sname:
                        style = _v
                        break
                if style is None:
                    # 顺序回退：导出表与源表 sheet 通常同序，防止名称细微差异导致样式整体丢失
                    _sm_vals = list(style_map.values())
                    if idx < len(_sm_vals):
                        style = _sm_vals[idx]
            default_font = (style or {}).get('default_font') or {'name': '宋体', 'size': 11}
            # 序号列不参与数字格式统一（避免 1 -> 1.00），仅对金额等数值列统一
            det = sh.get('detect') or {}
            seq_col = det.get('seq_col')
            # 识别身份证/账号等“应作为文本”的列（按表头关键字）
            header_row = (grid[hidx] if (variant == 'after' and 0 <= hidx < len(grid)) else (grid[0] if grid else []))
            text_cols = set()
            for j, hname in enumerate(header_row):
                hn = str(hname)
                if any(k in hn for k in ('身份证', '证件', '身份', '账号', '银行卡', '银行账号')):
                    text_cols.add(j)
            r = 1
            if variant == 'after' and hf.get('header_after'):
                header_fonts = hf.get('header_fonts') or []
                header_layouts = hf.get('header_layouts') or []
                header_aligns = hf.get('header_aligns') or []
                for i, line in enumerate(hf['header_after']):
                    hfont = None
                    if isinstance(header_fonts, list) and i < len(header_fonts) and header_fonts[i]:
                        hfont = header_fonts[i]
                    else:
                        hfont = font_style
                    hlayout = 'merge'
                    if isinstance(header_layouts, list) and i < len(header_layouts) and header_layouts[i]:
                        hlayout = header_layouts[i]
                    halign = 'center'
                    if isinstance(header_aligns, list) and i < len(header_aligns) and header_aligns[i]:
                        halign = header_aligns[i]
                    elif hf.get('header_align'):
                        halign = hf.get('header_align')
                    nonblank = [(j, v) for j, v in enumerate(line) if v != '']
                    if hlayout != 'merge' or not (len(nonblank) == 1 and nonblank[0][0] == 0):
                        # 逐列定位（不合并）：支持「列字母:文本」指定任意列；内容超出默认不换行（表头不画边框）
                        for col, text in _parse_col_text_cells(line):
                            if col < 1 or col > maxc:
                                continue
                            cell = ws.cell(row=r, column=col, value=text)
                            _apply_style(cell, hfont, halign, wrap=False)
                    else:
                        # 合并居中：仅在 A 列有内容时跨列合并（表头不画边框）
                        cell = ws.cell(row=r, column=1, value=nonblank[0][1])
                        if maxc > 1:
                            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=  maxc)
                        _apply_style(cell, hfont, halign)
                    r += 1
            # ===== 正文数据区：按导入表格格式呈现 =====
            # 数值单元格统一使用“常规”格式（不套用源表千分位/小数位等自定义格式），数值本身保留原有小数位精度
            for bi, row in enumerate(body):
                orig_row = (hidx + bi) if variant == 'after' else bi
                for j, v in enumerate(row, 1):
                    c0 = j - 1
                    raw_v = v
                    # 日期：去掉 datetime 转 str 带来的尾部 00:00:00
                    if isinstance(raw_v, str):
                        dm = _DATE_RE.match(raw_v)
                        if dm and dm.group(2):
                            raw_v = dm.group(1)
                    in_text_col = (c0 in text_cols) or (isinstance(raw_v, str) and _looks_like_id(raw_v))
                    if in_text_col:
                        # 身份证/账号等：按文本存储，避免丢失前导零或变科学计数
                        cell = ws.cell(row=r, column=j, value=raw_v)
                        cell.number_format = '@'
                    else:
                        f = _payslip_to_float(raw_v)
                        cell = ws.cell(row=r, column=j, value=(f if f is not None else raw_v))
                    # 套用导入表格的字体/对齐（边框改为正文区统一自动添加，不依赖源表是否带框线）
                    st = ((style or {}).get('cells') or {}).get((orig_row, c0))
                    if st:
                        cell.font = Font(name=st['name'], size=st['size'], bold=st['bold'])
                        cell.alignment = Alignment(horizontal=(st.get('align') or 'center'),
                                                   vertical='center', wrap_text=bool(st.get('wrap')))
                    else:
                        # 无逐格样式：字体沿用工作簿默认字体（多数中文表为宋体）
                        cell.font = Font(name=default_font['name'], size=default_font['size'], bold=False)
                        if not in_text_col and style is None and f is None:
                            cell.alignment = Alignment(horizontal='center')
                    # 数值格式：统一为常规（General），数值值本身按原样保留小数位精度；身份证/账号仍为文本
                    if (not in_text_col) and (f is not None):
                        cell.number_format = 'General'
                    # 正文数据区统一加细边框（无论源表是否带框线，保证有数值区域都有框线）
                    cell.border = border
                # 行高：复用导入表格的行高（显式行高优先；未显式设置时回退工作表默认行高，保持与导入表一致）
                if style:
                    h = (style.get('heights') or {}).get(orig_row)
                    if h:
                        ws.row_dimensions[r].height = h
                    else:
                        _drh = style.get('default_row_height')
                        if _drh:
                            ws.row_dimensions[r].height = _drh
                r += 1
            if variant == 'after' and hf.get('footer_after'):
                flayout = (hf.get('footer_layout') or 'cols')
                if flayout == 'merge':
                    # 合并居中：同一行表脚合并 A..maxc 并居中，多段用空格连接
                    for line in hf['footer_after']:
                        vals = [v for v in line if v != '']
                        text = ' '.join(vals).strip()
                        if text == '':
                            r += 1
                            continue
                        cell = ws.cell(row=r, column=1, value=text)
                        if maxc > 1:
                            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=maxc)
                        _apply_style(cell, font_style, (hf.get('footer_align') or 'center'))
                        r += 1
                else:
                    # 逐列定位：支持「列字母:文本」指定任意列；单元格内用空格分隔的多个片段也分别落位
                    for line in hf['footer_after']:
                        k = 1
                        for raw_cell in line:
                            if raw_cell is None or raw_cell == '':
                                k += 1
                                continue
                            # 一个单元格内可能含多个「列字母:文本」片段（用空格分隔），逐个落位
                            for tok in re.split(r'\s+', str(raw_cell).strip()):
                                if tok == '':
                                    continue
                                col = k
                                text = tok
                                m = re.match(r'^([A-Za-z]{1,2})[:：]\s*(.*)$', tok)
                                if m and 0 < _col_to_idx(m.group(1)) <= 16384:
                                    col = _col_to_idx(m.group(1))
                                    text = m.group(2)
                                cell = ws.cell(row=r, column=col, value=text)
                                _apply_style(cell, font_style, (hf.get('footer_align') or 'center'))
                                if not m:
                                    k = col + 1
                        r += 1
            # 列宽：复用导入表格的列宽（作用于整列，含表头表脚区）
            if style and style.get('widths'):
                for c0, w in (style['widths'] or {}).items():
                    if c0 < 0:
                        continue
                    ws.column_dimensions[get_column_letter(c0 + 1)].width = w
            # 隐藏列：复用导入表格的隐藏状态（被隐藏的列仍按源表写入数据，只是不显示）
            if style and style.get('hidden_cols'):
                for c0 in style['hidden_cols']:
                    if c0 < 0:
                        continue
                    ws.column_dimensions[get_column_letter(c0 + 1)].hidden = True
            # 默认列宽：正文中未显式设置宽度的列，套用工作表默认列宽（保持与导入表一致）
            _dcw = (style or {}).get('default_col_width') if style else None
            if _dcw:
                _explicit_w = (style or {}).get('widths') or {}
                _hide_c = (style or {}).get('hidden_cols') or set()
                for _c0 in range(maxc):
                    if _c0 in _explicit_w or _c0 in _hide_c:
                        continue
                    ws.column_dimensions[get_column_letter(_c0 + 1)].width = _dcw
            # 页面/打印设置：纸张方向、纸型、缩放、边距按导入表原样（打印观感与导入表一致）
            _psu = (style or {}).get('page_setup') if style else None
            if _psu:
                if _psu.get('orientation'):
                    ws.page_setup.orientation = _psu['orientation']
                if _psu.get('paperSize'):
                    ws.page_setup.paperSize = _psu['paperSize']
                if _psu.get('scale'):
                    ws.page_setup.scale = _psu['scale']
                if _psu.get('fitToWidth') is not None:
                    ws.page_setup.fitToWidth = _psu['fitToWidth']
                if _psu.get('fitToHeight') is not None:
                    ws.page_setup.fitToHeight = _psu['fitToHeight']
                if _psu.get('fitToPage') is not None:
                    try:
                        from openpyxl.worksheet.properties import PageSetupProperties
                        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=_psu['fitToPage'])
                    except Exception:
                        pass
            _pmg = (style or {}).get('page_margins') if style else None
            if _pmg:
                for _mk, _mv in _pmg.items():
                    if _mv is not None:
                        try:
                            setattr(ws.page_margins, _mk, _mv)
                        except Exception:
                            pass
        if not wb.sheetnames:
            wb.create_sheet('工资表')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf


    @app.route('/payslip-summary')
    @login_required
    def salary_summary_page():
        return render_template('payslip_summary.html')

    @app.route('/api/payslip/check', methods=['POST'])
    @login_required
    def api_payslip_check():
        f = request.files.get('file')
        if not f:
            return jsonify({'error': '未上传文件'}), 400
        name = (f.filename or '').lower()
        try:
            data = f.read()
            raw_sheets = _parse_payslip_grid(data, name)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {e}'}), 400
        sheets = []
        for sh in raw_sheets:
            grid = sh.get('grid') or []
            if not grid:
                continue
            det = _detect_payslip(grid)
            errors = _validate_payslip(grid, det)
            sheets.append({'name': sh['name'], 'grid': grid, 'detect': det, 'errors': errors})
        if not sheets:
            return jsonify({'error': '文件中没有数据'}), 400
        # 缓存源文件字节，供导出时复用其列宽/行高/字体/边框样式
        import uuid
        token = uuid.uuid4().hex
        PAYSLIP_SRC_STORE[token] = (name, data)
        if len(PAYSLIP_SRC_STORE) > 32:
            # 仅保留最近 32 份
            for old in list(PAYSLIP_SRC_STORE)[:len(PAYSLIP_SRC_STORE) - 32]:
                PAYSLIP_SRC_STORE.pop(old, None)
        return jsonify({'sheets': sheets, 'multi': len(sheets) > 1, 'src_token': token})

    @app.route('/api/payslip/export', methods=['POST'])
    @login_required
    def api_payslip_export():
        # 兼容 multipart（方式一：支持 file 字节 / template_id 样式模板）与 JSON（默认）
        # variant: 'after' = 校验后（应用模板表头表脚+个税）；'before' = 校验前（正文透传）
        sheets, header_rows, footer_rows, headfoot = None, [], [], None
        using_form = bool(request.files) or (bool(request.form) and 'sheets' in request.form)
        payload = request.get_json(silent=True) or {}
        variant = (request.form.get('variant') or payload.get('variant') or 'after')
        styled = None
        hf_v2 = None
        tax = None
        if using_form:
            raw = request.form.get('sheets')
            sheets = json.loads(raw) if raw else None
            header_rows = [h for h in (request.form.get('header_rows') or '').split('\n') if h.strip()]
            footer_rows = [f for f in (request.form.get('footer_rows') or '').split('\n') if f.strip()]
            hf = request.form.get('headfoot') or None
            try:
                headfoot = json.loads(hf) if hf else None
            except Exception:
                headfoot = None
            tid = request.form.get('template_id')
            if tid:
                t = db.session.get(PayslipTemplate, int(tid))
                if t:
                    styled = t.file_blob
                    if t.headfoot_json:
                        try:
                            hf_v2 = json.loads(t.headfoot_json)
                        except Exception:
                            hf_v2 = None
                    if isinstance(hf_v2, dict):
                        tax = hf_v2.get('tax')
            # 一次性模板：form 直接传 v2 表头表脚（含 per_sheet / tax）时按 v2 处理（覆盖模板库加载）
            if isinstance(headfoot, dict) and (headfoot.get('per_sheet') or headfoot.get('tax')):
                hf_v2 = headfoot
                tax = hf_v2.get('tax')
            if styled is None and request.files.get('file'):
                styled = request.files.get('file').read()
        else:
            sheets = payload.get('sheets')
            header_rows = payload.get('header_rows') or []
            footer_rows = payload.get('footer_rows') or []
            headfoot = payload.get('headfoot') or None
        if not sheets or not isinstance(sheets, list):
            return jsonify({'error': '缺少表格数据'}), 400
        # 复用导入表样式：优先用随导出请求上传的源文件（最可靠，不依赖缓存），其次用校验时缓存的 src_token
        style_map = None
        src_file = request.files.get('src_file')
        if src_file is not None:
            try:
                _sdata = src_file.read()
                _sname = (getattr(src_file, 'filename', '') or '工资表.xlsx').lower()
                style_map = {sh['name']: sh.get('style') for sh in _parse_payslip_grid(_sdata, _sname)}
            except Exception:
                style_map = None
        if style_map is None:
            src_token = request.form.get('src_token') or payload.get('src_token')
            if src_token and src_token in PAYSLIP_SRC_STORE:
                try:
                    sname, sdata = PAYSLIP_SRC_STORE[src_token]
                    style_map = {sh['name']: sh.get('style') for sh in _parse_payslip_grid(sdata, sname)}
                except Exception:
                    style_map = None
        filename = request.form.get('filename') or payload.get('filename') or '工资表.xlsx'
        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'
        try:
            decoration = _has_decoration(hf_v2) if isinstance(hf_v2, dict) else False
            tax_on = bool(tax and tax.get('enabled'))
            if variant == 'after' and (decoration or tax_on or style_map):
                buf = _build_report_workbook(sheets, hf_v2, 'after', style_map)
            elif styled and not decoration:
                buf = _build_styled_workbook_multi(sheets, styled)
            else:
                buf = _build_payslip_workbook(sheets, header_rows, footer_rows,
                                              headfoot if variant == 'after' else None)
        except Exception as e:
            return jsonify({'error': f'生成失败: {e}'}), 400
        return _send_xlsx(buf, filename)

    def _send_xlsx(buf, filename):
        """导出 xlsx 并保证中文下载文件名正确（修复 Flask fallback 文件名被剥离成 '().xlsx' 的问题）。
        filename* 携带完整中文名（现代浏览器优先采用）；filename 仅作老客户端兜底。"""
        from urllib.parse import quote
        resp = send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp.headers['Content-Disposition'] = "attachment; filename=\"export.xlsx\"; filename*=UTF-8''%s" % quote(filename)
        return resp

    # ===================== 工资表汇总制作：表头/表脚模板（含合并单元格） =====================
    @app.route('/api/payslip/headfoot/sample')
    @login_required
    def api_payslip_headfoot_sample():
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side
        wb = openpyxl.Workbook()
        thin = Side(style='thin', color='D0D0D0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # 表头：三行，均合并为整行（示例 8 列宽，导出时按数据表宽度钳制）
        hs = wb.active
        hs.title = '表头'
        for rr, txt in [(1, '空港航食有限责任公司工资表'), (2, '期间：2026-08'), (3, '制表人：　　　　审核人：')]:
            c = hs.cell(row=rr, column=1, value=txt)
            hs.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.font = Font(bold=True, size=12)
        fs = wb.create_sheet('表脚')
        for rr, txt in [(1, '制表：　　　　审核：　　　　负责人：'), (2, '2026-08-31')]:
            c = fs.cell(row=rr, column=1, value=txt)
            fs.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
            c.alignment = Alignment(horizontal='center', vertical='center')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='工资表表头表脚模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/payslip/headfoot/import', methods=['POST'])
    @login_required
    def api_payslip_headfoot_import():
        f = request.files.get('file')
        if not f:
            return jsonify({'error': '未上传文件'}), 400
        name = (f.filename or '').lower()
        try:
            hf = _parse_headfoot_template(f.read(), name)
        except Exception as e:
            return jsonify({'error': f'模板解析失败: {e}'}), 400
        Setting.set('payslip_headfoot', json.dumps(hf, ensure_ascii=False))
        db.session.commit()
        return jsonify(hf)

    @app.route('/api/payslip/headfoot')
    @login_required
    def api_payslip_headfoot_get():
        raw = Setting.get('payslip_headfoot', '')
        if not raw:
            return jsonify({'header': {'grid': [], 'merges': []}, 'footer': {'grid': [], 'merges': []}})
        try:
            return jsonify(json.loads(raw))
        except Exception:
            return jsonify({'header': {'grid': [], 'merges': []}, 'footer': {'grid': [], 'merges': []}})

    # ===================== 个税管理 =====================
    @app.route('/tax')
    @login_required
    def tax_page():
        return render_template('tax.html')

    def _build_tax_rows(year, keyword, units):
        """按身份证分组，用累计预扣法计算每人每月个税；units 为客户单位过滤（空=不过滤）。"""
        q = SalaryRecord.query.filter(SalaryRecord.period.like(year + '%'))
        if keyword:
            like = f'%{keyword}%'
            q = q.filter(or_(SalaryRecord.name.like(like), SalaryRecord.id_card.like(like)))
        if units:
            q = q.filter(SalaryRecord.client_unit.in_(units))
        recs = q.all()
        from collections import defaultdict
        people = defaultdict(list)
        for r in recs:
            people[r.id_card].append(r)
        rows = []
        for idc, lst in people.items():
            lst.sort(key=lambda x: x.period)
            months = {}
            cum_income = 0.0
            cum_deduct = 0.0
            cum_tax_due = 0.0
            total_calc = 0.0
            total_rec = 0.0
            remark = ''
            for r in lst:
                should = float(r.should_pay or 0)
                social = float(r.social_personal or 0)
                fund = float(r.fund_personal or 0)
                cum_income += should
                cum_deduct += 5000 + social + fund
                cum_taxable = max(0.0, cum_income - cum_deduct)
                tax_due = calc_income_tax(cum_taxable, 'cum')
                month_tax = max(0.0, tax_due - cum_tax_due)
                cum_tax_due = tax_due
                rec_tax = abs(float(r.tax or 0))
                months[r.period] = {
                    'calc': round(month_tax, 2), 'rec': round(rec_tax, 2),
                    'taxable': round(cum_taxable, 2), 'should': round(should, 2)
                }
                total_calc += month_tax
                total_rec += rec_tax
                # 取最新月份的非空备注作为人员级备注
                if r.remark and r.remark.strip():
                    remark = r.remark.strip()
            sample = lst[-1]  # 取最新月份记录用于客户单位/姓名
            dept = (sample.client_unit or '').strip() or _dept_of(sample.employee_id)
            rows.append({
                'id_card': idc, 'name': sample.name or '', 'department': dept,
                'client_unit': (sample.client_unit or '').strip(),
                'months': months, 'total_calc': round(total_calc, 2),
                'total_rec': round(total_rec, 2), 'diff': round(total_calc - total_rec, 2),
                'remark': remark
            })
        rows.sort(key=lambda x: (x['department'], x['name']))
        months_list = sorted({p for row in rows for p in row['months'].keys()})
        return rows, months_list

    @app.route('/api/tax/summary')
    @login_required
    def api_tax_summary():
        year = (request.args.get('year') or '').strip() or str(date.today().year)
        keyword = (request.args.get('keyword') or '').strip()
        units = [u for u in (request.args.get('units') or '').split(',') if u]
        try:
            page = int(request.args.get('page', 1))
            size = int(request.args.get('size', 10))
        except (ValueError, TypeError):
            page, size = 1, 10
        if size <= 0:
            size = 10
        rows, months_list = _build_tax_rows(year, keyword, units)
        total = len(rows)
        paged = rows[(page - 1) * size:page * size]
        return jsonify({'items': paged, 'total': total, 'page': page, 'size': size,
                        'year': year, 'months': months_list,
                        'units': client_unit_options()})

    @app.route('/api/tax/person/<id_card>')
    @login_required
    def api_tax_person(id_card):
        """获取某人在指定年份的逐月个税明细（含可编辑的原始字段）。"""
        year = (request.args.get('year') or '').strip() or str(date.today().year)
        recs = SalaryRecord.query.filter(
            SalaryRecord.id_card == id_card,
            SalaryRecord.period.like(year + '%')
        ).order_by(SalaryRecord.period.asc()).all()
        if not recs:
            return jsonify({'error': '未找到该人员记录'}), 404
        records = []
        for r in recs:
            records.append({
                'period': r.period,
                'should_pay': round(float(r.should_pay or 0), 2),
                'social_personal': round(float(r.social_personal or 0), 2),
                'fund_personal': round(float(r.fund_personal or 0), 2),
                'tax': round(float(r.tax or 0), 2),
                'remark': r.remark or '',
            })
        cum_income = 0.0
        cum_deduct = 0.0
        cum_tax_due = 0.0
        for rec in records:
            cum_income += rec['should_pay']
            cum_deduct += 5000 + rec['social_personal'] + rec['fund_personal']
            cum_taxable = max(0.0, cum_income - cum_deduct)
            tax_due = calc_income_tax(cum_taxable, 'cum')
            month_tax = max(0.0, tax_due - cum_tax_due)
            cum_tax_due = tax_due
            rec['calc'] = round(month_tax, 2)
            rec['rec'] = abs(rec['tax'])
        sample = recs[-1]
        return jsonify({
            'id_card': id_card,
            'name': recs[0].name or '',
            'department': (sample.client_unit or '').strip() or _dept_of(sample.employee_id),
            'client_unit': (sample.client_unit or '').strip(),
            'year': year,
            'units': client_unit_options(),
            'records': records,
        })

    @app.route('/api/tax/person', methods=['PUT'])
    @role_required('admin', 'hr')
    def api_tax_update_person():
        """修改某人在指定年份的逐月个税明细（客户单位 + 各月应发/社保/公积金/个税/备注）。"""
        data = request.get_json(silent=True) or {}
        id_card = (data.get('id_card') or '').strip()
        year = (data.get('year') or '').strip()
        if not id_card:
            return jsonify({'error': '缺少身份证号'}), 400
        q = SalaryRecord.query.filter(SalaryRecord.id_card == id_card)
        if year:
            q = q.filter(SalaryRecord.period.like(year + '%'))
        recs = q.all()
        if not recs:
            return jsonify({'error': '未找到该人员记录'}), 404

        def to_f(v):
            try:
                return float(v)
            except (ValueError, TypeError):
                return 0.0

        if 'client_unit' in data:
            cu = (data.get('client_unit') or '').strip()
            for r in recs:
                r.client_unit = cu
        by_period = {r.period: r for r in recs}
        for m in (data.get('months') or []):
            p = m.get('period')
            r = by_period.get(p)
            if not r:
                continue
            if 'should_pay' in m:
                r.should_pay = to_f(m['should_pay'])
            if 'social_personal' in m:
                r.social_personal = to_f(m['social_personal'])
            if 'fund_personal' in m:
                r.fund_personal = to_f(m['fund_personal'])
            if 'tax' in m:
                r.tax = abs(to_f(m['tax']))  # 个税统一为正（扣款金额）
            if 'remark' in m:
                r.remark = (m.get('remark') or '').strip()
        db.session.commit()
        log_operation('个税管理', '修改', f'{year} {id_card}', f'客户单位={data.get("client_unit", "")}')
        return jsonify({'msg': '已保存'})

    @app.route('/api/tax/export')
    @login_required
    def api_tax_export():
        year = (request.args.get('year') or '').strip() or str(date.today().year)
        keyword = (request.args.get('keyword') or '').strip()
        units = [u for u in (request.args.get('units') or '').split(',') if u]
        rows, months_list = _build_tax_rows(year, keyword, units)
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '个税明细'
        head = ['序号', '客户单位', '员工姓名', '身份证号'] \
               + [m[5:] + '月' for m in months_list] \
               + ['系统计算', '原记录', '差异', '备注']
        ws.append(head)
        hdr_fill = PatternFill('solid', fgColor='2C5AA0')
        hdr_font = Font(color='FFFFFF', bold=True)
        thin = Side(style='thin', color='D0D7DE')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c in ws[1]:
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        for i, d in enumerate(rows, 1):
            line = [i, d['department'], d['name'], d['id_card']]
            for m in months_list:
                c = d['months'].get(m)
                line.append(abs(c['calc']) if c else '')
            line += [abs(d['total_calc']), abs(d['total_rec']), abs(d['diff']), d.get('remark', '')]
            ws.append(line)
        # 数值列右对齐 + 边框
        num_start = 5  # 第5列起为各月个税
        num_end = 4 + len(months_list) + 3  # 含全年计算/系统记录/差异
        for r in range(2, ws.max_row + 1):
            for cidx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=cidx)
                cell.border = border
                if num_start <= cidx <= num_end and cidx != 4:  # 身份证号(4)左对齐
                    cell.alignment = Alignment(horizontal='right')
        # 列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        for cidx in range(5, 5 + len(months_list)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(cidx)].width = 10
        ws.column_dimensions[openpyxl.utils.get_column_letter(5 + len(months_list))].width = 12
        ws.column_dimensions[openpyxl.utils.get_column_letter(6 + len(months_list))].width = 12
        ws.column_dimensions[openpyxl.utils.get_column_letter(7 + len(months_list))].width = 12
        ws.column_dimensions[openpyxl.utils.get_column_letter(8 + len(months_list))].width = 20
        ws.freeze_panes = 'A2'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f'个税明细_{year}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/tax/recalculate', methods=['POST'])
    @role_required('admin', 'hr')
    def api_tax_recalculate():
        data = request.get_json(silent=True) or {}
        year = (data.get('year') or '').strip() or str(date.today().year)
        q = SalaryRecord.query.filter(SalaryRecord.period.like(year + '%'))
        recs = q.all()
        from collections import defaultdict
        people = defaultdict(list)
        for r in recs:
            people[r.id_card].append(r)
        updated = 0
        for idc, lst in people.items():
            lst.sort(key=lambda x: x.period)
            cum_income = 0.0
            cum_deduct = 0.0
            cum_tax_due = 0.0
            for r in lst:
                should = float(r.should_pay or 0)
                social = float(r.social_personal or 0)
                fund = float(r.fund_personal or 0)
                cum_income += should
                cum_deduct += 5000 + social + fund
                cum_taxable = max(0.0, cum_income - cum_deduct)
                tax_due = calc_income_tax(cum_taxable, 'cum')
                month_tax = max(0.0, tax_due - cum_tax_due)
                cum_tax_due = tax_due
                if round(float(r.tax or 0), 2) != round(month_tax, 2):
                    r.tax = round(month_tax, 2)
                    updated += 1
        db.session.commit()
        log_operation('个税管理', '按年累计重算', year, f'更新 {updated} 条个税')
        return jsonify({'msg': f'已按 {year} 年累计预扣法重算，更新 {updated} 条', 'updated': updated})

    # ===================== 工资表汇总制作：生成模板 & 社保关联 =====================
    PAYSLIP_TPL_KEY = 'payslip_gen_template'
    # 社保列定义（kind=insurance；src 对应 InsuranceDetail 字段）
    INSURANCE_COLUMNS = [
        {'key': 'ins_base', 'label': '社保缴费基数', 'kind': 'insurance', 'src': 'base'},
        {'key': 'ins_pension_per', 'label': '养老个人', 'kind': 'insurance', 'src': 'pension_per'},
        {'key': 'ins_pension_emp', 'label': '养老单位', 'kind': 'insurance', 'src': 'pension_emp'},
        {'key': 'ins_medical_per', 'label': '医疗个人', 'kind': 'insurance', 'src': 'medical_per'},
        {'key': 'ins_medical_emp', 'label': '医疗单位', 'kind': 'insurance', 'src': 'medical_emp'},
        {'key': 'ins_extra_medical_per', 'label': '大额医疗个人', 'kind': 'insurance', 'src': 'extra_medical_per'},
        {'key': 'ins_extra_medical_emp', 'label': '大额医疗单位', 'kind': 'insurance', 'src': 'extra_medical_emp'},
        {'key': 'ins_unemployment_per', 'label': '失业个人', 'kind': 'insurance', 'src': 'unemployment_per'},
        {'key': 'ins_unemployment_emp', 'label': '失业单位', 'kind': 'insurance', 'src': 'unemployment_emp'},
        {'key': 'ins_injury_emp', 'label': '工伤单位', 'kind': 'insurance', 'src': 'injury_emp'},
        {'key': 'ins_maternity_emp', 'label': '生育单位', 'kind': 'insurance', 'src': 'maternity_emp'},
        {'key': 'ins_fund_base', 'label': '公积金缴费基数', 'kind': 'insurance', 'src': 'fund_base'},
        {'key': 'ins_fund_rate', 'label': '公积金缴费比例%', 'kind': 'insurance', 'src': 'fund_rate'},
        {'key': 'ins_fund_per', 'label': '公积金个人', 'kind': 'insurance', 'src': 'fund_per'},
        {'key': 'ins_fund_emp', 'label': '公积金单位', 'kind': 'insurance', 'src': 'fund_emp'},
    ]

    def _fmt_num(v):
        try:
            f = round(float(v), 2)
        except (TypeError, ValueError):
            return '' if v is None else str(v)
        if f == int(f):
            return str(int(f))
        return ('%.2f' % f).rstrip('0').rstrip('.')

    def payslip_template_columns():
        """当前生效的生成模板列（持久化于 Setting 'payslip_gen_template'，导入即覆盖）"""
        raw = Setting.get(PAYSLIP_TPL_KEY, '')
        if raw:
            try:
                tpl = json.loads(raw)
                if isinstance(tpl, list) and tpl:
                    return tpl
            except Exception:
                pass
        return [{'key': c['key'], 'label': c['label'], 'kind': 'salary'} for c in SALARY_COLUMNS]

    def _match_template_col(header):
        """表头文本 -> 匹配列定义；无法识别返回 None"""
        h = str(header).strip()
        if h == '序号':
            return {'key': 'seq', 'label': '序号', 'kind': 'seq'}
        aliases = {'姓名': ('name', '员工姓名'), '身份证': ('id_card', '身份证号'),
                   '部门': ('department', '客户单位'), '客户单位': ('department', '客户单位'),
                   '月份': ('period', '期间'), '缴费基数': ('ins_base', '社保缴费基数', 'insurance', 'base')}
        if h in aliases:
            a = aliases[h]
            if len(a) == 4:
                return {'key': a[0], 'label': a[1], 'kind': a[2], 'src': a[3]}
            return {'key': a[0], 'label': a[1], 'kind': 'salary'}
        for c in SALARY_COLUMNS:
            if h == c['label'] or h == c['key']:
                m = {'key': c['key'], 'label': c['label'], 'kind': 'salary'}
                # 若与社保列重名（如"公积金个人"），标记社保来源：关联时优先取五险一金数据
                for ic in INSURANCE_COLUMNS:
                    if ic['label'] == c['label']:
                        m['ambig_src'] = ic['src']
                        break
                return m
        for c in INSURANCE_COLUMNS:
            if h == c['label'] or h == c['key']:
                return dict(c)
        return None

    @app.route('/api/payslip/template')
    @login_required
    def api_payslip_template_get():
        return jsonify({'columns': payslip_template_columns()})

    @app.route('/api/payslip/template/sample')
    @login_required
    def api_payslip_template_sample():
        from openpyxl import Workbook
        sal_labels = {c['label'] for c in SALARY_COLUMNS}
        labels = [c['label'] for c in SALARY_COLUMNS] + \
                 [c['label'] for c in INSURANCE_COLUMNS if c['label'] not in sal_labels]
        wb = Workbook()
        ws = wb.active
        ws.title = '生成模板'
        for j, lbl in enumerate(labels, 1):
            ws.cell(row=1, column=j, value=lbl)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='工资表生成模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    _KNOWN_HDR = ['序号', '姓名', '身份证', '部门', '客户单位', '月份', '期间', '应发', '实发',
                  '社保', '公积金', '养老', '医疗', '失业', '工伤', '生育', '基数', '个人',
                  '单位', '服务费', '备注', '科室', '人次', '小计']

    def _find_header_row(rows):
        """在模板各行中定位表头行：与已知列关键字/生成列匹配最多的行"""
        best_i, best = 0, -1
        for i, row in enumerate(rows):
            score = 0
            nonempty = 0
            for c in row:
                s = str(c).strip()
                if not s:
                    continue
                nonempty += 1
                if s in _KNOWN_HDR or any(k in s for k in _KNOWN_HDR):
                    score += 1
            # 表头行应同时“匹配多”且“单元格较多”
            if score > best and nonempty >= 2:
                best = score
                best_i = i
        return best_i

    def _parse_template_columns(file_bytes, name):
        """从模板文件解析生成列定义，自动跳过标题行定位真正的表头行。返回 (tpl, unknown)"""
        if name.endswith('.csv'):
            import csv as _csv
            text = file_bytes.decode('utf-8-sig', errors='ignore')
            rows = [[c for c in r] for r in _csv.reader(io.StringIO(text))]
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            rows = [[c for c in r] for r in wb.active.iter_rows(values_only=True)]
        if not rows:
            raise ValueError('模板未识别到内容')
        hi = _find_header_row(rows)
        headers = [str(c).strip() for c in rows[hi] if str(c).strip() != '']
        if not headers:
            headers = [str(c).strip() for c in rows[0] if str(c).strip() != '']
        if not headers:
            raise ValueError('模板未识别到列名')
        tpl, unknown = [], []
        for h in headers:
            m = _match_template_col(h)
            if m:
                tpl.append(m)
            else:
                tpl.append({'key': 'lit_%d' % len(tpl), 'label': h, 'kind': 'lit'})
                unknown.append(h)
        return tpl, unknown

    @app.route('/api/payslip/template/import', methods=['POST'])
    @role_required('admin', 'hr')
    def api_payslip_template_import():
        """方式二：导入模板覆盖全局生效（持久化于 Setting 'payslip_gen_template'），新模板从导入开始生效"""
        f = request.files.get('file')
        if not f:
            return jsonify({'error': '未上传文件'}), 400
        name = (f.filename or '').lower()
        try:
            data = f.read()
            tpl, unknown = _parse_template_columns(data, name)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {e}'}), 400
        Setting.set(PAYSLIP_TPL_KEY, json.dumps(tpl, ensure_ascii=False))
        db.session.commit()
        log_operation('薪酬绩效', '模板导入', '工资表生成模板', '列数 %d，未识别 %d' % (len(tpl), len(unknown)))
        return jsonify({'columns': tpl, 'unknown': unknown, 'file_name': f.filename})

    @app.route('/api/payslip/template/preview', methods=['POST'])
    @login_required
    def api_payslip_template_preview():
        """方式一：仅解析模板列定义与按 sheet 表头表脚返回，不保存（供预览/临时使用）"""
        f = request.files.get('file')
        if not f:
            return jsonify({'error': '未上传文件'}), 400
        name = (f.filename or '').lower()
        try:
            data = f.read()
            tpl, unknown = _parse_template_columns(data, name)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {e}'}), 400
        # 自动提取按 sheet 表头表脚 + 个税建议
        per_sheet, tax = {}, None
        try:
            for sh in _parse_payslip_grid(data, name):
                per_sheet[sh['name']] = _extract_report_hf(sh['grid'])
                if tax is None:
                    s = _suggest_tax(sh['grid'])
                    if s:
                        tax = s
        except Exception:
            pass
        headfoot = {'per_sheet': per_sheet, 'tax': tax or {'enabled': False, 'target': '税额-所得税', 'source': '税前合计', 'method': 'cum', 'threshold': 5000, 'mode': 'keep'}}
        return jsonify({'columns': tpl, 'unknown': unknown, 'file_name': f.filename, 'headfoot': headfoot})

    def _build_hf_v2(file_bytes, fname, explicit_headfoot, auto_hf):
        """构造 v2 表头表脚：优先显式传入；否则按 auto_hf 从文件自动提取；返回 (dict|None, tax)"""
        if explicit_headfoot:
            try:
                p = json.loads(explicit_headfoot)
                if isinstance(p, dict):
                    return p, p.get('tax')
            except Exception:
                pass
        if auto_hf == '1':
            per_sheet, tax = {}, None
            try:
                for sh in _parse_payslip_grid(file_bytes, fname):
                    per_sheet[sh['name']] = _extract_report_hf(sh['grid'])
                    if tax is None:
                        s = _suggest_tax(sh['grid'])
                        if s:
                            tax = s
            except Exception:
                pass
            return {'per_sheet': per_sheet, 'tax': tax or {'enabled': False, 'target': '税额-所得税', 'source': '税前合计', 'method': 'cum', 'threshold': 5000, 'mode': 'keep'}}, (tax or {'enabled': False, 'target': '税额-所得税', 'source': '税前合计', 'method': 'cum', 'threshold': 5000, 'mode': 'keep'})
        return None, None

    @app.route('/api/payslip/templates', methods=['POST'])
    @role_required('admin', 'hr')
    def api_payslip_templates_create():
        """保存为固定模板（命名、可带样式文件、可带按 sheet 表头表脚 + 个税）"""
        f = request.files.get('file')
        name = (request.form.get('name') or '').strip()
        if not f:
            return jsonify({'error': '请上传模板文件'}), 400
        if not name:
            name = (f.filename or '未命名模板').rsplit('.', 1)[0]
        fname = (f.filename or '').lower()
        try:
            data = f.read()
            tpl, unknown = _parse_template_columns(data, fname)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {e}'}), 400
        headfoot_v2, tax = _build_hf_v2(data, fname, request.form.get('headfoot') or '',
                                        request.form.get('auto_hf'))
        headfoot_json = json.dumps(headfoot_v2, ensure_ascii=False) if headfoot_v2 else None
        t = PayslipTemplate(name=name, columns_json=json.dumps(tpl, ensure_ascii=False),
                            file_blob=data, file_name=f.filename,
                            headfoot_json=headfoot_json)
        db.session.add(t)
        db.session.commit()
        log_operation('薪酬绩效', '保存生成模板', name, '列数 %d，未识别 %d，含样式=%s，含表头表脚=%s，个税=%s'
                      % (len(tpl), len(unknown), bool(data), bool(headfoot_json), bool(tax and tax.get('enabled'))))
        return jsonify(t.to_dict())

    @app.route('/api/payslip/templates', methods=['GET'])
    @login_required
    def api_payslip_templates_list():
        items = [t.to_dict() for t in PayslipTemplate.query.order_by(PayslipTemplate.updated_at.desc()).all()]
        return jsonify({'items': items})

    @app.route('/api/payslip/templates/<int:tid>', methods=['GET'])
    @login_required
    def api_payslip_templates_get(tid):
        t = db.session.get(PayslipTemplate, tid)
        if not t:
            return jsonify({'error': '模板不存在'}), 404
        return jsonify(t.to_dict())

    @app.route('/api/payslip/templates/<int:tid>', methods=['PUT'])
    @role_required('admin', 'hr')
    def api_payslip_templates_update(tid):
        """改名 / 重新导入样式文件 / 更新表头表脚"""
        t = db.session.get(PayslipTemplate, tid)
        if not t:
            return jsonify({'error': '模板不存在'}), 404
        name = (request.form.get('name') or '').strip()
        f = request.files.get('file')
        if name:
            t.name = name
        if f and f.filename:
            fname = f.filename.lower()
            try:
                data = f.read()
                tpl, unknown = _parse_template_columns(data, fname)
            except Exception as e:
                return jsonify({'error': f'文件解析失败: {e}'}), 400
            t.columns_json = json.dumps(tpl, ensure_ascii=False)
            t.file_blob = data
            t.file_name = f.filename
        # 表头/表脚：显式传 headfoot（v2）则更新；auto_hf=1 从文件重新提取；headfoot_clear=1 则清空
        if 'headfoot' in request.form:
            hf = request.form.get('headfoot') or ''
            headfoot_json = None
            if hf:
                try:
                    parsed = json.loads(hf)
                    if isinstance(parsed, dict):
                        headfoot_json = json.dumps(parsed, ensure_ascii=False)
                except Exception:
                    pass
            t.headfoot_json = headfoot_json
        elif request.form.get('auto_hf') == '1' and f and f.filename:
            headfoot_v2, _ = _build_hf_v2(data, f.filename.lower(), '', '1')
            t.headfoot_json = json.dumps(headfoot_v2, ensure_ascii=False) if headfoot_v2 else None
        elif request.form.get('headfoot_clear') == '1':
            t.headfoot_json = None
        t.updated_at = datetime.now()
        db.session.commit()
        return jsonify(t.to_dict())

    @app.route('/api/payslip/templates/<int:tid>', methods=['DELETE'])
    @role_required('admin', 'hr')
    def api_payslip_templates_delete(tid):
        t = db.session.get(PayslipTemplate, tid)
        if not t:
            return jsonify({'error': '模板不存在'}), 404
        db.session.delete(t)
        db.session.commit()
        log_operation('薪酬绩效', '删除生成模板', t.name, '')
        return jsonify({'msg': '已删除'})

    def _resolve_template_columns(data):
        """优先 columns > template_id 对应固定模板 > 默认全列"""
        cols = data.get('columns')
        if isinstance(cols, list) and cols:
            return cols
        tid = data.get('template_id')
        if tid:
            t = db.session.get(PayslipTemplate, int(tid))
            if t and t.columns_json:
                try:
                    return json.loads(t.columns_json)
                except Exception:
                    pass
        return payslip_template_columns()

    def _generate_payslip_grid(period, associate, ins_period, tpl):
        q = SalaryRecord.query
        if period:
            q = q.filter(SalaryRecord.period == period)
        rows = q.order_by(SalaryRecord.id.asc()).all()
        emp_cache = {}
        def dept_of(eid):
            if not eid:
                return '-'
            if eid not in emp_cache:
                e = db.session.get(Employee, eid)
                emp_cache[eid] = e.department if e else '-'
            return emp_cache[eid]
        ins_map = {}
        if associate:
            iq = InsuranceDetail.query
            if ins_period:
                iq = iq.filter(InsuranceDetail.period == ins_period)
            for d in iq.all():
                ins_map[d.employee_id] = d
        emp_by_card = {}
        def ins_of(rec):
            ins = ins_map.get(rec.employee_id)
            if ins is None and rec.id_card:
                if rec.id_card not in emp_by_card:
                    e = Employee.query.filter_by(id_card=rec.id_card).first()
                    emp_by_card[rec.id_card] = e.id if e else None
                eid = emp_by_card.get(rec.id_card)
                if eid:
                    ins = ins_map.get(eid)
            return ins
        grid = [[c['label'] for c in tpl]]
        for i, r in enumerate(rows, 1):
            sal = serialize_salary(r, dept_of(r.employee_id))
            ins = ins_of(r) if associate else None
            row = []
            for c in tpl:
                k = c['key']
                if c['kind'] == 'seq':
                    row.append(str(i))
                elif c['kind'] == 'salary':
                    v = sal.get(k)
                    # 重名列（如"公积金个人"）：关联且五险一金有值时取社保值，否则取导入工资值
                    if associate and ins is not None and c.get('ambig_src'):
                        iv = getattr(ins, c['ambig_src'], None)
                        if iv is not None:
                            v = iv
                    row.append('' if v is None else (_fmt_num(v) if isinstance(v, (int, float)) else str(v)))
                elif c['kind'] == 'insurance':
                    v = getattr(ins, c['src'], None) if ins else None
                    row.append('' if v is None else _fmt_num(v))
                else:
                    row.append('')
            grid.append(row)
        return grid, len(rows)

    @app.route('/api/payslip/generate', methods=['POST'])
    @login_required
    def api_payslip_generate():
        data = request.get_json(silent=True) or {}
        period = (data.get('period') or '').strip()
        associate = bool(data.get('associate'))
        # 社保关联月份：优先使用用户选择的 ins_period，否则回退到工资期间
        ins_period = (data.get('ins_period') or '').strip() or period
        tpl = _resolve_template_columns(data)
        grid, total = _generate_payslip_grid(period, associate, ins_period, tpl)
        return jsonify({'grid': grid, 'total': total})

    def _fill_styled_sheet(ws, grid, thin, border):
        """在以带样式模板为骨架的 ws 上，按列位置写入 grid 数据（保留表头行字体/列宽等格式）"""
        from openpyxl.styles import Font, Alignment
        labels = [str(c) for c in grid[0]] if grid else []
        # 定位表头行：与 grid 表头标签匹配最多的行
        header_row, best = 1, -1
        for ri in range(1, min(ws.max_row, 30) + 1):
            score = 0
            for ci in range(1, ws.max_column + 1):
                if str(ws.cell(row=ri, column=ci).value or '').strip() in labels:
                    score += 1
            if score > best:
                best = score
                header_row = ri
        # 清空表头行以下所有值
        for ri in range(header_row + 1, ws.max_row + 1):
            for ci in range(1, ws.max_column + 1):
                ws.cell(row=ri, column=ci).value = None
        # 写入数据：沿用表头行的字体/数字格式
        for r, row in enumerate(grid[1:], start=header_row + 1):
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=ci)
                hc = ws.cell(row=header_row, column=ci)
                f = _payslip_to_float(val)
                cell.value = f if f is not None else (val if val != '' else None)
                try:
                    cell.font = Font(name=hc.font.name, size=hc.font.size or 11,
                                     bold=hc.font.bold, italic=hc.font.italic, color=hc.font.color)
                except Exception:
                    pass
                cell.alignment = Alignment(horizontal=('right' if f is not None else 'center'),
                                           vertical='center', wrap_text=True)
                cell.border = border
                if hc.number_format and hc.number_format not in ('General',):
                    cell.number_format = hc.number_format

    def _build_styled_workbook_multi(sheets, template_bytes):
        """多 sheet 按同一带样式模板骨架导出：每张表套用模板格式，表名取原表名"""
        from openpyxl import load_workbook
        from openpyxl.styles import Side, Border
        thin = Side(style='thin', color='D0D0D0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        wb_tpl = load_workbook(io.BytesIO(template_bytes), data_only=False)
        ws_tpl = wb_tpl.active
        n = len(sheets)
        # 先复制 N 份干净骨架（基于未修改的模板），再逐张填数据，避免互相污染
        sheets_ws = [ws_tpl] + [wb_tpl.copy_worksheet(ws_tpl) for _ in range(1, n)] if n else []
        for idx, sh in enumerate(sheets):
            grid = sh.get('grid') or []
            title = (sh.get('name') or ('工资表%d' % (idx + 1)))[:31]
            ws = sheets_ws[idx]
            ws.title = title
            if grid:
                _fill_styled_sheet(ws, grid, thin, border)
        if not sheets_ws:
            wb_tpl.create_sheet('工资表')
        buf = io.BytesIO()
        wb_tpl.save(buf)
        buf.seek(0)
        return buf

    @app.route('/api/payslip/generate/export', methods=['POST'])
    @login_required
    def api_payslip_generate_export():
        # multipart：period/associate/ins_period/columns/header_rows/footer_rows/headfoot
        period = (request.form.get('period') or '').strip()
        associate = request.form.get('associate') in ('1', 'true', 'True')
        ins_period = (request.form.get('ins_period') or '').strip() or period
        columns = request.form.get('columns')
        tpl = None
        if columns:
            try:
                tpl = json.loads(columns)
                if not isinstance(tpl, list) or not tpl:
                    tpl = None
            except Exception:
                tpl = None
        # 未传列则使用全局生效模板（导入覆盖生效）
        if not isinstance(tpl, list) or not tpl:
            tpl = payslip_template_columns()
        grid, total = _generate_payslip_grid(period, associate, ins_period, tpl)
        filename = request.form.get('filename') or '工资表(生成).xlsx'
        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'
        try:
            header_rows = [h for h in (request.form.get('header_rows') or '').split('\n') if h.strip()]
            footer_rows = [f for f in (request.form.get('footer_rows') or '').split('\n') if f.strip()]
            hf = request.form.get('headfoot') or None
            try:
                hf = json.loads(hf) if hf else None
            except Exception:
                hf = None
            buf = _build_payslip_workbook([{'name': '工资表', 'grid': grid}],
                                         header_rows, footer_rows, hf)
        except Exception as e:
            return jsonify({'error': f'生成失败: {e}'}), 400
        return _send_xlsx(buf, filename)

    @app.route('/api/insurance/<int:did>', methods=['DELETE'])
    @role_required('admin', 'hr')
    def api_insurance_delete(did):
        d = db.session.get(InsuranceDetail, did)
        if not d:
            return jsonify({'error': '不存在'}), 404
        if insurance_period_locked(d.period) and current_user.role != 'admin':
            return jsonify({'error': f'{d.period} 社保数据已锁定，仅管理员可修改'}), 403
        db.session.delete(d)
        db.session.commit()
        return jsonify({'msg': '已删除'})

    @app.route('/api/insurance/<int:did>', methods=['PUT'])
    @role_required('admin', 'hr')
    def api_insurance_update(did):
        """修改导入的保险明细：缴费基数/各险种金额/公积金基数与比例/人员编号/备注"""
        d = db.session.get(InsuranceDetail, did)
        if not d:
            return jsonify({'error': '不存在'}), 404
        if insurance_period_locked(d.period) and current_user.role != 'admin':
            return jsonify({'error': f'{d.period} 社保数据已锁定，仅管理员可修改'}), 403
        data = request.get_json(silent=True) or {}

        def _f(v):
            try:
                return round(float(v), 2)
            except (ValueError, TypeError):
                return 0.0

        # 金额字段（含公积金基数/比例）
        num_fields = ['base', 'pension_emp', 'pension_per', 'medical_emp', 'medical_per',
                      'extra_medical_emp', 'extra_medical_per', 'unemployment_emp',
                      'unemployment_per', 'injury_emp', 'maternity_emp',
                      'fund_base', 'fund_rate', 'fund_emp', 'fund_per']
        for k in num_fields:
            if k in data:
                setattr(d, k, _f(data[k]))
        # 人员编号 / 备注
        if 'person_no' in data:
            d.person_no = str(data.get('person_no') or '').strip()
        if 'remark' in data:
            d.remark = str(data.get('remark') or '').strip()
        db.session.commit()
        log_operation('五险一金', '修改', f'{d.period}',
                      f'修改 {d.employee.name}(id={d.employee_id}) 的社保明细')
        return jsonify({'msg': f'{d.employee.name} 的 {d.period} 社保明细已更新',
                        'item': serialize_insurance(d)})

    # ---------- 数据备份（整包：全局库 data.db + 各公司账套库 companies/*.db） ----------
    @app.route('/api/backup/export')
    @role_required('admin', 'hr')
    def api_backup_export():
        """导出全部数据为 zip 整包：data.db + company_core.db + 各公司账套库 companies/*.db + 上传文件 uploads/"""
        import io, zipfile, json
        # 先提交所有未写变更（含各公司库）
        try:
            db.session.commit()
        except Exception:
            pass
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f'hrms_backup_{ts}.zip'
        data_db = os.path.join(BASE_DIR, 'data.db')
        core_db = os.path.join(BASE_DIR, 'company_core.db')
        comp_dir = os.path.join(BASE_DIR, 'companies')
        up_dir = os.path.join(BASE_DIR, 'uploads')
        mem = io.BytesIO()
        company_files = []
        upload_count = 0
        with zipfile.ZipFile(mem, 'w', zipfile.ZIP_DEFLATED) as zf:
            if os.path.exists(data_db):
                zf.write(data_db, 'data.db')
            if os.path.exists(core_db):
                zf.write(core_db, 'company_core.db')
            if os.path.isdir(comp_dir):
                for fn in sorted(os.listdir(comp_dir)):
                    if fn.endswith('.db'):
                        full = os.path.join(comp_dir, fn)
                        zf.write(full, os.path.join('companies', fn))
                        company_files.append(fn)
            # 上传文件（员工照片/附件/公司 Logo 等），按 uploads/ 相对路径打包
            if os.path.isdir(up_dir):
                for root, _dirs, files in os.walk(up_dir):
                    for fn in files:
                        full = os.path.join(root, fn)
                        rel = os.path.relpath(full, BASE_DIR)  # uploads/...
                        zf.write(full, rel)
                        upload_count += 1
            manifest = {
                'app': 'hrms',
                'version': APP_VERSION,
                'created': ts,
                'type': 'full',
                'global_db': 'data.db' if os.path.exists(data_db) else None,
                'core_db': 'company_core.db' if os.path.exists(core_db) else None,
                'company_dbs': company_files,
                'uploads': upload_count,
            }
            zf.writestr('backup_manifest.json', json.dumps(manifest, ensure_ascii=False, indent=2))
        mem.seek(0)
        log_operation('系统', '导出备份', fname,
                      f'全量备份：{len(company_files)} 个公司账套库 + {upload_count} 个上传文件，'
                      f'{round(mem.getbuffer().nbytes / 1024 / 1024, 2)} MB')
        return send_file(mem, as_attachment=True, download_name=fname, mimetype='application/zip')

    @app.route('/api/backup/import', methods=['POST'])
    @role_required('admin', 'hr')
    def api_backup_import():
        """导入备份恢复：支持新版 .zip 整包，兼容旧版单库 .db/.sqlite/.sqlite3"""
        import shutil, tempfile, zipfile, io
        f = request.files.get('file')
        if not f or not f.filename:
            return jsonify({'error': '请选择备份文件'}), 400
        fname = f.filename.lower()
        if not (fname.endswith('.zip') or fname.endswith('.db') or fname.endswith('.sqlite') or fname.endswith('.sqlite3')):
            return jsonify({'error': '仅支持 .zip / .db / .sqlite / .sqlite3 备份文件'}), 400
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        data_db = os.path.join(BASE_DIR, 'data.db')
        core_db = os.path.join(BASE_DIR, 'company_core.db')
        comp_dir = os.path.join(BASE_DIR, 'companies')
        up_dir = os.path.join(BASE_DIR, 'uploads')
        # 先把当前数据整体备份一份（便于失败回滚）：数据库 + 上传文件
        rollback_dir = os.path.join(BASE_DIR, f'data_backup_{ts}')
        try:
            os.makedirs(rollback_dir, exist_ok=True)
            if os.path.exists(data_db):
                shutil.copy2(data_db, os.path.join(rollback_dir, 'data.db'))
            if os.path.exists(core_db):
                shutil.copy2(core_db, os.path.join(rollback_dir, 'company_core.db'))
            if os.path.isdir(comp_dir):
                for fn in os.listdir(comp_dir):
                    if fn.endswith('.db'):
                        shutil.copy2(os.path.join(comp_dir, fn), os.path.join(rollback_dir, fn))
            if os.path.isdir(up_dir):
                shutil.copytree(up_dir, os.path.join(rollback_dir, 'uploads'))
        except Exception as e:
            return jsonify({'error': f'备份当前数据失败: {e}'}), 500
        # 关闭并释放所有连接，避免文件被占用
        try:
            db.session.close()
            db.engine.dispose()
        except Exception:
            pass
        for eng in list(_CDB_ENGINES.values()):
            try:
                eng.dispose()
            except Exception:
                pass
        _CDB_ENGINES.clear()
        restored_company_files = set()
        try:
            if fname.endswith('.zip'):
                # 解压整包并按结构还原
                tmpd = tempfile.mkdtemp(dir=BASE_DIR)
                try:
                    f.save(os.path.join(tmpd, '_upload.zip'))
                    with zipfile.ZipFile(os.path.join(tmpd, '_upload.zip')) as zf:
                        zf.extractall(tmpd)
                except Exception as e:
                    raise RuntimeError(f'压缩包解析失败: {e}')
                # 还原全局库
                src_data = os.path.join(tmpd, 'data.db')
                if os.path.exists(src_data):
                    shutil.copy2(src_data, data_db)
                # 还原兜底库 company_core.db（若备份中存在）
                src_core = os.path.join(tmpd, 'company_core.db')
                if os.path.exists(src_core):
                    shutil.copy2(src_core, core_db)
                # 还原各公司账套库
                os.makedirs(comp_dir, exist_ok=True)
                src_comp = os.path.join(tmpd, 'companies')
                if os.path.isdir(src_comp):
                    for fn in os.listdir(src_comp):
                        if fn.endswith('.db'):
                            shutil.copy2(os.path.join(src_comp, fn), os.path.join(comp_dir, fn))
                            restored_company_files.add(fn)
                # 彻底替换：删除当前存在但备份中不含的公司库（已留存于 rollback_dir）
                if os.path.isdir(comp_dir):
                    for fn in os.listdir(comp_dir):
                        if fn.endswith('.db') and fn not in restored_company_files:
                            try:
                                os.remove(os.path.join(comp_dir, fn))
                            except Exception:
                                pass
                # 还原上传文件（员工照片/附件/公司 Logo 等）
                src_up = os.path.join(tmpd, 'uploads')
                if os.path.isdir(src_up):
                    if os.path.isdir(up_dir):
                        shutil.rmtree(up_dir, ignore_errors=True)
                    shutil.copytree(src_up, up_dir)
                shutil.rmtree(tmpd, ignore_errors=True)
            else:
                # 旧版单库备份：仅还原全局库 data.db（兼容历史备份）
                f.save(data_db)
        except Exception as e:
            # 回滚到导入前的状态
            try:
                if os.path.exists(data_db):
                    os.remove(data_db)
                if os.path.isdir(comp_dir):
                    for fn in os.listdir(comp_dir):
                        if fn.endswith('.db'):
                            os.remove(os.path.join(comp_dir, fn))
                for fn in os.listdir(rollback_dir):
                    if fn.endswith('.db'):
                        dst = data_db if fn == 'data.db' else (core_db if fn == 'company_core.db' else os.path.join(comp_dir, fn))
                        shutil.copy2(os.path.join(rollback_dir, fn), dst)
                # 恢复上传文件
                rb_up = os.path.join(rollback_dir, 'uploads')
                if os.path.isdir(rb_up):
                    if os.path.isdir(up_dir):
                        shutil.rmtree(up_dir, ignore_errors=True)
                    shutil.copytree(rb_up, up_dir)
            except Exception:
                pass
            return jsonify({'error': f'写入数据库失败，已回滚: {e}'}), 500
        msg_suffix = ''
        if not fname.endswith('.zip'):
            msg_suffix = '（旧版单库备份：仅还原全局库，各公司账套数据未变动）'
        log_operation('系统', '数据恢复', '导入备份数据库',
                     f'文件 {f.filename}，原数据已备份至 {os.path.basename(rollback_dir)}{msg_suffix}')
        return jsonify({'msg': '数据恢复成功，请重新登录以生效',
                        'backup': os.path.basename(rollback_dir),
                        'restored_company_dbs': sorted(restored_company_files),
                        'legacy_single_db': not fname.endswith('.zip')})

    @app.route('/api/backup/info')
    @role_required('admin', 'hr')
    def api_backup_info():
        """获取全部数据文件信息（含各公司账套库业务数据量与上传文件数）"""
        from sqlalchemy import text
        data_db = os.path.join(BASE_DIR, 'data.db')
        core_db = os.path.join(BASE_DIR, 'company_core.db')
        comp_dir = os.path.join(BASE_DIR, 'companies')
        up_dir = os.path.join(BASE_DIR, 'uploads')
        dbfiles = []
        total_size = 0
        mtime = None
        if os.path.exists(data_db):
            total_size += os.path.getsize(data_db)
            mtime = os.path.getmtime(data_db)
            dbfiles.append('data.db')
        if os.path.exists(core_db):
            total_size += os.path.getsize(core_db)
            mtime = max(mtime, os.path.getmtime(core_db)) if mtime is not None else os.path.getmtime(core_db)
            dbfiles.append('company_core.db')
        comp_db_count = 0
        if os.path.isdir(comp_dir):
            for fn in sorted(os.listdir(comp_dir)):
                if fn.endswith('.db'):
                    p = os.path.join(comp_dir, fn)
                    total_size += os.path.getsize(p)
                    mtime = max(mtime, os.path.getmtime(p)) if mtime is not None else os.path.getmtime(p)
                    dbfiles.append(fn)
                    comp_db_count += 1
        # 上传文件统计（员工照片/附件/公司 Logo）
        upload_files = 0
        if os.path.isdir(up_dir):
            for root, _dirs, files in os.walk(up_dir):
                upload_files += len(files)
        # 跨全部公司账套库统计真实业务数据（员工/社保/工资含绩效工资）
        stats = {}
        try:
            stats['公司账套'] = Company.query.count()
            stats['用户'] = User.query.count()
            total_emp = total_ins = total_sal = 0
            for c in Company.query.all():
                eng = company_engine(c.id)
                if not eng:
                    continue
                try:
                    with eng.connect() as conn:
                        total_emp += conn.execute(text('SELECT COUNT(*) FROM employees')).scalar() or 0
                        total_ins += conn.execute(text('SELECT COUNT(*) FROM insurance_details')).scalar() or 0
                        total_sal += conn.execute(text('SELECT COUNT(*) FROM salary_records')).scalar() or 0
                except Exception:
                    pass
            stats['员工(合计)'] = total_emp
            stats['社保明细(合计)'] = total_ins
            stats['工资/绩效(合计)'] = total_sal
        except Exception:
            pass
        return jsonify({
            'size': total_size,
            'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S') if mtime else '',
            'db_files': dbfiles,
            'company_db_count': comp_db_count,
            'company_dbs': [fn for fn in dbfiles if fn not in ('data.db', 'company_core.db')],
            'stats': stats,
            'upload_files': upload_files,
        })

    # ---------- 待办 ----------
    @app.route('/api/todos')
    @login_required
    def api_todos_list():
        cid = current_company().id if current_company() else None
        q = Todo.query.filter_by(user_id=current_user.id)
        if cid is not None:
            q = q.filter(Todo.company_id == cid)
        done = request.args.get('done')
        if done is not None:
            q = q.filter(Todo.done == (done == '1'))
        rows = q.order_by(Todo.due_date.asc()).all()
        return jsonify({'items': [serialize_todo(t) for t in rows]})

    @app.route('/api/todos', methods=['POST'])
    @login_required
    def api_todos_create():
        data = request.get_json(force=True)
        remind_count = int(data.get('remind_count') or 1)
        if remind_count < 1:
            remind_count = 1
        if remind_count > 50:
            remind_count = 50
        # 提醒时间段：开始/结束时间；只填其一则两端相同；起>止自动交换
        rs = parse_datetime(data.get('remind_start'))
        re_ = parse_datetime(data.get('remind_end'))
        if rs and re_ and rs > re_:
            rs, re_ = re_, rs
        elif rs and not re_:
            re_ = rs
        elif re_ and not rs:
            rs = re_
        t = Todo(
            user_id=current_user.id,
            company_id=current_company().id if current_company() else None,
            title=data['title'],
            due_date=parse_date(data.get('due_date')) or date.today(),
            priority=data.get('priority', '中'),
            remind_start=rs,
            remind_end=re_,
            remind_count=remind_count,
            reminded_count=0
        )
        db.session.add(t)
        db.session.commit()
        log_operation('待办', '新增', t.title, f'截止 {t.due_date}，提醒 {t.remind_count} 次')
        return jsonify({'id': t.id, 'msg': '已创建'}), 201

    @app.route('/api/todos/<int:tid>', methods=['PUT'])
    @login_required
    def api_todos_update(tid):
        t = db.session.get(Todo, tid)
        if not t or t.user_id != current_user.id:
            return jsonify({'error': '不存在或无权'}), 404
        data = request.get_json(force=True)
        for f in ['title', 'priority']:
            if f in data:
                setattr(t, f, data[f])
        if 'due_date' in data:
            t.due_date = parse_date(data['due_date']) or t.due_date
        if 'remind_start' in data or 'remind_end' in data:
            # 重新设定提醒时间段：归一化（只填其一则两端相同；起>止交换）并重置已提醒次数
            rs = parse_datetime(data.get('remind_start')) if 'remind_start' in data else t.remind_start
            re_ = parse_datetime(data.get('remind_end')) if 'remind_end' in data else t.remind_end
            if rs and re_ and rs > re_:
                rs, re_ = re_, rs
            elif rs and not re_:
                re_ = rs
            elif re_ and not rs:
                rs = re_
            t.remind_start = rs
            t.remind_end = re_
            t.reminded_count = 0
        if 'remind_count' in data:
            n = int(data['remind_count'] or 1)
            t.remind_count = max(1, min(50, n))
        if 'done' in data:
            t.done = bool(data['done'])
            if t.done:
                # 完成后不再提醒
                t.reminded_count = t.remind_count
        db.session.commit()
        log_operation('待办', '更新', t.title, f'截止 {t.due_date}')
        return jsonify({'msg': '已更新'})

    @app.route('/api/todos/<int:tid>', methods=['DELETE'])
    @login_required
    def api_todos_delete(tid):
        t = db.session.get(Todo, tid)
        if not t or t.user_id != current_user.id:
            return jsonify({'error': '不存在或无权'}), 404
        db.session.delete(t)
        db.session.commit()
        log_operation('待办', '删除', t.title, f'截止 {t.due_date}')
        return jsonify({'msg': '已删除'})

    @app.route('/api/todos/<int:tid>/ack', methods=['POST'])
    @login_required
    def api_todos_ack(tid):
        """确认收到提醒：已提醒次数 +1（用于右下角弹窗每次弹出后记录）"""
        t = db.session.get(Todo, tid)
        if not t or t.user_id != current_user.id:
            return jsonify({'error': '不存在或无权'}), 404
        t.reminded_count = (t.reminded_count or 0) + 1
        db.session.commit()
        return jsonify({'msg': 'ok'})

    @app.route('/api/todos/bell')
    @login_required
    def api_todos_bell():
        """顶部铃铛：所有过期未办待办（未完成且截止日期<=今天）"""
        items = due_todo_query(current_user.id).order_by(Todo.due_date.asc()).all()
        return jsonify({
            'count': len(items),
            'items': [serialize_todo(t) for t in items[:20]]
        })

    def reminder_slots(t):
        """按提醒开始/结束时间，将提醒次数均匀分配到时间段内（不连续）"""
        if not t.remind_start or not t.remind_end or not t.remind_count or t.remind_count < 1:
            return []
        n = t.remind_count
        if n == 1:
            return [t.remind_start]
        total = (t.remind_end - t.remind_start).total_seconds()
        return [t.remind_start + timedelta(seconds=total * i / (n - 1)) for i in range(n)]

    @app.route('/api/todos/reminders')
    @login_required
    def api_todos_reminders():
        """右下角弹窗提醒：在提醒时间段内、且当前已到下一次应提醒时刻的待办。
        待办被完成(done)或删除后自然停止；reminded_count 达 remind_count 后停止。"""
        now = datetime.now()
        out = []
        cid = current_company().id if current_company() else None
        _tq = Todo.query.filter_by(user_id=current_user.id, done=False)
        if cid is not None:
            _tq = _tq.filter(Todo.company_id == cid)
        todos = _tq.all()
        for t in todos:
            if not t.remind_start or not t.remind_end:
                continue
            if now < t.remind_start or now > t.remind_end:
                continue
            slots = reminder_slots(t)
            idx = t.reminded_count or 0
            if idx < len(slots) and slots[idx] <= now:
                out.append(serialize_todo(t))
            if len(out) >= 10:
                break
        return jsonify({'count': len(out), 'items': out})

    # ---------- 操作记录查询 ----------
    @app.route('/api/logs')
    @role_required('admin', 'hr')
    def api_logs_list():
        q = OperationLog.query
        module = request.args.get('module')
        if module:
            q = q.filter(OperationLog.module == module)
        action = request.args.get('action')
        if action:
            q = q.filter(OperationLog.action == action)
        op = request.args.get('operator')
        if op:
            q = q.filter(OperationLog.operator.contains(op) | OperationLog.operator_name.contains(op))
        kw = request.args.get('keyword')
        if kw:
            q = q.filter(OperationLog.target.contains(kw) | OperationLog.detail.contains(kw))
        s = request.args.get('start')
        e = request.args.get('end')
        if s:
            sd = parse_date(s)
            if sd:
                q = q.filter(OperationLog.created_at >= datetime.combine(sd, datetime.min.time()))
        if e:
            ed = parse_date(e)
            if ed:
                q = q.filter(OperationLog.created_at <= datetime.combine(ed, datetime.max.time()))
        rows = q.order_by(OperationLog.created_at.desc(), OperationLog.id.desc()).limit(500).all()
        return jsonify({
            'items': [{
                'id': r.id,
                'operator': r.operator, 'operator_name': r.operator_name or '',
                'module': r.module, 'action': r.action,
                'target': r.target or '', 'detail': r.detail or '',
                'ip': r.ip or '',
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
            } for r in rows]
        })

    # ---------- 用户/账号管理 ----------
    @app.route('/api/users')
    @login_required
    def api_users_list():
        # 范围：管理员看全部；HR 看本人 + 员工；员工只看本人
        if current_user.role == 'admin':
            rows = User.query.order_by(User.id.asc()).all()
        elif current_user.role == 'hr':
            rows = User.query.filter(
                (User.id == current_user.id) | (User.role == 'employee')
            ).order_by(User.id.asc()).all()
        else:  # employee
            rows = User.query.filter_by(id=current_user.id).all()
        return jsonify({'items': [serialize_user(u) for u in rows]})

    @app.route('/api/users', methods=['POST'])
    @role_required('admin')
    def api_users_create():
        data = request.get_json(force=True)
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'error': '用户名已存在'}), 400
        u = User(
            username=data['username'],
            name=data.get('name', ''),
            role=data.get('role', 'employee'),
            employee_id=data.get('employee_id') or None
        )
        if u.role == 'hr':
            for mod in MODULES:
                setattr(u, f'can_{mod}_view', bool(data.get(f'can_{mod}_view', True)))
                setattr(u, f'can_{mod}_manage', bool(data.get(f'can_{mod}_manage', True)))
        elif u.role == 'employee':
            u.emp_view_mode = data.get('emp_view_mode', 'self')
            depts = data.get('emp_depts', [])
            u.emp_depts = json.dumps(depts) if isinstance(depts, list) else str(depts)
        u.set_password(data.get('password', '123456'))
        db.session.add(u)
        db.session.commit()
        # 关联可访问公司（仅 hr/employee 生效；admin 恒拥有全部）
        if u.role in ('hr', 'employee') and data.get('companies'):
            for cid in [int(x) for x in data['companies'] if str(x).isdigit()]:
                if Company.query.get(cid) and not UserCompany.query.filter_by(user_id=u.id, company_id=cid).first():
                    db.session.add(UserCompany(user_id=u.id, company_id=cid))
            db.session.commit()
        log_operation('账号', '新增', u.username, f'姓名 {u.name}，角色 {u.role}')
        return jsonify({'id': u.id, 'msg': '已创建'}), 201

    @app.route('/api/users/<int:uid>', methods=['PUT'])
    @login_required
    def api_users_update(uid):
        u = db.session.get(User, uid)
        if not u:
            return jsonify({'error': '不存在'}), 404
        # 范围校验：员工只能改本人；HR 只能改本人或员工账号
        if current_user.role == 'employee':
            if u.id != current_user.id:
                return jsonify({'error': '只能修改本人账号'}), 403
        elif current_user.role == 'hr':
            if u.id != current_user.id and u.role != 'employee':
                return jsonify({'error': 'HR 只能管理本人及员工账号'}), 403
        data = request.get_json(force=True)
        # 角色修改限制：仅管理员可改他人角色，且不能改自己角色（防自锁/提权）
        if 'role' in data and (current_user.role != 'admin' or u.id == current_user.id):
            data.pop('role')
        # 用户名不可修改；尤其保护系统账号 hr（防止改名后无法识别/登录异常）
        if 'username' in data:
            if u.username == 'hr':
                return jsonify({'error': '系统账号 hr 的用户名不可修改'}), 403
            data.pop('username')
        for f in ['name', 'role', 'employee_id']:
            if f in data:
                setattr(u, f, data[f] or None)
        if u.role == 'hr':
            for mod in MODULES:
                if f'can_{mod}_view' in data:
                    setattr(u, f'can_{mod}_view', bool(data[f'can_{mod}_view']))
                if f'can_{mod}_manage' in data:
                    setattr(u, f'can_{mod}_manage', bool(data[f'can_{mod}_manage']))
        elif u.role == 'employee':
            if 'emp_view_mode' in data:
                u.emp_view_mode = data['emp_view_mode']
            if 'emp_depts' in data:
                depts = data['emp_depts']
                u.emp_depts = json.dumps(depts) if isinstance(depts, list) else str(depts)
        if data.get('password'):
            u.set_password(data['password'])
        # 同步可访问公司（仅 hr/employee；admin 忽略）
        if u.role in ('hr', 'employee') and 'companies' in data:
            want = set(int(x) for x in data['companies'] if str(x).isdigit())
            have = {uc.company_id for uc in UserCompany.query.filter_by(user_id=u.id).all()}
            for cid in (want - have):
                if Company.query.get(cid):
                    db.session.add(UserCompany(user_id=u.id, company_id=cid))
            for uc in UserCompany.query.filter_by(user_id=u.id).all():
                if uc.company_id not in want:
                    db.session.delete(uc)
        db.session.commit()
        log_operation('账号', '编辑', u.username, f'姓名 {u.name}，角色 {u.role}')
        return jsonify({'msg': '已更新'})

    @app.route('/api/users/<int:uid>', methods=['DELETE'])
    @role_required('admin')
    def api_users_delete(uid):
        u = db.session.get(User, uid)
        if not u:
            return jsonify({'error': '不存在'}), 404
        if u.id == current_user.id:
            return jsonify({'error': '不能删除自己'}), 400
        log_operation('账号', '删除', u.username, f'姓名 {u.name}')
        db.session.delete(u)
        db.session.commit()
        return jsonify({'msg': '已删除'})

    # ---------- 看板统计 ----------
    @app.route('/api/dashboard/stats')
    @login_required
    def api_stats():
        emp_base = employee_scope_filter(Employee.query)
        in_service = emp_base.filter_by(status='在职').count()
        resigned = emp_base.filter_by(status='离职').count()
        retired = emp_base.filter_by(status='退休').count()
        if current_user.role == 'employee':
            allowed_ids = [e.id for e in employee_scope_filter(Employee.query).all()]
            active_contracts = Contract.query.filter(
                Contract.status == '生效',
                Contract.employee_id.in_(allowed_ids)).count() if allowed_ids else 0
        else:
            active_contracts = Contract.query.filter_by(status='生效').count()
        return jsonify({
            'in_service': in_service, 'resigned': resigned, 'retired': retired,
            'active_contracts': active_contracts
        })

    # ---------- 代码字段（字典）维护 ----------
    @app.route('/api/dicts/categories')
    @login_required
    def api_dict_categories():
        rows = db.session.query(DictItem.category).distinct().all()
        return jsonify({'items': [r[0] for r in rows]})

    @app.route('/api/dicts')
    @login_required
    def api_dict_list():
        cat = request.args.get('category')
        q = DictItem.query.filter(DictItem.enabled == True)  # noqa: E712
        if cat:
            q = q.filter(DictItem.category == cat)
        rows = q.order_by(DictItem.sort.asc(), DictItem.id.asc()).all()
        return jsonify({'items': [serialize_dict(d) for d in rows]})

    @app.route('/api/dicts', methods=['POST'])
    @role_required('admin', 'hr')
    def api_dict_create():
        data = request.get_json(force=True)
        if not data.get('category') or not data.get('label'):
            return jsonify({'error': '分类和选项值不能为空'}), 400
        if DictItem.query.filter_by(category=data['category'], label=data['label']).first():
            return jsonify({'error': '该选项已存在'}), 400
        d = DictItem(
            category=data['category'],
            label=data['label'],
            sort=data.get('sort', 0),
            enabled=data.get('enabled', True),
            remark=data.get('remark', '')
        )
        db.session.add(d)
        db.session.commit()
        log_operation('字典', '新增', f'{d.category}/{d.label}', d.remark or '')
        return jsonify({'id': d.id, 'msg': '已新增'}), 201

    @app.route('/api/dicts/<int:did>', methods=['PUT'])
    @role_required('admin', 'hr')
    def api_dict_update(did):
        d = db.session.get(DictItem, did)
        if not d:
            return jsonify({'error': '不存在'}), 404
        data = request.get_json(force=True)
        for f in ['label', 'sort', 'remark']:
            if f in data:
                setattr(d, f, data[f])
        if 'enabled' in data:
            d.enabled = bool(data['enabled'])
        db.session.commit()
        log_operation('字典', '编辑', f'{d.category}/{d.label}', d.remark or '')
        return jsonify({'msg': '已更新'})

    @app.route('/api/dicts/<int:did>', methods=['DELETE'])
    @role_required('admin', 'hr')
    def api_dict_delete(did):
        d = db.session.get(DictItem, did)
        if not d:
            return jsonify({'error': '不存在'}), 404
        log_operation('字典', '删除', f'{d.category}/{d.label}', d.remark or '')
        db.session.delete(d)
        db.session.commit()
        return jsonify({'msg': '已删除'})

    # ---------- 公司（账套）管理：仅管理员，已并入「系统设置」页 ----------
    @app.route('/companies')
    @role_required('admin')
    def companies_page():
        return redirect(url_for('settings_page'))

    @app.route('/api/companies')
    @role_required('admin')
    def api_companies():
        companies = Company.query.order_by(Company.id.asc()).all()
        items = []
        for c in companies:
            items.append({
                'id': c.id, 'name': c.name, 'code': c.code,
                'enabled': c.enabled, 'remark': c.remark or '',
                'created_at': c.created_at.strftime('%Y-%m-%d') if c.created_at else '',
                'user_count': UserCompany.query.filter_by(company_id=c.id).count(),
            })
        return jsonify({'items': items})

    @app.route('/api/company-create', methods=['POST'])
    @role_required('admin')
    def api_company_create():
        data = request.get_json(force=True)
        name = (data.get('name') or '').strip()
        code = (data.get('code') or '').strip()
        if not name:
            return jsonify({'error': '公司名称不能为空'}), 400
        if not code:
            code = re.sub(r'[^A-Za-z0-9_-]', '', name) or ('c%d' % (Company.query.count() + 1))
        if Company.query.filter_by(code=code).first():
            return jsonify({'error': '该公司编码已存在'}), 400
        comp = Company(name=name, code=code, remark=(data.get('remark') or '').strip())
        db.session.add(comp)
        db.session.commit()
        # 建库建表
        company_engine(comp.id)
        return jsonify({'id': comp.id, 'msg': '已创建公司', 'name': name, 'code': code}), 201

    @app.route('/api/company/<int:cid>', methods=['PUT'])
    @role_required('admin')
    def api_company_update(cid):
        c = db.session.get(Company, cid)
        if not c:
            return jsonify({'error': '不存在'}), 404
        data = request.get_json(force=True)
        if 'name' in data and data['name'].strip():
            c.name = data['name'].strip()
        if 'enabled' in data:
            c.enabled = bool(data['enabled'])
        if 'remark' in data:
            c.remark = data['remark']
        db.session.commit()
        return jsonify({'msg': '已更新'})

    @app.route('/api/company-logo/<int:cid>', methods=['POST', 'DELETE'])
    @role_required('admin')
    def api_company_logo(cid):
        """上传 / 删除公司 Logo。文件存于 uploads/company_logos/，并在 Company.logo 记录文件名。"""
        c = db.session.get(Company, cid)
        if not c:
            return jsonify({'error': '不存在'}), 404
        import uuid
        logo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'company_logos')
        os.makedirs(logo_dir, exist_ok=True)
        if request.method == 'DELETE':
            if c.logo:
                try:
                    os.remove(os.path.join(logo_dir, c.logo))
                except Exception:
                    pass
            c.logo = ''
            db.session.commit()
            return jsonify({'msg': '已删除公司 Logo'})
        f = request.files.get('logo')
        if not f or not f.filename:
            return jsonify({'error': '未选择文件'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.ico'):
            return jsonify({'error': '不支持的图片格式'}), 400
        if c.logo:
            try:
                os.remove(os.path.join(logo_dir, c.logo))
            except Exception:
                pass
        fn = '%d_%s%s' % (cid, uuid.uuid4().hex[:8], ext)
        f.save(os.path.join(logo_dir, fn))
        c.logo = fn
        db.session.commit()
        return jsonify({'msg': '已上传公司 Logo', 'logo': fn})

    # ---------- 在职管理：转岗/调动 ----------
    @app.route('/api/employees/<int:eid>/transfer', methods=['POST'])
    @role_required('admin', 'hr')
    def api_employees_transfer(eid):
        e = db.session.get(Employee, eid)
        if not e:
            return jsonify({'error': '不存在'}), 404
        if e.status != '在职':
            return jsonify({'error': '仅在职员工可转岗'}), 400
        data = request.get_json(force=True)
        old_dept, old_pos = e.department, e.position
        if 'department' in data:
            e.department = data['department'] or e.department
        if 'position' in data:
            e.position = data['position'] or e.position
        # 记录转岗历史
        rec = TransferRecord(
            employee_id=e.id,
            old_department=old_dept,
            new_department=e.department,
            old_position=old_pos,
            new_position=e.position,
            remark=data.get('remark', ''),
            transfer_date=parse_date(data.get('transfer_date')) or date.today(),
            operator=current_user.name or current_user.username
        )
        db.session.add(rec)
        db.session.commit()
        log_operation('员工', '转岗', e.name, f'{old_dept}/{old_pos} → {e.department}/{e.position}')
        return jsonify({'msg': '转岗/调动成功'})

    @app.route('/api/employees/<int:eid>/transfers')
    @role_required('admin', 'hr', 'employee')
    def api_employee_transfers(eid):
        if not can_view_employee(eid):
            return _deny('无权限查看该员工信息')
        rows = TransferRecord.query.filter_by(employee_id=eid).order_by(
            TransferRecord.id.desc()).all()
        return jsonify({'items': [serialize_transfer(t) for t in rows]})

    # ---------- 全局：未读提醒数 ----------
    @app.context_processor
    def inject_globals():
        bell_count = 0
        my_companies = []
        cur_company = None
        if current_user.is_authenticated:
            bell_count = due_todo_query(current_user.id).count()
            my_companies = accessible_companies(current_user)
            cur_company = current_company()
        # 登录页品牌（系统名/Logo）取自全局 SysSetting，与左上角的「当前公司」展示解耦
        sys_title = SysSetting.get('sys_title', '人力资源管理系统')
        sys_logo = SysSetting.get('logo', '')
        # 左上角公司切换区使用的当前公司 Logo（无则路由自动返回内置默认图）
        company_logo_url = ('/company-logo/%d' % cur_company.id) if cur_company else ''
        # 登录后多公司选择弹窗：是否在页面上叠加显示
        show_company_modal = bool(getattr(g, 'show_company_modal', False)) and current_user.is_authenticated and not cur_company
        can_create_company = current_user.is_authenticated and current_user.role == 'admin'
        has_any_company = Company.query.filter_by(enabled=True).first() is not None
        return dict(current_user=current_user, bell_count=bell_count,
                    my_companies=my_companies, current_company=cur_company,
                    sys_title=sys_title, logo_url='/logo' if sys_logo else '',
                    has_logo=bool(sys_logo), company_logo_url=company_logo_url,
                    app_version=APP_VERSION,
                    can_view_module=can_view_module,
                    can_manage_module=can_manage_module,
                    show_company_modal=show_company_modal,
                    can_create_company=can_create_company,
                    has_any_company=has_any_company)

    # ---------- 404 ----------
    @app.errorhandler(404)
    def not_found(e):
        if request.path.startswith('/api/'):
            return jsonify({'error': '不存在'}), 404
        return render_template('base.html', error_404=True), 404

    @app.errorhandler(403)
    def forbidden(e):
        if request.path.startswith('/api/'):
            return jsonify({'error': '权限不足'}), 403
        return render_template('base.html', error_403=True), 403

    # ---------------- 权限中央拦截 ----------------
    # 模块前缀归类：命中后按写方法做权限校验
    MODULE_PREFIXES = {
        'employee': ('/lifecycle', '/contracts', '/api/contracts', '/api/contract_renews',
                     '/api/attachments', '/api/employees'),
        'insurance': ('/insurance', '/api/insurance'),
        'salary': ('/salary', '/tax', '/payslip-summary', '/api/salary', '/api/tax', '/api/payslip'),
        'todos': ('/todos', '/api/todos'),
        'logs': ('/logs', '/api/logs'),
        'users': ('/users', '/api/users'),
        'dicts': ('/dicts', '/api/dicts'),
        'backup': ('/backup', '/api/backup'),
    }

    @app.before_request
    def enforce_permissions():
        path = request.path
        # 白名单：静态资源、登录登出、公开 logo
        if path.startswith('/static') or path in ('/login', '/logout') or path == '/logo':
            return
        # 个人提醒角标：任何已登录用户可读（不依赖模块权限）
        if path == '/api/todos/bell':
            if not current_user.is_authenticated:
                return jsonify({'error': '未登录'}), 401
            return
        if not current_user.is_authenticated:
            if path.startswith('/api/'):
                return jsonify({'error': '未登录'}), 401
            return redirect(url_for('login'))
        module = None
        for mod, prefixes in MODULE_PREFIXES.items():
            if any(path.startswith(p) for p in prefixes):
                module = mod
                break
        if module:
            if request.method in ('POST', 'PUT', 'DELETE', 'PATCH'):
                if not can_manage_module(module):
                    return _deny(f'无权限：需要管理员账号或【{MODULES[module]}】模块的管理权限')
            else:
                if not can_view_module(module):
                    return _deny(f'无权限：需要【{MODULES[module]}】模块的查看权限')
            return
        # 其它路径（工作看板 / 系统设置 等）由各自 role_required 与登录态处理
        return

    # ---------------- 系统设置 ----------------
    @app.route('/settings')
    @role_required('admin')
    def settings_page():
        companies = Company.query.order_by(Company.id.asc()).all()
        stats = {c.id: UserCompany.query.filter_by(company_id=c.id).count() for c in companies}
        return render_template('settings.html', companies=companies, stats=stats)

    @app.route('/api/settings')
    @role_required('admin')
    def api_settings_get():
        return jsonify({
            'sys_title': SysSetting.get('sys_title', '人力资源管理系统'),
            'logo': SysSetting.get('logo', ''),
        })

    @app.route('/api/settings', methods=['POST'])
    @role_required('admin')
    def api_settings_update():
        sys_title = (request.form.get('sys_title') or '').strip()
        if sys_title:
            SysSetting.set('sys_title', sys_title)
        logo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'logo')
        # 恢复默认 logo
        if request.form.get('reset_logo'):
            old = SysSetting.get('logo', '')
            if old:
                try:
                    os.remove(os.path.join(logo_dir, old))
                except Exception:
                    pass
            SysSetting.set('logo', '')
        f = request.files.get('logo')
        if f and f.filename:
            ext = os.path.splitext(f.filename)[1].lower()
            if ext not in ('.png', '.jpg', '.jpeg', '.gif', '.svg', '.ico'):
                return jsonify({'error': '不支持的图片格式'}), 400
            os.makedirs(logo_dir, exist_ok=True)
            old = SysSetting.get('logo', '')
            if old:
                try:
                    os.remove(os.path.join(logo_dir, old))
                except Exception:
                    pass
            fn = 'logo' + ext
            f.save(os.path.join(logo_dir, fn))
            SysSetting.set('logo', fn)
        db.session.commit()
        return jsonify({'msg': '保存成功'})

    @app.route('/logo')
    def serve_logo():
        logo = SysSetting.get('logo', '')
        logo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'logo')
        if logo and os.path.exists(os.path.join(logo_dir, logo)):
            return send_file(os.path.join(logo_dir, logo))
        svg = ('<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="40" height="40">'
               '<circle cx="12" cy="12" r="11" fill="#2c5aa0"/>'
               '<text x="12" y="16" font-size="12" fill="#fff" text-anchor="middle" '
               'font-family="sans-serif">HR</text></svg>')
        return Response(svg, mimetype='image/svg+xml')

    # 内置默认公司 Logo（圆角色块 + 公司名首字），按公司 id 取模分配 4 种配色；公司未上传时用
    _COMPANY_LOGO_COLORS = ['#2c5aa0', '#10b981', '#f59e0b', '#8b5cf6']

    def default_company_logo_svg(cid, name):
        color = _COMPANY_LOGO_COLORS[(cid or 1) % len(_COMPANY_LOGO_COLORS)]
        ch = (name or '企')[0] if (name or '').strip() else '企'
        return ('<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 48 48" width="48" height="48">'
                '<rect width="48" height="48" rx="11" fill="%s"/>'
                '<text x="24" y="32" font-size="22" fill="#fff" text-anchor="middle" '
                'font-family="PingFang SC,Microsoft YaHei,sans-serif" font-weight="700">%s</text>'
                '</svg>' % (color, ch))

    @app.route('/company-logo/<int:cid>')
    def serve_company_logo(cid):
        c = db.session.get(Company, cid)
        logo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'company_logos')
        fn = c.logo if c and c.logo else ''
        if fn and os.path.exists(os.path.join(logo_dir, fn)):
            return send_file(os.path.join(logo_dir, fn))
        return Response(default_company_logo_svg(cid, c.name if c else ''), mimetype='image/svg+xml')

    return app


# ---------------- 工具函数 ----------------
def sync_contract_no(emp):
    """员工档案编号 ↔ 合同档案编号同步：
    仅在该员工已存在合同记录时，将档案编号写入合同；
    若无合同则不自动创建——合同起止时间须由用户手动录入后新增，避免默认生成合同。"""
    if not emp.archive_no:
        return
    contract = emp.contracts.filter_by(status='生效').first()
    if not contract:
        contract = emp.contracts.first()
    if contract:
        contract.contract_no = emp.archive_no
        db.session.flush()
    # 无合同：不自动创建，交由用户在合同模块手动录入起止时间后新增


def parse_date(s):
    # 空值 / NaN / NaT 一律视为无日期（避免 NaT 流入 Date 列触发 "cannot convert float NaN to integer"）
    if s is None or pd.isna(s):
        return None
    sv = str(s).strip()
    if not sv or sv.lower() == 'nan':
        return None
    if isinstance(s, date):
        return s
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d'):
        try:
            return datetime.strptime(sv, fmt).date()
        except ValueError:
            continue
    # pandas Timestamp / Excel 序列号等；解析失败或得到 NaT 均视为无日期
    try:
        d = pd.to_datetime(sv)
        return d.date() if not pd.isna(d) else None
    except Exception:
        return None


def parse_datetime(s):
    """解析 datetime-local 输入（YYYY-MM-DDTHH:MM 等）为 datetime，空返回 None"""
    if not s:
        return None
    if isinstance(s, datetime):
        return s
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M:%S',
                '%Y-%m-%d', '%Y/%m/%d %H:%M'):
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(str(s)).to_pydatetime()
    except Exception:
        return None


def get_client_ip():
    """获取操作者IP"""
    try:
        from flask import request as _req
        xff = _req.headers.get('X-Forwarded-For')
        if xff:
            return xff.split(',')[0].strip()
        return _req.remote_addr or ''
    except Exception:
        return ''


def log_operation(module, action, target='', detail=''):
    """记录操作日志（自动带上当前登录用户）"""
    from flask_login import current_user as _cu
    try:
        name = _cu.name if _cu.is_authenticated else ''
        rec = OperationLog(
            operator=_cu.username if _cu.is_authenticated else '匿名',
            operator_name=name,
            module=module,
            action=action,
            target=target,
            detail=detail,
            ip=get_client_ip(),
        )
        db.session.add(rec)
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass


def serialize_attachment(a):
    return {
        'id': a.id, 'employee_id': a.employee_id,
        'orig_name': a.orig_name,
        'size': a.size or 0,
        'category': a.category or '',
        'uploader': a.uploader or '',
        'created_at': a.created_at.strftime('%Y-%m-%d') if a.created_at else '',
    }


def serialize_contract(c):
    return {
        'id': c.id, 'employee_id': c.employee_id,
        'emp_name': c.employee.name,
        'department': c.employee.department or '',
        'contract_no': c.contract_no or '',
        'start_date': c.start_date.strftime('%Y-%m-%d') if c.start_date else '',
        'end_date': c.end_date.strftime('%Y-%m-%d') if c.end_date else '',
        'contract_type': c.contract_type,
        'status': c.status, 'remark': c.remark or '',
        'renew_count': c.renew_count or 0,
        'renews': [serialize_renew(r) for r in c.renews.all()],
        'days': c.days_to_expiry(),
    }


def serialize_renew(r):
    return {
        'id': r.id, 'contract_id': r.contract_id,
        'renew_date': r.renew_date.strftime('%Y-%m-%d') if r.renew_date else '',
        'old_end_date': r.old_end_date.strftime('%Y-%m-%d') if r.old_end_date else '',
        'new_end_date': r.new_end_date.strftime('%Y-%m-%d') if r.new_end_date else '',
        'remark': r.remark or '',
        'operator': r.operator or '',
    }


def serialize_employee(e):
    # 该员工的合同信息（优先生效合同）
    contract = e.contracts.filter_by(status='生效').first() or e.contracts.first()
    contract_info = None
    if contract:
        contract_info = {
            'contract_no': contract.contract_no or '',
            'contract_type': contract.contract_type,
            'start_date': contract.start_date.strftime('%Y-%m-%d') if contract.start_date else '',
            'end_date': contract.end_date.strftime('%Y-%m-%d') if contract.end_date else '',
            'status': contract.status,
            'renew_count': contract.renew_count or 0,
            'days': contract.days_to_expiry(),
        }
    return {
        'id': e.id, 'name': e.name, 'id_card': e.id_card,
        'phone': e.phone or '', 'department': e.department or '',
        'position': e.position or '', 'gender': e.gender or '',
        'hire_date': e.hire_date.strftime('%Y-%m-%d') if e.hire_date else '',
        'leave_date': e.leave_date.strftime('%Y-%m-%d') if e.leave_date else '',
        'status': e.status, 'address': e.address or '',
        'birthday': e.birthday.strftime('%Y-%m-%d') if e.birthday else '',
        'archive_no': e.archive_no or '',
        'native_place': e.native_place or '',
        'ethnicity': e.ethnicity or '',
        'education': e.education or '',
        'school': e.school or '',
        'political_status': e.political_status or '',
        'marital_status': e.marital_status or '',
        'email': e.email or '',
        'emergency_contact': e.emergency_contact or '',
        'emergency_phone': e.emergency_phone or '',
        'remark': e.remark or '',
        'photo': e.photo or '',
        # 参保信息
        'pension_enroll_date': e.pension_enroll_date.strftime('%Y-%m-%d') if e.pension_enroll_date else '',
        'pension_person_no': e.pension_person_no or '',
        'injury_enroll_date': e.injury_enroll_date.strftime('%Y-%m-%d') if e.injury_enroll_date else '',
        'injury_person_no': e.injury_person_no or '',
        'medical_enroll_date': e.medical_enroll_date.strftime('%Y-%m-%d') if e.medical_enroll_date else '',
        'medical_person_no': e.medical_person_no or '',
        'maternity_enroll_date': e.maternity_enroll_date.strftime('%Y-%m-%d') if e.maternity_enroll_date else '',
        'maternity_person_no': e.maternity_person_no or '',
        'unemployment_enroll_date': e.unemployment_enroll_date.strftime('%Y-%m-%d') if e.unemployment_enroll_date else '',
        'unemployment_person_no': e.unemployment_person_no or '',
        'fund_enroll_date': e.fund_enroll_date.strftime('%Y-%m-%d') if e.fund_enroll_date else '',
        'fund_person_no': e.fund_person_no or '',
        'contract_info': contract_info,
    }


def serialize_transfer(t):
    return {
        'id': t.id, 'employee_id': t.employee_id,
        'old_department': t.old_department or '',
        'new_department': t.new_department or '',
        'old_position': t.old_position or '',
        'new_position': t.new_position or '',
        'remark': t.remark or '',
        'transfer_date': t.transfer_date.strftime('%Y-%m-%d') if t.transfer_date else '',
        'operator': t.operator or '',
    }


def valid_id_card(s):
    """校验身份证号格式：15 位数字 或 18 位（末位可为 X/x）。"""
    return bool(re.match(r'^(?:\d{15}|\d{17}[\dXx])$', str(s or '').strip()))


def is_masked_id_card(s):
    """是否为脱敏身份证号（含 * 号，如 532224********1125）。"""
    return '*' in str(s or '').strip()


def match_employee_by_masked(name, masked, cache=None):
    """按「姓名 + 未脱敏部分身份证号」匹配员工（用于导入文件中身份证被脱敏的场景）。

    masked 形如 532224********1125；* 号位不参与比对，其余位必须与库中身份证逐位一致。
    返回 (emp, reason)：匹配成功 emp 为 Employee，reason 为空串；失败 emp 为 None。
    cache 可传入 dict 以复用同名同脱敏号的查询结果。
    """
    m = str(masked or '').strip().lower().replace(' ', '')
    nm = str(name or '').strip()
    if cache is not None:
        ck = (nm, m)
        if ck in cache:
            return cache[ck]
    res = (None, '')
    if not re.match(r'^[\dXx*]+$', m) or not (15 <= len(m) <= 18) or m.count('*') == 0:
        res = (None, '身份证号格式不正确（脱敏号应为15或18位，隐藏位用*）')
    else:
        cands = Employee.query.filter(Employee.name == nm).all()
        if not cands:
            res = (None, '未匹配到人员（员工表中无此姓名）')
        else:
            hits = []
            for e in cands:
                v = str(e.id_card or '').strip().lower()
                if len(v) != len(m):
                    continue
                if all(m[i] == '*' or m[i] == v[i] for i in range(len(m))):
                    hits.append(e)
            if len(hits) == 1:
                res = (hits[0], '')
            elif len(hits) > 1:
                res = (None, f'按姓名+部分身份证匹配到 {len(hits)} 人，请补全身份证号')
            else:
                res = (None, '未匹配到人员（姓名或部分身份证号与员工表不符）')
    if cache is not None:
        cache[(nm, m)] = res
    return res


def serialize_insurance(d):
    return {
        'id': d.id, 'employee_id': d.employee_id,
        'period': d.period, 'base': d.base or 0,
        'person_no': d.person_no or '',
        'remark': d.remark or '',
        'pension_emp': d.pension_emp or 0, 'pension_per': d.pension_per or 0,
        'medical_emp': d.medical_emp or 0, 'medical_per': d.medical_per or 0,
        'extra_medical_emp': d.extra_medical_emp or 0, 'extra_medical_per': d.extra_medical_per or 0,
        'unemployment_emp': d.unemployment_emp or 0, 'unemployment_per': d.unemployment_per or 0,
        'injury_emp': d.injury_emp or 0, 'maternity_emp': d.maternity_emp or 0,
        'fund_base': d.fund_base or 0, 'fund_rate': d.fund_rate or 0,
        'fund_emp': d.fund_emp or 0, 'fund_per': d.fund_per or 0,
    }


def serialize_todo(t):
    return {
        'id': t.id, 'title': t.title,
        'due_date': t.due_date.strftime('%Y-%m-%d'),
        'priority': t.priority, 'done': t.done,
        'remind_start': t.remind_start.strftime('%Y-%m-%d %H:%M') if t.remind_start else '',
        'remind_end': t.remind_end.strftime('%Y-%m-%d %H:%M') if t.remind_end else '',
        'remind_count': t.remind_count or 1,
        'reminded_count': t.reminded_count or 0,
    }


def serialize_user(u):
    return {
        'id': u.id, 'username': u.username, 'name': u.name,
        'role': u.role, 'employee_id': u.employee_id,
        'can_employee_view': bool(u.can_employee_view),
        'can_employee_manage': bool(u.can_employee_manage),
        'can_insurance_view': bool(u.can_insurance_view),
        'can_insurance_manage': bool(u.can_insurance_manage),
        'can_todos_view': bool(u.can_todos_view),
        'can_todos_manage': bool(u.can_todos_manage),
        'can_logs_view': bool(u.can_logs_view),
        'can_logs_manage': bool(u.can_logs_manage),
        'can_users_view': bool(u.can_users_view),
        'can_users_manage': bool(u.can_users_manage),
        'can_dicts_view': bool(u.can_dicts_view),
        'can_dicts_manage': bool(u.can_dicts_manage),
        'can_backup_view': bool(u.can_backup_view),
        'can_backup_manage': bool(u.can_backup_manage),
        'emp_view_mode': u.emp_view_mode or 'self',
        'emp_depts': parse_depts(u.emp_depts),
        'companies': [uc.company_id for uc in UserCompany.query.filter_by(user_id=u.id).all()],
        'created_at': u.created_at.strftime('%Y-%m-%d') if u.created_at else '',
    }


def serialize_dict(d):
    return {
        'id': d.id, 'category': d.category, 'label': d.label,
        'sort': d.sort, 'enabled': d.enabled, 'remark': d.remark or '',
    }


app = create_app()


def run_migrations():
    """为已存在的 SQLite 表补充新增列（幂等）"""
    from sqlalchemy import text, inspect
    with app.app_context():
        db.create_all()
        insp = inspect(db.engine)
        # companies 表新增 logo 列（公司 Logo 文件名）；旧表需补列
        try:
            comp_cols = {c['name'] for c in insp.get_columns('companies')}
            if 'logo' not in comp_cols:
                db.session.execute(text("ALTER TABLE companies ADD COLUMN logo VARCHAR(255) DEFAULT ''"))
                db.session.commit()
        except Exception:
            pass
        emp_cols = {c['name'] for c in insp.get_columns('employees')}
        add = {
            'archive_no': 'VARCHAR(64)',
            'native_place': 'VARCHAR(64)',
            'ethnicity': 'VARCHAR(16)',
            'education': 'VARCHAR(32)',
            'school': 'VARCHAR(128)',
            'political_status': 'VARCHAR(16)',
            'marital_status': 'VARCHAR(16)',
            'email': 'VARCHAR(128)',
            'emergency_contact': 'VARCHAR(64)',
            'emergency_phone': 'VARCHAR(32)',
            'remark': 'VARCHAR(512)',
            'photo': 'VARCHAR(256)',
            'pension_enroll_date': 'DATE',
            'pension_person_no': 'VARCHAR(32)',
            'injury_enroll_date': 'DATE',
            'injury_person_no': 'VARCHAR(32)',
            'medical_enroll_date': 'DATE',
            'medical_person_no': 'VARCHAR(32)',
            'maternity_enroll_date': 'DATE',
            'maternity_person_no': 'VARCHAR(32)',
            'unemployment_enroll_date': 'DATE',
            'unemployment_person_no': 'VARCHAR(32)',
            'fund_enroll_date': 'DATE',
            'fund_person_no': 'VARCHAR(32)',
        }
        for col, typ in add.items():
            if col not in emp_cols:
                db.session.execute(text(
                    f'ALTER TABLE employees ADD COLUMN {col} {typ}'))
        # 字典分类历史数据迁移：部门 -> 客户单位
        try:
            db.session.execute(text("UPDATE dict_items SET category='客户单位' WHERE category='部门'"))
        except Exception:
            pass
        # contracts 表新增列
        try:
            ctr_cols = {c['name'] for c in insp.get_columns('contracts')}
            if 'renew_count' not in ctr_cols:
                db.session.execute(text('ALTER TABLE contracts ADD COLUMN renew_count INTEGER DEFAULT 0'))
        except Exception:
            pass
        # todos 表新增列（提醒时间/次数）
        try:
            todo_cols = {c['name'] for c in insp.get_columns('todos')}
            if 'remind_time' not in todo_cols:
                db.session.execute(text('ALTER TABLE todos ADD COLUMN remind_time VARCHAR(8)'))
            if 'remind_start' not in todo_cols:
                db.session.execute(text('ALTER TABLE todos ADD COLUMN remind_start DATETIME'))
            if 'remind_end' not in todo_cols:
                db.session.execute(text('ALTER TABLE todos ADD COLUMN remind_end DATETIME'))
            if 'remind_count' not in todo_cols:
                db.session.execute(text('ALTER TABLE todos ADD COLUMN remind_count INTEGER DEFAULT 1'))
            if 'reminded_count' not in todo_cols:
                db.session.execute(text('ALTER TABLE todos ADD COLUMN reminded_count INTEGER DEFAULT 0'))
            if 'company_id' not in todo_cols:
                db.session.execute(text('ALTER TABLE todos ADD COLUMN company_id INTEGER'))
        except Exception:
            pass
        # 历史待办归属默认公司（按公司分开管理/提醒），仅回填一次
        try:
            if Company.query.count() > 0 and 'todos' in {t for t in inspect(db.engine).get_table_names()}:
                _def = Company.query.filter_by(code='default').first() or Company.query.first()
                db.session.execute(text(
                    'UPDATE todos SET company_id=:cid WHERE company_id IS NULL'), {'cid': _def.id})
                db.session.commit()
        except Exception:
            pass
        # insurance_details 表新增列
        try:
            ins_cols = {c['name'] for c in insp.get_columns('insurance_details')}
            ins_add = {
                'person_no': 'VARCHAR(32)',
                'remark': 'VARCHAR(512)',
                'extra_medical_emp': 'FLOAT DEFAULT 0',
                'extra_medical_per': 'FLOAT DEFAULT 0',
                'fund_base': 'FLOAT DEFAULT 0',
                'fund_rate': 'FLOAT DEFAULT 0',
            }
            for col, typ in ins_add.items():
                if col not in ins_cols:
                    db.session.execute(text(f'ALTER TABLE insurance_details ADD COLUMN {col} {typ}'))
        except Exception:
            pass
        # payslip_templates 表新增 headfoot_json（表头/表脚模板）
        try:
            tpl_cols = {c['name'] for c in insp.get_columns('payslip_templates')}
            if 'headfoot_json' not in tpl_cols:
                db.session.execute(text('ALTER TABLE payslip_templates ADD COLUMN headfoot_json TEXT'))
        except Exception:
            pass
        # users 表新增 HR 模块权限（查看/管理）与员工查看范围列
        try:
            user_cols = {c['name'] for c in insp.get_columns('users')}
            new_user_cols = {
                'can_employee_view': 'BOOLEAN DEFAULT 1',
                'can_employee_manage': 'BOOLEAN DEFAULT 1',
                'can_insurance_view': 'BOOLEAN DEFAULT 1',
                'can_insurance_manage': 'BOOLEAN DEFAULT 1',
                'can_todos_view': 'BOOLEAN DEFAULT 1',
                'can_todos_manage': 'BOOLEAN DEFAULT 1',
                'can_logs_view': 'BOOLEAN DEFAULT 1',
                'can_logs_manage': 'BOOLEAN DEFAULT 1',
                'can_salary_view': 'BOOLEAN DEFAULT 1',
                'can_salary_manage': 'BOOLEAN DEFAULT 1',
                'can_users_view': 'BOOLEAN DEFAULT 1',
                'can_users_manage': 'BOOLEAN DEFAULT 1',
                'can_dicts_view': 'BOOLEAN DEFAULT 1',
                'can_dicts_manage': 'BOOLEAN DEFAULT 1',
                'can_backup_view': 'BOOLEAN DEFAULT 1',
                'can_backup_manage': 'BOOLEAN DEFAULT 1',
                'emp_view_mode': "VARCHAR(8) DEFAULT 'self'",
                'emp_depts': "TEXT DEFAULT ''",
            }
            for col, typ in new_user_cols.items():
                if col not in user_cols:
                    db.session.execute(text(f'ALTER TABLE users ADD COLUMN {col} {typ}'))
            # 默认全部置 1（admin/hr 默认拥有；employee 忽略）；员工查看范围默认 self
            db.session.execute(text(
                "UPDATE users SET "
                "can_employee_view=1,can_employee_manage=1,can_insurance_view=1,can_insurance_manage=1,"
                "can_todos_view=1,can_todos_manage=1,can_logs_view=1,can_logs_manage=1,"
                "can_users_view=1,can_users_manage=1,can_dicts_view=1,can_dicts_manage=1,"
                "can_salary_view=1,can_salary_manage=1,"
                "can_backup_view=1,can_backup_manage=1,"
                "emp_view_mode=COALESCE(NULLIF(emp_view_mode,''),'self')"
            ))
            # 从旧权限列回填 HR 的「管理」权限（查看恒为 1）
            if 'perm_employee' in user_cols:
                db.session.execute(text(
                    "UPDATE users SET can_employee_manage=(CASE WHEN perm_employee THEN 1 ELSE 0 END) WHERE role='hr'"))
            if 'perm_insurance' in user_cols:
                db.session.execute(text(
                    "UPDATE users SET can_insurance_manage=(CASE WHEN perm_insurance THEN 1 ELSE 0 END) WHERE role='hr'"))
            db.session.commit()
        except Exception:
            pass
        # salary_records 表新增 client_unit（客户单位）与原始表格内容列
        try:
            sal_cols = {c['name'] for c in insp.get_columns('salary_records')}
            if 'client_unit' not in sal_cols:
                db.session.execute(text("ALTER TABLE salary_records ADD COLUMN client_unit VARCHAR(128) DEFAULT ''"))
            if 'headers_json' not in sal_cols:
                db.session.execute(text("ALTER TABLE salary_records ADD COLUMN headers_json TEXT DEFAULT ''"))
            if 'values_json' not in sal_cols:
                db.session.execute(text("ALTER TABLE salary_records ADD COLUMN values_json TEXT DEFAULT ''"))
            db.session.commit()
        except Exception:
            pass
        db.session.commit()
        # 默认字典改为「每公司独立库」初始化，见 migrate_dicts_to_companies()
        # 系统级品牌（系统名 / 登录页 Logo）迁移到全局 sys_settings 表，与左侧「当前公司」展示解耦
        try:
            if not SysSetting.query.filter_by(key='sys_title').first():
                title = '人力资源管理系统'
                default_c = Company.query.filter_by(code='default').first()
                if default_c and os.path.exists(default_c.db_path()):
                    try:
                        from sqlalchemy import create_engine as _ce
                        _eng = _ce('sqlite:///' + default_c.db_path())
                        _row = _eng.execute(text("SELECT value FROM settings WHERE key='sys_title'")).fetchone()
                        if _row and _row[0]:
                            title = _row[0]
                        _eng.dispose()
                    except Exception:
                        pass
                SysSetting.set('sys_title', title)
            db.session.commit()
        except Exception:
            pass
        # 操作记录改为「每公司独立库」：为已有公司补建 operation_logs 表，
        # 并将全局库（data.db）中历史操作记录迁入默认公司库，时间由 UTC 校正为本地时间。
        try:
            for _c in Company.query.all():
                try:
                    company_engine(_c.id)   # 建库建表（含 operation_logs）
                except Exception:
                    pass
            # 仅当默认公司存在且全局仍有旧 operation_logs 时迁移一次
            if Company.query.count() > 0 and 'operation_logs' in {t for t in inspect(db.engine).get_table_names()}:
                _comp = Company.query.filter_by(code='default').first() or Company.query.first()
                _eng = company_engine(_comp.id)
                with _eng.begin() as _conn:
                    _existing = _conn.execute(text('SELECT COUNT(*) FROM operation_logs')).scalar() or 0
                    if _existing == 0:
                        _rows = db.session.execute(text(
                            'SELECT operator,operator_name,module,action,target,detail,ip,created_at '
                            'FROM operation_logs ORDER BY id ASC')).fetchall()
                        for _r in _rows:
                            _ca = _r[7]
                            # 历史 created_at 为 UTC（datetime.utcnow）；SQLite 原始查询返回字符串，需解析后校正为本地时间（+8h）
                            if isinstance(_ca, str):
                                try:
                                    _ca = datetime.strptime(_ca, '%Y-%m-%d %H:%M:%S.%f')
                                except Exception:
                                    try:
                                        _ca = datetime.strptime(_ca, '%Y-%m-%d %H:%M:%S')
                                    except Exception:
                                        _ca = None
                            if isinstance(_ca, datetime):
                                _ca = (_ca + timedelta(hours=8)).replace(tzinfo=None)
                            _conn.execute(text(
                                'INSERT INTO operation_logs '
                                '(operator,operator_name,module,action,target,detail,ip,created_at) '
                                'VALUES (:o,:on,:m,:a,:t,:d,:ip,:ca)'),
                                {'o': _r[0], 'on': _r[1], 'm': _r[2], 'a': _r[3],
                                 't': _r[4], 'd': _r[5], 'ip': _r[6], 'ca': _ca})
                # 清理全局库中的旧操作记录表（数据已迁入公司库，避免重复/混淆）
                try:
                    db.session.execute(text('DROP TABLE IF EXISTS operation_logs'))
                    db.session.commit()
                except Exception:
                    pass
        except Exception:
            pass
        # 旧单体数据迁移到首个公司（仅首次、且存在旧业务表时）
        migrate_to_companies()
        # 代码字段（字典）改为「每公司独立库」：迁移全局历史字典到各公司库，并补充默认字典
        migrate_dicts_to_companies()


def migrate_to_companies():
    """首次启动：若全局库仍存在旧业务表（员工/工资/合同等），则创建默认公司，
    将其整体迁入该公司独立库，并把现有 hr/employee 用户关联到该公司；随后清理全局旧业务表。
    幂等：已存在公司则跳过。"""
    from sqlalchemy import text, inspect as sa_inspect
    with app.app_context():
        if Company.query.count() > 0:
            return
        insp = sa_inspect(db.engine)
        tables = set(insp.get_table_names())
        old_biz = ['employees', 'contracts', 'contract_renews', 'attachments',
                   'insurance_details', 'salary_records', 'transfer_records',
                   'payslip_templates', 'settings']
        if not (set(old_biz) & tables):
            # 全新安装：自动创建默认公司，保证开箱即用
            comp = Company(name='默认公司', code='default')
            db.session.add(comp)
            db.session.commit()
            company_engine(comp.id)
            return

        # 默认公司名/编码取自原系统设置
        name = '默认公司'
        code = 'default'
        try:
            row = db.session.execute(text("SELECT value FROM settings WHERE key='sys_title'")).fetchone()
            if row and row[0]:
                name = row[0]
            row = db.session.execute(text("SELECT value FROM settings WHERE key='sys_code'")).fetchone()
            if row and row[0]:
                code = re.sub(r'[^A-Za-z0-9_-]', '', str(row[0])) or 'default'
        except Exception:
            pass

        comp = Company(name=name, code=code)
        db.session.add(comp)
        db.session.commit()
        eng = company_engine(comp.id)  # 建库+建表

        counts = {}
        for tbl in old_biz:
            if tbl not in tables:
                continue
            cols = [c['name'] for c in insp.get_columns(tbl)]
            rows = db.session.execute(text('SELECT ' + ','.join(cols) + f' FROM {tbl}')).fetchall()
            counts[tbl] = len(rows)
            if not rows:
                continue
            with eng.begin() as conn:
                for r in rows:
                    params = {f'p{i}': v for i, v in enumerate(r)}
                    ph = ','.join(f':p{i}' for i in range(len(cols)))
                    conn.execute(text(f'INSERT INTO {tbl} ({",".join(cols)}) VALUES ({ph})'), params)

        # 关联现有 hr/employee 用户到默认公司（admin 靠角色自动拥有全部）
        for u in User.query.all():
            if u.role in ('hr', 'employee'):
                if not UserCompany.query.filter_by(user_id=u.id, company_id=comp.id).first():
                    db.session.add(UserCompany(user_id=u.id, company_id=comp.id))
        db.session.commit()

        # 清理全局旧业务表（数据已迁入公司库）
        for tbl in old_biz:
            if tbl in tables:
                try:
                    db.session.execute(text(f'DROP TABLE IF EXISTS {tbl}'))
                except Exception:
                    pass
        db.session.commit()


def migrate_dicts_to_companies():
    """代码字段（字典）改为「每公司独立库」：
    - 为已存在的公司建表（company_engine 已含 dict_items）；
    - 将全局库（data.db）中历史 dict_items 迁入各公司库（仅当本公司尚无数据，幂等）；
    - 补充默认代码字段；
    - 随后清理全局旧字典表，避免与「每公司独立库」混淆。"""
    from sqlalchemy import text, inspect as sa_inspect
    with app.app_context():
        global_has = 'dict_items' in {t for t in sa_inspect(db.engine).get_table_names()}
        g_rows = []
        if global_has:
            try:
                # 全局字典表仍在主库（data.db）；读取历史数据用于迁入各公司库
                g_rows = db.session.execute(
                    text('SELECT category,label,sort,enabled,remark FROM dict_items')).fetchall()
            except Exception:
                g_rows = []
        for _c in Company.query.all():
            eng = company_engine(_c.id)  # 建库建表（含 dict_items）
            token = current_company_engine.set(eng)
            try:
                # 仅当本公司尚无字典数据时迁入全局历史数据
                if DictItem.query.count() == 0 and g_rows:
                    for r in g_rows:
                        if not DictItem.query.filter_by(category=r[0], label=r[1]).first():
                            db.session.add(DictItem(
                                category=r[0], label=r[1],
                                sort=int(r[2] or 0), enabled=bool(r[3]), remark=r[4] or ''))
                seed_default_dicts()  # 确保默认代码字段存在（写入当前公司引擎）
                db.session.commit()
            finally:
                current_company_engine.reset(token)
        # 清理全局旧字典表
        if global_has:
            try:
                db.session.execute(text('DROP TABLE IF EXISTS dict_items'))
                db.session.commit()
            except Exception:
                pass


def seed_default_dicts():
    """填充默认代码字段字典（已存在则跳过）"""
    defaults = [
        ('合同类型', '固定期限', 1), ('合同类型', '无固定期限', 2), ('合同类型', '以完成一定工作任务为期限', 3),
        ('婚姻状况', '未婚', 1), ('婚姻状况', '已婚', 2), ('婚姻状况', '离异', 3), ('婚姻状况', '丧偶', 4),
        ('学历', '小学', 0), ('学历', '初中', 1), ('学历', '中专', 2), ('学历', '职业高中', 2), ('学历', '普通高中', 2), ('学历', '大学专科', 3), ('学历', '大学本科', 4), ('学历', '硕士', 5), ('学历', '博士', 6),
        ('客户单位', '空港航食', 1), ('客户单位', '宁洱荞樽', 2), ('客户单位', '铁路局（乐道停车场）', 3), ('客户单位', '彤骏', 4), ('客户单位', '曦堃垚', 5),
        ('政治面貌', '中国共产党党员', 1), ('政治面貌', '中国共产主义青年团团员', 2), ('政治面貌', '群众', 3), ('政治面貌', '民主党派', 4),
        ('民族', '汉族', 1), ('民族', '壮族', 2), ('民族', '回族', 3), ('民族', '满族', 4), ('民族', '维吾尔族', 5), ('民族', '苗族', 6), ('民族', '彝族', 7), ('民族', '土家族', 8), ('民族', '白族', 9), ('民族', '傣族', 10), ('民族', '哈尼族', 11), ('民族', '基诺族', 12), ('民族', '傈僳族', 13), ('民族', '纳西族', 14), ('民族', '佤族', 15), ('民族', '其他', 99),
        ('职位', '普工', 0), ('职位', '工勤', 0), ('职位', '总经理', 1), ('职位', '部门经理', 2), ('职位', '客户单位经理', 2), ('职位', '工程师', 3), ('职位', '架构师', 4), ('职位', 'HR专员', 5), ('职位', '会计', 6), ('职位', '销售代表', 7), ('职位', '行政助理', 8),
    ]
    for cat, label, sort in defaults:
        if not DictItem.query.filter_by(category=cat, label=label).first():
            db.session.add(DictItem(category=cat, label=label, sort=sort))
    db.session.commit()


# 模块加载即执行迁移（无论是 `python3 app.py` 直接运行，还是被 import 以启动服务）
run_migrations()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
